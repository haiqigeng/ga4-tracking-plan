from __future__ import annotations

import io
import json
import sys
import tempfile
import unittest
from contextlib import redirect_stdout
from pathlib import Path
from unittest.mock import patch
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as WorkbookImage
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table
from PIL import Image as PillowImage

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "skill" / "scripts"))

from adapt_tracking_plan_workbook import (  # noqa: E402
    _cell_value,
    _value_at_path,
    adapt_workbook,
    build_fidelity_report,
    load_mapping,
    validate_mapping,
)
from adapt_tracking_plan_workbook import main as adapt_main  # noqa: E402
from inspect_tracking_plan_template import (  # noqa: E402
    compare_workbook_fidelity,
    file_sha256,
    inspect_workbook,
    unsupported_strict_package_parts,
)
from inspect_tracking_plan_template import main as inspect_main  # noqa: E402


class TemplateToolTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.plan = json.loads(
            (ROOT / "skill" / "references" / "03-rules" / "example-ga4-tracking-plan.json").read_text(encoding="utf-8")
        )

    @staticmethod
    def strict_mapping(template: Path, writes: list[dict]) -> dict:
        return {
            "mode": "strict_client_template",
            "mapping_id": "test_mapping",
            "template_sha256": file_sha256(template),
            "cell_writes": writes,
        }

    def test_inspector_reports_structure_hash_and_hazards_without_cell_dump(self) -> None:
        with tempfile.TemporaryDirectory() as raw:
            path = Path(raw) / "client.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Client Overview"
            sheet.append(["Name", "Value"])
            sheet.append(["Total", "=1+1"])
            sheet["A2"].comment = Comment("Keep this analyst note", "Analyst")
            sheet.protection.sheet = True
            sheet.add_table(Table(displayName="ClientTable", ref="A1:B2"))
            validation = DataValidation(type="list", formula1='"ok,ko"')
            sheet.add_data_validation(validation)
            validation.add("B2")
            workbook.save(path)

            inventory = inspect_workbook(path)
            expected_hash = file_sha256(path)

        self.assertEqual(inventory["template_sha256"], expected_hash)
        inspected = inventory["sheets"][0]
        self.assertEqual(inspected["formula_count"], 1)
        self.assertTrue(inspected["sheet_protected"])
        self.assertEqual(inspected["table_count"], 1)
        self.assertEqual(inspected["data_validation_count"], 1)
        self.assertIn("formula cells", inspected["destructive_overwrite_hazards"])
        self.assertNotIn("cells", inspected)

    def test_mapping_requires_structured_mode_hash_and_unique_cells(self) -> None:
        with self.assertRaisesRegex(ValueError, "explicit mapping"):
            load_mapping(None)
        with tempfile.TemporaryDirectory() as raw:
            path = Path(raw) / "mapping.json"
            invalid_values = [
                ["not", "a", "mapping"],
                {"00 Overview": "Client Overview"},
                {"mode": "strict_client_template", "template_sha256": "x", "cell_writes": []},
                {
                    "mode": "strict_client_template",
                    "template_sha256": "0" * 64,
                    "cell_writes": [
                        {"sheet": "Events", "cell": "A1", "value": "x"},
                        {"sheet": "Events", "cell": "A1", "value": "y"},
                    ],
                },
            ]
            for value in invalid_values:
                path.write_text(json.dumps(value), encoding="utf-8")
                with self.subTest(value=value), self.assertRaises(ValueError):
                    load_mapping(path)

    def test_mapping_contract_rejects_every_ambiguous_structure(self) -> None:
        write = {"sheet": "Events", "cell": "A1", "value": "x"}
        clone = {
            "source_sheet": "Events",
            "target_sheet": "Events 2",
            "approved_reason": "Client approved this additional event block",
        }
        base = {
            "mode": "strict_client_template",
            "template_sha256": "0" * 64,
            "cell_writes": [write],
        }
        invalid_mappings = [
            {**base, "mode": "legacy"},
            {**base, "cell_writes": []},
            {**base, "cell_writes": ["not an object"]},
            {**base, "cell_writes": [{**write, "unknown": True}]},
            {**base, "cell_writes": [{**write, "cell": "a1"}]},
            {**base, "cell_writes": [{**write, "value_path": "$.document.title"}]},
            {**base, "sheet_clones": [clone]},
            {**base, "mode": "approved_structural_extension"},
            {**base, "mode": "approved_structural_extension", "sheet_clones": ["not an object"]},
            {
                **base,
                "mode": "approved_structural_extension",
                "sheet_clones": [{"source_sheet": "Events"}],
            },
            {
                **base,
                "mode": "approved_structural_extension",
                "sheet_clones": [{**clone, "target_sheet": "Events"}],
            },
            {
                **base,
                "mode": "approved_structural_extension",
                "sheet_clones": [clone, clone],
            },
        ]
        for mapping in invalid_mappings:
            with self.subTest(mapping=mapping), self.assertRaises(ValueError):
                validate_mapping(mapping)

    def test_mapping_value_paths_and_transforms_fail_closed(self) -> None:
        value = {"events": [{"journey_ids": ["one", "two"]}]}
        self.assertIs(_value_at_path(value, "$"), value)
        self.assertEqual(_value_at_path(value, "$.events[0].journey_ids"), ["one", "two"])
        for path in ("events", "$.events.bad-key", "$.missing", "$.events[2]"):
            with self.subTest(path=path), self.assertRaises(ValueError):
                _value_at_path(value, path)

        self.assertEqual(_cell_value(["one", "two"], "join_pipe"), "one | two")
        self.assertEqual(_cell_value({"status": "ok"}, "json"), '{"status":"ok"}')
        with self.assertRaisesRegex(ValueError, "requires a list"):
            _cell_value("one", "join_pipe")
        with self.assertRaisesRegex(ValueError, "Unknown mapping transform"):
            _cell_value("one", "unsupported")

    def test_strict_mapping_paths_transforms_and_formula_guardrail(self) -> None:
        with tempfile.TemporaryDirectory() as raw:
            template = Path(raw) / "client.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Events"
            sheet.merge_cells("F1:G1")
            sheet["F1"] = "Merged"
            sheet["H1"] = "=1+1"
            workbook.save(template)

            mapping = self.strict_mapping(template, [
                {"sheet": "Events", "cell": "A1", "value_path": "$.events[0].journey_ids", "transform": "join_pipe"},
                {"sheet": "Events", "cell": "B1", "value": {"state": "ok"}},
                {"sheet": "Events", "cell": "C1", "value": ["one", "two"]},
                {"sheet": "Events", "cell": "D1", "value": 7, "transform": "string"},
                {"sheet": "Events", "cell": "H1", "value": "approved", "allow_formula_overwrite": True},
            ])
            adapted = adapt_workbook(self.plan, template, mapping)
            self.assertEqual(adapted["Events"]["A1"].value, "homepage_discovery")
            self.assertEqual(adapted["Events"]["B1"].value, '{"state":"ok"}')
            self.assertEqual(adapted["Events"]["C1"].value, "one | two")
            self.assertEqual(adapted["Events"]["D1"].value, "7")
            self.assertEqual(adapted["Events"]["H1"].value, "approved")

            guarded = self.strict_mapping(template, [{"sheet": "Events", "cell": "H1", "value": "no"}])
            with self.assertRaisesRegex(ValueError, "replace formula"):
                adapt_workbook(self.plan, template, guarded)

    def test_template_hash_and_missing_sheet_fail_closed(self) -> None:
        with tempfile.TemporaryDirectory() as raw:
            template = Path(raw) / "client.xlsx"
            workbook = Workbook()
            workbook.active.title = "Events"
            workbook.save(template)
            mismatched = self.strict_mapping(template, [{"sheet": "Events", "cell": "A1", "value": "x"}])
            mismatched["template_sha256"] = "0" * 64
            with self.assertRaisesRegex(ValueError, "different template"):
                adapt_workbook(self.plan, template, mismatched)
            missing = self.strict_mapping(template, [{"sheet": "Missing", "cell": "A1", "value": "x"}])
            with self.assertRaisesRegex(ValueError, "missing sheet"):
                adapt_workbook(self.plan, template, missing)

    def test_strict_adapter_blocks_unsupported_chart_parts(self) -> None:
        with tempfile.TemporaryDirectory() as raw:
            template = Path(raw) / "client.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Events"
            sheet.append(["Value"])
            sheet.append([1])
            chart = BarChart()
            chart.add_data(Reference(sheet, min_col=1, min_row=1, max_row=2), titles_from_data=True)
            sheet.add_chart(chart, "C1")
            workbook.save(template)
            mapping = self.strict_mapping(template, [{"sheet": "Events", "cell": "A2", "value": 2}])
            with self.assertRaisesRegex(ValueError, "cannot preserve"):
                adapt_workbook(self.plan, template, mapping)

        with tempfile.TemporaryDirectory() as raw:
            template = Path(raw) / "client-with-shape.xlsx"
            workbook = Workbook()
            workbook.active["A1"] = "Value"
            workbook.save(template)
            with ZipFile(template, "a") as archive:
                archive.writestr(
                    "xl/drawings/drawing1.xml",
                    '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"><xdr:sp/></xdr:wsDr>',
                )
            mapping = self.strict_mapping(template, [{"sheet": "Sheet", "cell": "A1", "value": "Changed"}])
            with self.assertRaisesRegex(ValueError, "cannot preserve"):
                adapt_workbook(self.plan, template, mapping)

        with tempfile.TemporaryDirectory() as raw:
            corrupt = Path(raw) / "corrupt.xlsx"
            corrupt.write_bytes(b"not an OOXML package")
            self.assertTrue(unsupported_strict_package_parts(corrupt)[0].startswith("unreadable-workbook-package:"))

    def test_strict_adapter_preserves_complete_surface_and_image_anchor(self) -> None:
        with tempfile.TemporaryDirectory() as raw:
            folder = Path(raw)
            template = folder / "client.xlsx"
            output = folder / "adapted.xlsx"
            image_path = folder / "evidence.png"
            with PillowImage.new("RGB", (24, 16), "red") as source_image:
                source_image.save(image_path)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Events"
            sheet.append(["Event", "Version", "Formula"])
            sheet.append(["page_view", "old", "=1+1"])
            sheet["A2"].comment = Comment("Keep this analyst note", "Analyst")
            sheet["B2"].fill = PatternFill("solid", fgColor="00FF00")
            sheet.merge_cells("D1:E1")
            sheet["D1"] = "Client heading"
            sheet.add_table(Table(displayName="EventTable", ref="A1:C2"))
            validation = DataValidation(type="list", formula1='"draft,published"')
            sheet.add_data_validation(validation)
            validation.add("B2")
            with PillowImage.open(image_path) as source_image:
                sheet.add_image(WorkbookImage(source_image), "F1")
                workbook.save(template)

            mapping = self.strict_mapping(template, [{"sheet": "Events", "cell": "B2", "value_path": "$.document.version"}])
            adapted = adapt_workbook(self.plan, template, mapping)
            adapted.save(output)
            report = compare_workbook_fidelity(template, output, {"Events": {"B2"}})

            self.assertEqual(report["status"], "passed")
            self.assertEqual(adapted["Events"]["C2"].value, "=1+1")
            self.assertEqual(adapted["Events"]["B2"].fill.fgColor.rgb, "0000FF00")
            self.assertEqual(len(adapted["Events"]._images), 1)

    def test_fidelity_detects_image_extent_and_workbook_metadata_changes(self) -> None:
        with tempfile.TemporaryDirectory() as raw:
            folder = Path(raw)
            template = folder / "client.xlsx"
            changed = folder / "changed.xlsx"
            image_path = folder / "image.png"
            with PillowImage.new("RGB", (20, 20), "blue") as source_image:
                source_image.save(image_path)
            workbook = Workbook()
            workbook.properties.title = "Client template"
            with PillowImage.open(image_path) as source_image:
                workbook.active.add_image(WorkbookImage(source_image), "C3")
                workbook.save(template)
            edited = load_workbook(template)
            edited.properties.title = "Changed"
            edited.active._images[0].width = 80
            edited.save(changed)
            report = compare_workbook_fidelity(template, changed, {})
            self.assertEqual(report["status"], "failed")
            self.assertTrue(any("workbook_properties" in item or "images" in item for item in report["unexpected_differences"]))

    def test_approved_extension_clones_only_declared_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as raw:
            folder = Path(raw)
            template = folder / "client.xlsx"
            output = folder / "adapted.xlsx"
            mapping_path = folder / "mapping.json"
            workbook = Workbook()
            workbook.active.title = "Event Block"
            workbook.active["A1"] = "Event"
            workbook.active["B1"] = "Value"
            workbook.save(template)
            mapping = {
                "mode": "approved_structural_extension",
                "mapping_id": "approved_event_block",
                "template_sha256": file_sha256(template),
                "sheet_clones": [{
                    "source_sheet": "Event Block",
                    "target_sheet": "Event Block 2",
                    "approved_reason": "The client approved one additional event block sheet.",
                }],
                "cell_writes": [{"sheet": "Event Block 2", "cell": "B1", "value_path": "$.document.version"}],
            }
            mapping_path.write_text(json.dumps(mapping), encoding="utf-8")
            adapted = adapt_workbook(self.plan, template, mapping)
            adapted.save(output)
            report = build_fidelity_report(template, output, mapping, mapping_path)
            self.assertEqual(report["status"], "passed")
            self.assertEqual(load_workbook(output).sheetnames, ["Event Block", "Event Block 2"])

    def test_cli_always_writes_artifact_bound_fidelity_report(self) -> None:
        with tempfile.TemporaryDirectory() as raw:
            folder = Path(raw)
            template = folder / "client.xlsx"
            output = folder / "adapted.xlsx"
            inventory = folder / "inventory.json"
            mapping_path = folder / "mapping.json"
            workbook = Workbook()
            workbook.active.title = "Events"
            workbook.active["A1"] = "Old version"
            workbook.save(template)
            mapping = self.strict_mapping(template, [{"sheet": "Events", "cell": "A1", "value_path": "$.document.version"}])
            mapping_path.write_text(json.dumps(mapping), encoding="utf-8")

            with patch.object(sys, "argv", ["inspect_tracking_plan_template.py", str(template), "--output", str(inventory)]):
                with redirect_stdout(io.StringIO()):
                    self.assertEqual(inspect_main(), 0)
            self.assertEqual(json.loads(inventory.read_text(encoding="utf-8"))["template_sha256"], file_sha256(template))

            plan_path = ROOT / "skill" / "references" / "03-rules" / "example-ga4-tracking-plan.json"
            with patch.object(sys, "argv", [
                "adapt_tracking_plan_workbook.py",
                str(plan_path),
                str(template),
                "--mapping",
                str(mapping_path),
                "--output",
                str(output),
            ]):
                with redirect_stdout(io.StringIO()):
                    self.assertEqual(adapt_main(), 0)
            report_path = folder / "adapted.fidelity.json"
            report = json.loads(report_path.read_text(encoding="utf-8"))
            self.assertEqual(report["status"], "passed")
            self.assertEqual(report["template_sha256"], file_sha256(template))
            self.assertEqual(report["output_sha256"], file_sha256(output))
            self.assertEqual(len(report["mapping_sha256"]), 64)


if __name__ == "__main__":
    unittest.main()
