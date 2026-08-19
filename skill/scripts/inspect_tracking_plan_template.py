from __future__ import annotations

import argparse
import json
import re
import unicodedata
from pathlib import Path
from typing import Any

from contract_utils import sha256_file
from jsonschema import Draft202012Validator
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string
from template_preflight import inspect_template_richness

ROOT = Path(__file__).resolve().parents[1]
MAPPING_SCHEMA = ROOT / "references" / "schema-template-map.json"

ALIASES = {
    "event": {
        "event",
        "event name",
        "event_name",
        "evenement",
        "nom evenement",
        "nom de levenement",
        "nom de l evenement",
    },
    "journey": {"journey", "parcours", "funnel", "etape du parcours"},
    "classification": {"classification", "type evenement", "event type"},
    "definition": {"definition", "description", "explication"},
    "trigger": {
        "trigger",
        "declencheur",
        "condition de declenchement",
        "regle de declenchement",
    },
    "locations": {
        "locations",
        "pages routes components",
        "emplacement",
        "page",
        "url",
        "pages routes composants",
    },
    "variables": {
        "variables",
        "parameters",
        "parametres",
        "variables propres a levenement",
        "event specific variables",
    },
    "variable": {
        "variable",
        "parameter",
        "parametre",
        "nom variable",
        "nom du parametre",
    },
    "scope": {"scope", "portee", "niveau"},
    "type": {"type", "format", "type format"},
    "requirement": {"requirement", "exigence", "statut", "obligation"},
    "condition": {"condition", "condition dexigence", "required when"},
    "values": {
        "possible values rule",
        "valeurs possibles regle",
        "valeurs des variables",
        "valeurs possibles",
    },
    "rule": {"rule", "value rule", "regle", "regle de valeur"},
    "possible_values_or_examples": {
        "possible values or examples",
        "possible values or example",
        "valeurs possibles ou exemples",
        "valeurs possibles ou exemple",
    },
    "example": {"example", "exemple", "example value", "valeur exemple"},
    "concerned_events": {
        "concerned events",
        "evenements concernes",
        "disponibilite par evenement",
    },
    "source_path": {
        "datalayer path source",
        "chemin datalayer source",
        "source",
        "chemin datalayer",
    },
    "notes": {"notes", "note"},
    "datalayer": {
        "datalayer specification",
        "specification datalayer",
        "datalayer",
        "exemple datalayer",
    },
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Inspect a supplied tracking-plan workbook and propose semantic regions.")
    parser.add_argument("template", type=Path)
    parser.add_argument("--output", "-o", type=Path, required=True)
    return parser.parse_args()


def normalize(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]+", " ", text.casefold()).strip()


NORMALIZED_ALIASES = {field: {normalize(value) for value in values} for field, values in ALIASES.items()}


def field_for(value: Any) -> str | None:
    normalized = normalize(value)
    for field, aliases in NORMALIZED_ALIASES.items():
        if normalized in aliases:
            return field
    return None


def sha256(path: Path) -> str:
    """Compatibility name retained for template-adaptation imports."""
    return sha256_file(path)


def sheet_inventory(sheet) -> dict[str, Any]:
    formulas = []
    comments = []
    for cell in sheet._cells.values():
        if isinstance(cell.value, str) and cell.value.startswith("=") and len(formulas) < 100:
            formulas.append(cell.coordinate)
        if cell.comment and len(comments) < 100:
            comments.append(cell.coordinate)
    return {
        "title": sheet.title,
        "state": sheet.sheet_state,
        "max_row": sheet.max_row,
        "max_column": sheet.max_column,
        "merged_ranges": sorted(str(value) for value in sheet.merged_cells.ranges),
        "formula_cells": sorted(formulas),
        "comment_cells": sorted(comments),
        "table_names": sorted(sheet.tables.keys()),
        "data_validation_count": len(sheet.data_validations.dataValidation),
        "conditional_formatting_count": len(sheet.conditional_formatting),
        "image_count": len(sheet._images),
        "chart_count": len(sheet._charts),
        "freeze_panes": str(sheet.freeze_panes or ""),
        "auto_filter": str(sheet.auto_filter.ref or ""),
        "print_area": str(sheet.print_area or ""),
    }


def header_candidates(sheet) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    max_row = min(sheet.max_row, 120)
    max_column = min(sheet.max_column, 50)
    for row in range(1, max_row + 1):
        columns: dict[str, int] = {}
        for column in range(1, max_column + 1):
            field = field_for(sheet.cell(row, column).value)
            if field and field not in columns:
                columns[field] = column
        if len(columns) >= 2:
            candidates.append(
                {
                    "sheet": sheet.title,
                    "header_row": row,
                    "data_start_row": row + 1,
                    "columns": columns,
                    "identity": f"{sheet.title}!{row}",
                }
            )
    return candidates


def _candidate_score(candidate: dict[str, Any], role: str) -> int:
    fields = set(candidate["columns"])
    sheet_name = normalize(candidate["sheet"])
    if role == "event_matrix":
        if not {"event", "definition"} <= fields:
            return -1
        score = 20 + 2 * len(fields & {"journey", "classification", "trigger", "locations", "variables", "notes"})
        score += 5 if any(token in sheet_name for token in ("event matrix", "events", "evenements")) else 0
        score -= 3 if "datalayer" in fields else 0
        return score
    if role == "parameter_reference":
        if "variable" not in fields:
            return -1
        supporting = fields & {
            "scope",
            "type",
            "definition",
            "values",
            "rule",
            "possible_values_or_examples",
            "example",
            "concerned_events",
        }
        if len(supporting) < 3:
            return -1
        score = 20 + 2 * len(supporting)
        score += 5 if any(token in sheet_name for token in ("parameter", "parametre", "variable", "dimension")) else 0
        return score
    if role == "data_layer_table":
        if not {"event", "datalayer"} <= fields:
            return -1
        return 24 + 2 * len(fields & {"definition", "trigger", "locations"}) + (4 if "datalayer" in sheet_name else 0)
    return -1


def _select_candidate(
    candidates: list[dict[str, Any]],
    role: str,
) -> tuple[dict[str, Any], dict[str, Any]]:
    ranked = sorted(
        ((candidate, _candidate_score(candidate, role)) for candidate in candidates),
        key=lambda item: (-item[1], item[0]["identity"]),
    )
    ranked = [item for item in ranked if item[1] >= 0]
    if not ranked:
        return {}, {"semantic_role": role, "selected": None, "confidence": "missing", "alternatives": []}
    top, top_score = ranked[0]
    equally_plausible = [item for item in ranked[1:] if top_score - item[1] < 3]
    confidence = "ambiguous" if equally_plausible else "unique"
    return dict(top), {
        "semantic_role": role,
        "selected": str(top["identity"]),
        "confidence": confidence,
        "alternatives": [str(item[0]["identity"]) for item in equally_plausible],
    }


def classify_regions(
    candidates: list[dict[str, Any]],
) -> tuple[dict[str, Any], dict[str, Any], dict[str, Any], list[dict[str, Any]]]:
    event_matrix, event_preview = _select_candidate(candidates, "event_matrix")
    parameter_reference, parameter_preview = _select_candidate(candidates, "parameter_reference")
    data_layer_table, data_layer_preview = _select_candidate(candidates, "data_layer_table")
    return event_matrix, parameter_reference, data_layer_table, [event_preview, parameter_preview, data_layer_preview]


def _table_binding(sheet: Any, candidate: dict[str, Any]) -> tuple[dict[str, Any], list[str]]:
    matches: list[Any] = []
    header_row = int(candidate["header_row"])
    columns = {int(value) for value in candidate["columns"].values()}
    for table in sheet.tables.values():
        min_col, min_row, max_col, _max_row = range_boundaries(str(table.ref))
        if min_row == header_row and columns and min(columns) >= min_col and max(columns) <= max_col:
            matches.append(table)
    if len(matches) != 1:
        reasons = [] if not matches else [f"Multiple Excel tables contain {candidate['identity']}: " + ", ".join(str(item.name) for item in matches)]
        return {}, reasons
    table = matches[0]
    return {
        "name": str(table.name),
        "ref": str(table.ref),
        "totals_row_shown": bool(table.totalsRowShown),
    }, []


def _existing_data_end(sheet: Any, start_row: int, columns: set[int]) -> int:
    last = start_row - 1
    for row in range(start_row, min(sheet.max_row, start_row + 5000) + 1):
        if any(
            merged.min_row <= row <= merged.max_row
            and len(columns & set(range(merged.min_col, merged.max_col + 1))) >= 2
            for merged in sheet.merged_cells.ranges
        ):
            break
        if any(sheet.cell(row, column).value not in (None, "") for column in columns):
            last = row
            continue
        break
    return last


def _is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _enrich_region(
    workbook: Any,
    candidate: dict[str, Any],
    semantic_role: str,
) -> tuple[dict[str, Any], list[str]]:
    if not candidate:
        return {}, []
    sheet = workbook[str(candidate["sheet"])]
    columns = {str(key): int(value) for key, value in candidate["columns"].items()}
    mapped_columns = set(columns.values())
    table, review = _table_binding(sheet, candidate)
    start = int(candidate["data_start_row"])
    if table:
        _min_col, _min_row, _max_col, table_end = range_boundaries(str(table["ref"]))
        data_end = table_end - (1 if table["totals_row_shown"] else 0)
    else:
        data_end = _existing_data_end(sheet, start, mapped_columns)
    capacity = max(0, data_end - start + 1)
    formatted_prototype = any(sheet.cell(start, column).has_style for column in mapped_columns)
    if not table and formatted_prototype:
        formatted_end = max(data_end, start - 1)
        scan_from = max(start, data_end + 1)
        for row in range(scan_from, min(sheet.max_row, start + 5000) + 1):
            merged_boundary = any(
                merged.min_row <= row <= merged.max_row
                and len(mapped_columns & set(range(merged.min_col, merged.max_col + 1))) >= 2
                for merged in sheet.merged_cells.ranges
            )
            if merged_boundary or any(sheet.cell(row, column).value not in (None, "") for column in mapped_columns):
                break
            if not any(sheet.cell(row, column).has_style for column in mapped_columns):
                break
            formatted_end = row
        data_end = max(data_end, formatted_end)
        capacity = max(0, data_end - start + 1)
    prototype = start if capacity > 0 else None
    inspect_end = max(data_end, start if prototype else start - 1)
    writable: list[str] = []
    protected: list[str] = []
    for row in range(start, inspect_end + 1):
        for column in sorted(mapped_columns):
            cell = sheet.cell(row, column)
            if _is_formula(cell.value):
                protected.append(cell.coordinate)
            else:
                writable.append(cell.coordinate)
    if protected:
        review.append(
            f"Mapped value columns in {candidate['identity']} contain formulas ({', '.join(protected[:8])}); formula cells cannot be writable."
        )
    merged_conflict = any(
        merged.min_row <= start <= merged.max_row and merged.min_col <= max(mapped_columns) and merged.max_col >= min(mapped_columns)
        for merged in sheet.merged_cells.ranges
    )
    downstream_conflict = any(
        any(sheet.cell(row, column).value not in (None, "") for column in mapped_columns)
        or any(
            merged.min_row <= row <= merged.max_row
            and len(mapped_columns & set(range(merged.min_col, merged.max_col + 1))) >= 2
            for merged in sheet.merged_cells.ranges
        )
        for row in range(data_end + 1, sheet.max_row + 1)
    )
    growth_policy = (
        "excel_table"
        if table
        else ("prototype_row" if prototype and not merged_conflict and not downstream_conflict else "fixed_capacity")
    )
    has_validations = bool(sheet.data_validations.dataValidation)
    has_conditional_formatting = bool(sheet.conditional_formatting)
    escaped_sheet_title = sheet.title.replace("'", "''")
    sheet_tokens = {f"'{escaped_sheet_title}'!", f"{sheet.title}!"}
    has_defined_names = any(any(token in str(item.attr_text or "") for token in sheet_tokens) for item in workbook.defined_names.values())
    permitted = {
        "values": True,
        "hyperlinks": semantic_role == "event_matrix",
        "auto_filter": bool(sheet.auto_filter.ref),
        "table_ref": bool(table),
        "data_validation_ranges": bool(table) or (growth_policy == "prototype_row" and has_validations),
        "conditional_formatting_ranges": bool(table) or (growth_policy == "prototype_row" and has_conditional_formatting),
        "defined_names": bool(table) or (growth_policy == "prototype_row" and has_defined_names),
        "row_layout": growth_policy in {"excel_table", "prototype_row"},
        "sheet_title": False,
        "sheet_visibility": False,
    }
    return {
        "semantic_role": semantic_role,
        "sheet": str(candidate["sheet"]),
        "header_row": int(candidate["header_row"]),
        "data_start_row": start,
        "columns": columns,
        "writable_value_cells": sorted(writable),
        "protected_formula_cells": sorted(protected),
        "prototype_data_row": prototype,
        "existing_capacity": capacity,
        "row_growth_policy": growth_policy,
        "table": table,
        "permitted_changes": permitted,
        "overflow_policy": "wrap_preserve_layout",
    }, review


def _right_coordinate(coordinate: str) -> str:
    letters, row = coordinate_from_string(coordinate)
    return f"{get_column_letter(column_index_from_string(letters) + 1)}{row}"


def event_tab_candidate(
    workbook: Any,
    sheet: Any | None = None,
) -> tuple[dict[str, Any] | None, list[str]]:
    # Keep the v2.8 single-sheet helper call valid while allowing workbook-level
    # prototype and defined-name inspection in the v2 mapping contract.
    if sheet is None:
        sheet = workbook
        workbook = sheet.parent
    field_cells: dict[str, str] = {}
    for row in range(1, min(sheet.max_row, 80) + 1):
        for column in range(1, min(sheet.max_column, 20) + 1):
            field = field_for(sheet.cell(row, column).value)
            if field and field not in field_cells:
                field_cells[field] = sheet.cell(row, column).coordinate
    if "event" not in field_cells or len(set(field_cells) & {"definition", "trigger", "locations", "datalayer"}) < 2:
        return None, []
    field_rows = {sheet[coordinate].row for field, coordinate in field_cells.items() if field in {"event", "definition", "trigger", "locations", "datalayer"}}
    if len(field_rows) < 3:
        return None, []
    event_label = sheet[field_cells["event"]]
    value_cell = sheet.cell(event_label.row, event_label.column + 1).coordinate
    parameter_region: dict[str, Any] = {}
    review: list[str] = []
    for candidate in header_candidates(sheet):
        if (
            "variable" in candidate["columns"]
            and len(
                set(candidate["columns"])
                & {
                    "scope",
                    "type",
                    "requirement",
                    "definition",
                    "values",
                    "rule",
                    "possible_values_or_examples",
                }
            )
            >= 3
        ):
            parameter_region, region_review = _enrich_region(workbook, candidate, "event_parameters")
            review.extend(region_review)
            break
    data_layer_cell = ""
    if "datalayer" in field_cells:
        label_cell = sheet[field_cells["datalayer"]]
        for row in range(label_cell.row, min(sheet.max_row, label_cell.row + 5) + 1):
            for column in range(1, min(sheet.max_column, 15) + 1):
                value = sheet.cell(row, column).value
                if isinstance(value, str) and "window.dataLayer" in value:
                    data_layer_cell = sheet.cell(row, column).coordinate
                    break
            if data_layer_cell:
                break
        if not data_layer_cell:
            data_layer_cell = sheet.cell(label_cell.row + 1, 1).coordinate
    direct_value_cells = sorted({_right_coordinate(coordinate) for coordinate in field_cells.values()})
    protected = [coordinate for coordinate in direct_value_cells if _is_formula(sheet[coordinate].value)]
    writable = [coordinate for coordinate in direct_value_cells if coordinate not in protected]
    if data_layer_cell and data_layer_cell not in protected and data_layer_cell not in writable:
        if _is_formula(sheet[data_layer_cell].value):
            protected.append(data_layer_cell)
        else:
            writable.append(data_layer_cell)
    if protected:
        review.append(f"Mapped event-detail values in {sheet.title} contain protected formulas: {', '.join(protected[:8])}.")
    return {
        "semantic_role": "event_detail",
        "sheet": sheet.title,
        "event_name_cell": value_cell,
        "existing_event_name": str(sheet[value_cell].value or "").strip(),
        "reusable": not bool(str(sheet[value_cell].value or "").strip()),
        "field_labels": field_cells,
        "parameter_region": parameter_region,
        "data_layer_cell": data_layer_cell,
        "writable_value_cells": sorted(writable),
        "protected_formula_cells": sorted(protected),
        "permitted_changes": {
            "values": True,
            "hyperlinks": False,
            "auto_filter": bool(parameter_region and sheet.auto_filter.ref),
            "table_ref": bool((parameter_region or {}).get("table")),
            "data_validation_ranges": bool((parameter_region or {}).get("table")),
            "conditional_formatting_ranges": bool((parameter_region or {}).get("table")),
            "defined_names": bool((parameter_region or {}).get("table")),
            "row_layout": bool((parameter_region or {}).get("table")),
            "sheet_title": True,
            "sheet_visibility": True,
        },
    }, review


def _prototype_event_tab(event_tabs: list[dict[str, Any]], workbook: Any) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    candidates = [
        item
        for item in event_tabs
        if item.get("reusable") is True
        and workbook[str(item["sheet"])].sheet_state != "visible"
        and any(token in normalize(item["sheet"]) for token in ("event template", "event prototype", "modele evenement"))
    ]
    if len(candidates) != 1:
        return {}, event_tabs
    prototype = candidates[0]
    return {
        "sheet": str(prototype["sheet"]),
        "cloning_allowed": True,
        "title_policy": "event_name",
        "mapping": prototype,
    }, event_tabs


def validate_template_mapping(mapping: dict[str, Any]) -> list[str]:
    schema = json.loads(MAPPING_SCHEMA.read_text(encoding="utf-8"))
    validator = Draft202012Validator(schema)
    return [
        f"{'/'.join(str(part) for part in error.absolute_path) or '<root>'}: {error.message}"
        for error in sorted(validator.iter_errors(mapping), key=lambda item: list(item.absolute_path))
    ]


def normalize_template_mapping(mapping: dict[str, Any], template: Path) -> dict[str, Any]:
    """Reinspect and safely enrich deterministic v1.1 maps; v2 maps pass through unchanged."""
    if str(mapping.get("mapping_version", "")) == "2.0.0":
        return mapping
    if str(mapping.get("mapping_version", "")) != "1.1":
        raise ValueError("Unsupported template mapping version; inspect the current workbook again.")
    upgraded = inspect(template)
    workbook = load_workbook(
        template,
        data_only=False,
        read_only=False,
        keep_links=True,
        keep_vba=template.suffix.lower() == ".xlsm",
    )
    old_regions = mapping.get("regions", {}) if isinstance(mapping.get("regions"), dict) else {}
    role_names = {
        "event_matrix": "event_matrix",
        "parameter_reference": str((old_regions.get("parameter_reference") or {}).get("semantic_role", "all_used_parameters")),
        "data_layer_table": "data_layer_examples",
    }
    resolved_roles: set[str] = set()
    for name, semantic_role in role_names.items():
        old = old_regions.get(name) or {}
        if not old:
            continue
        candidate = {
            "sheet": str(old["sheet"]),
            "header_row": int(old["header_row"]),
            "data_start_row": int(old["data_start_row"]),
            "columns": {str(key): int(value) for key, value in old["columns"].items()},
            "identity": f"{old['sheet']}!{old['header_row']}",
        }
        enriched, review = _enrich_region(workbook, candidate, semantic_role)
        upgraded["regions"][name] = enriched
        upgraded["review_required"].extend(review)
        resolved_roles.add(name)
    if old_regions.get("event_tabs"):
        by_sheet = {str(item["sheet"]): item for item in upgraded["regions"]["event_tabs"]}
        selected = [by_sheet[str(item["sheet"])] for item in old_regions["event_tabs"] if str(item.get("sheet", "")) in by_sheet]
        if len(selected) == len(old_regions["event_tabs"]):
            upgraded["regions"]["event_tabs"] = selected
    prefixes = {
        "event_matrix": ("Ambiguous event_matrix", "No semantic Event Matrix"),
        "parameter_reference": ("Ambiguous parameter_reference", "No semantic Parameter Reference"),
        "data_layer_table": ("Ambiguous data_layer_table",),
    }
    upgraded["review_required"] = [
        item
        for item in upgraded["review_required"]
        if not any(item.startswith(prefix) for role in resolved_roles for prefix in prefixes[role])
    ]
    errors = validate_template_mapping(upgraded)
    if errors:
        raise ValueError("Legacy template mapping could not be safely upgraded:\n- " + "\n- ".join(errors))
    return upgraded


def inspect(path: Path) -> dict[str, Any]:
    workbook = load_workbook(
        path,
        data_only=False,
        read_only=False,
        keep_links=True,
        keep_vba=path.suffix.lower() == ".xlsm",
    )
    candidates = [candidate for sheet in workbook.worksheets for candidate in header_candidates(sheet)]
    event_matrix_raw, parameter_reference_raw, data_layer_table_raw, mapping_preview = classify_regions(candidates)
    review: list[str] = []
    event_matrix, region_review = _enrich_region(workbook, event_matrix_raw, "event_matrix")
    review.extend(region_review)
    parameter_reference, region_review = _enrich_region(workbook, parameter_reference_raw, "all_used_parameters")
    review.extend(region_review)
    data_layer_table, region_review = _enrich_region(workbook, data_layer_table_raw, "data_layer_examples")
    review.extend(region_review)
    event_tabs: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        candidate, event_review = event_tab_candidate(workbook, sheet)
        review.extend(event_review)
        if candidate is not None:
            event_tabs.append(candidate)
    event_tab_prototype, event_tabs = _prototype_event_tab(event_tabs, workbook)
    for preview in mapping_preview:
        if preview["confidence"] == "ambiguous":
            review.append(
                f"Ambiguous {preview['semantic_role']} mapping: selected {preview['selected']} but similarly plausible regions are "
                + ", ".join(preview["alternatives"])
                + "."
            )
    if not event_matrix:
        review.append("No semantic Event Matrix region was recognized.")
    if not parameter_reference:
        review.append("No semantic Parameter Reference region was recognized.")
    elif "custom dimension" in normalize(parameter_reference.get("sheet", "")):
        review.append(
            "The recognized parameter sheet is named like a custom-dimensions registry. "
            "The mapping defaults to semantic_role=all_used_parameters; change it to "
            "custom_parameters_only only when the template owner explicitly confirms that contract."
        )
    if not event_tabs and not data_layer_table and not event_tab_prototype:
        review.append("No event-tab or dataLayer-table region was recognized; do not add sections without template approval.")
    if event_tab_prototype and not data_layer_table and not event_tab_prototype.get("mapping", {}).get("data_layer_cell"):
        review.append("The event-tab prototype has no mapped dataLayer example location; cloning it cannot produce implementation-complete event tabs.")
    preflight = inspect_template_richness(path)
    result = {
        "mapping_version": "2.0.0",
        "template": {
            "path": str(path.resolve()),
            "sha256": sha256(path),
            "extension": path.suffix.lower(),
        },
        "sheets": [sheet_inventory(sheet) for sheet in workbook.worksheets],
        "regions": {
            "event_matrix": event_matrix,
            "parameter_reference": parameter_reference,
            "data_layer_table": data_layer_table,
            "event_tabs": event_tabs,
            "event_tab_prototype": event_tab_prototype,
        },
        "policy": {
            "preserve_unmapped_content": True,
            "embed_internal_model": True,
            "writer": "auto",
            "allow_reduced_fidelity_fallback": False,
            "rendered_layout_check": "when_deterministic_renderer_available",
        },
        "review_required": review,
        "mapping_preview": mapping_preview,
        "preflight": preflight,
    }
    errors = validate_template_mapping(result)
    if errors:
        raise ValueError("Generated template mapping failed its contract:\n- " + "\n- ".join(errors))
    return result


def main() -> int:
    args = parse_args()
    result = inspect(args.template)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(result, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    print(args.output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
