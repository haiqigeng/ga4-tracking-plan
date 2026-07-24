from __future__ import annotations

import argparse
import copy
import json
import sys
from pathlib import Path
from typing import Any

from create_default_template import (
    ACCENT,
    BORDER,
    PALE,
    TEAL,
    TEXT,
    apply_header,
    apply_table_row,
    apply_title,
    fill,
    set_widths,
)
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from openpyxl.utils import quote_sheetname
from tracking_plan_model import (
    classification_label,
    compact_value,
    datalayer_code,
    event_journey_names,
    label,
    load_json,
    location_text,
    parameter_reference_rows,
    requirement_label,
    safe_sheet_title,
    scope_label,
    value_rule_text,
    workbook_language,
)
from validate_tracking_plan import render_text, validate_plan

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_TEMPLATE = ROOT / "assets" / "default-tracking-plan.xlsx"
MODEL_SHEET = "__tracking_plan_model"
MODEL_MARKER = "ga4-tracking-plan/model"
MODEL_CELL_LIMIT = 30000
MATRIX_WIDTHS = [20, 20, 15, 36, 40, 32, 28]
REFERENCE_WIDTHS = [22, 11, 11, 38, 22, 38, 28]
EVENT_WIDTHS = [20, 10, 10, 12, 23, 38, 36, 22, 32]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Render a validated GA4 tracking plan through the approved default XLSX asset.")
    parser.add_argument("plan", type=Path)
    parser.add_argument("--output", "-o", type=Path, required=True)
    parser.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE)
    parser.add_argument("--changes", type=Path)
    parser.add_argument("--screenshot-dir", type=Path)
    return parser.parse_args()


def clear_data_rows(ws, start_row: int) -> None:
    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)


def copy_row_style(ws, source_row: int, target_row: int, columns: int) -> None:
    for column in range(1, columns + 1):
        source = ws.cell(source_row, column)
        target = ws.cell(target_row, column)
        if source.has_style:
            target._style = copy.copy(source._style)
        target.number_format = source.number_format
        target.protection = copy.copy(source.protection)
        target.alignment = copy.copy(source.alignment)
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def content_row_height(
    values: list[Any],
    widths: list[int],
    *,
    minimum: int,
    maximum: int,
) -> int:
    max_lines = 1
    for value, width in zip(values, widths):
        text = compact_value(value)
        lines = 0
        for line in text.splitlines() or [""]:
            lines += max(1, (len(line) + max(8, width) - 1) // max(8, width))
        max_lines = max(max_lines, lines)
    return min(maximum, max(minimum, 10 + max_lines * 14))


def link(cell, sheet_name: str) -> None:
    cell.hyperlink = f"#{quote_sheetname(sheet_name)}!A1"
    cell.font = Font(color=TEAL, bold=True, underline="single")


def locations_for_matrix(event: dict[str, Any]) -> str:
    return location_text(event)


def variables_for_matrix(plan: dict[str, Any], event: dict[str, Any]) -> str:
    return "\n".join(
        f"{parameter.get('name')} ({requirement_label(plan, str(parameter.get('requirement', '')))})"
        for parameter in event.get("parameters", [])
        if isinstance(parameter, dict)
    )


def fill_guide(wb, plan: dict[str, Any], event_sheet_names: dict[str, str]) -> None:
    ws = wb["Guide"]
    doc = plan["document"]
    language = workbook_language(plan)
    apply_title(
        ws,
        str(doc["title"]),
        "Human contract for analyst review, maintenance, and dataLayer implementation."
        if language == "en"
        else "Contrat humain pour la revue, la maintenance et l'implémentation dataLayer.",
        7,
    )
    guide_rows = [
        ("document", doc["title"]),
        ("version", doc["version"]),
        ("date", doc["date"]),
        ("scope", doc["scope"]),
        ("target_state", doc["target_state"]),
        ("language", doc["language"]),
        ("analyst_entry", label(plan, "analyst_entry_value")),
        ("developer_entry", label(plan, "developer_entry_value")),
        (
            "datalayer_convention",
            f"{plan['data_layer_convention']['name']} ({plan['data_layer_convention']['origin']})"
            + (f"\n{plan['data_layer_convention'].get('notes')}" if plan["data_layer_convention"].get("notes") else ""),
        ),
    ]
    for row, (key, value) in enumerate(guide_rows, 4):
        ws.cell(row, 1, label(plan, key))
        ws.cell(row, 2, value)
    ws.cell(14, 1, label(plan, "journeys"))
    journey_headers = [
        label(plan, "journey"),
        label(plan, "scope"),
        label(plan, "business_goal"),
        label(plan, "urls"),
    ]
    for column, header_value in enumerate(journey_headers, 1):
        ws.cell(15, column, header_value)
    if ws.max_row > 16:
        ws.delete_rows(17, ws.max_row - 16)
    journeys = plan.get("journeys", [])
    for offset, journey in enumerate(journeys):
        row = 16 + offset
        if row > 16:
            copy_row_style(ws, 16, row, 4)
        ws.cell(row, 1, journey.get("name", ""))
        ws.cell(row, 2, journey.get("scope", ""))
        ws.cell(row, 3, journey.get("business_goal", ""))
        ws.cell(row, 4, "\n".join(str(value) for value in journey.get("urls", [])))
    ws.auto_filter.ref = f"A15:D{max(16, 15 + len(journeys))}"
    link(ws.cell(10, 2), "Event Matrix")
    if event_sheet_names:
        link(ws.cell(11, 2), next(iter(event_sheet_names.values())))


def fill_event_matrix(wb, plan: dict[str, Any], event_sheet_names: dict[str, str]) -> None:
    ws = wb["Event Matrix"]
    apply_title(
        ws,
        label(plan, "event_matrix"),
        "One row per manually implemented event or context push."
        if workbook_language(plan) == "en"
        else "Une ligne par événement implémenté manuellement ou push de contexte.",
        7,
    )
    headers = [
        label(plan, "journey"),
        label(plan, "event"),
        label(plan, "classification"),
        label(plan, "definition"),
        label(plan, "trigger"),
        label(plan, "locations"),
        label(plan, "variables"),
    ]
    for column, header_value in enumerate(headers, 1):
        ws.cell(4, column, header_value)
    clear_data_rows(ws, 5)
    for index, event in enumerate(plan.get("events", []), 5):
        apply_table_row(ws, index, 7)
        event_name = str(event.get("event_name", ""))
        ws.cell(index, 1, " | ".join(event_journey_names(plan, event)))
        ws.cell(index, 2, event_name)
        ws.cell(index, 3, classification_label(plan, str(event.get("classification", ""))))
        ws.cell(index, 4, event.get("definition", ""))
        ws.cell(index, 5, event.get("trigger", ""))
        ws.cell(index, 6, locations_for_matrix(event))
        ws.cell(index, 7, variables_for_matrix(plan, event))
        if event_name in event_sheet_names:
            link(ws.cell(index, 2), event_sheet_names[event_name])
        ws.row_dimensions[index].height = content_row_height(
            [ws.cell(index, column).value for column in range(1, 8)],
            MATRIX_WIDTHS,
            minimum=70,
            maximum=170,
        )
    ws.auto_filter.ref = f"A4:G{max(5, 4 + len(plan.get('events', [])))}"


def fill_parameter_reference(wb, plan: dict[str, Any]) -> None:
    ws = wb["Parameter Reference"]
    apply_title(
        ws,
        label(plan, "parameter_reference"),
        "Definitions and values for parameters actually used by the event specifications."
        if workbook_language(plan) == "en"
        else "Définitions et valeurs des variables réellement utilisées par les spécifications.",
        7,
    )
    headers = [
        label(plan, "variable"),
        label(plan, "scope_label"),
        label(plan, "type"),
        label(plan, "definition"),
        label(plan, "example"),
        label(plan, "values"),
        label(plan, "concerned_events"),
    ]
    for column, header_value in enumerate(headers, 1):
        ws.cell(4, column, header_value)
    clear_data_rows(ws, 5)
    rows = parameter_reference_rows(plan)
    for index, row in enumerate(rows, 5):
        apply_table_row(ws, index, 7)
        ws.cell(index, 1, row["name"])
        ws.cell(index, 2, scope_label(plan, str(row["scope"])))
        ws.cell(index, 3, row["type"])
        ws.cell(index, 4, row["definition"])
        ws.cell(index, 5, row["example"])
        ws.cell(index, 6, row["values"])
        ws.cell(index, 7, " | ".join(row["events"]))
        ws.row_dimensions[index].height = content_row_height(
            [ws.cell(index, column).value for column in range(1, 8)],
            REFERENCE_WIDTHS,
            minimum=60,
            maximum=190,
        )
    ws.auto_filter.ref = f"A4:G{max(5, 4 + len(rows))}"


def add_section(ws, row: int, text: str, columns: int = 9) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=columns)
    cell = ws.cell(row, 1, text)
    fill(cell, ACCENT)
    cell.font = Font(color=TEAL, bold=True, size=12)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 26


def add_code_block(ws, row: int, code: str, columns: int = 9) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=columns)
    cell = ws.cell(row, 1, code)
    fill(cell, PALE)
    cell.font = Font(name="Consolas", color=TEXT, size=10)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.border = BORDER
    line_count = max(5, code.count("\n") + 1)
    ws.row_dimensions[row].height = min(390, 15 * line_count)


def maybe_add_screenshot(ws, row: int, event: dict[str, Any], screenshot_dir: Path | None) -> int:
    screenshots = event.get("screenshots", [])
    if not screenshots or screenshot_dir is None:
        return row
    existing: list[Path] = []
    for value in screenshots:
        path = Path(str(value))
        candidate = path if path.is_absolute() else screenshot_dir / path
        if candidate.exists():
            existing.append(candidate)
    if not existing:
        return row
    add_section(ws, row, "Visual reference", 9)
    row += 1
    for path in existing:
        image = XLImage(str(path))
        image.width = min(640, image.width)
        image.height = min(360, image.height)
        ws.add_image(image, f"A{row}")
        ws.row_dimensions[row].height = 280
        row += 1
    return row


def fill_event_sheet(
    ws,
    plan: dict[str, Any],
    event: dict[str, Any],
    screenshot_dir: Path | None,
) -> None:
    event_name = str(event.get("event_name", ""))
    apply_title(
        ws,
        f'Event: "{event_name}"',
        "Authoritative event, parameter, and dataLayer specification."
        if workbook_language(plan) == "en"
        else "Spécification de référence de l'événement, de ses variables et du dataLayer.",
        9,
    )
    values = [
        (label(plan, "event"), event_name),
        (label(plan, "classification"), classification_label(plan, str(event.get("classification", "")))),
        (label(plan, "journey"), " | ".join(event_journey_names(plan, event))),
        (label(plan, "definition"), event.get("definition", "")),
        (label(plan, "trigger"), event.get("trigger", "")),
        (label(plan, "locations"), location_text(event)),
        (label(plan, "notes"), event.get("notes", "")),
    ]
    for row, (row_label, value) in enumerate(values, 3):
        ws.cell(row, 1, row_label)
        ws.cell(row, 2, value)
    if not str(event.get("notes", "")).strip():
        ws.row_dimensions[9].hidden = True
    headers = [
        label(plan, "variable"),
        label(plan, "scope_label"),
        label(plan, "type"),
        label(plan, "requirement"),
        label(plan, "condition"),
        label(plan, "definition"),
        label(plan, "values"),
        label(plan, "example"),
        label(plan, "source_path"),
    ]
    for column, header_value in enumerate(headers, 1):
        ws.cell(11, column, header_value)
    parameters = event.get("parameters", [])
    for offset, parameter in enumerate(parameters):
        row = 12 + offset
        if row > 12:
            copy_row_style(ws, 12, row, 9)
        ws.cell(row, 1, parameter.get("name", ""))
        ws.cell(row, 2, scope_label(plan, str(parameter.get("scope", ""))))
        ws.cell(row, 3, parameter.get("type", ""))
        ws.cell(row, 4, requirement_label(plan, str(parameter.get("requirement", ""))))
        ws.cell(row, 5, parameter.get("condition", ""))
        ws.cell(row, 6, parameter.get("definition", ""))
        ws.cell(row, 7, value_rule_text(parameter, plan))
        ws.cell(row, 8, compact_value(parameter.get("example")))
        source_text = str(parameter.get("data_layer_path", ""))
        if parameter.get("source"):
            source_text += f"\n{parameter.get('source')}"
        ws.cell(row, 9, source_text)
        ws.row_dimensions[row].height = content_row_height(
            [ws.cell(row, column).value for column in range(1, 10)],
            EVENT_WIDTHS,
            minimum=64,
            maximum=210,
        )
    if not parameters:
        for column in range(1, 10):
            ws.cell(12, column, "")
    data_layer_title_row = 13 + max(1, len(parameters))
    add_section(ws, data_layer_title_row, label(plan, "datalayer"), 9)
    add_code_block(ws, data_layer_title_row + 1, datalayer_code(event), 9)
    maybe_add_screenshot(ws, data_layer_title_row + 3, event, screenshot_dir)
    ws.auto_filter.ref = f"A11:I{11 + max(1, len(parameters))}"


def add_change_log(wb, plan: dict[str, Any], changes: list[dict[str, Any]]) -> None:
    name = label(plan, "change_log")
    ws = wb.create_sheet(name, 3)
    apply_title(
        ws,
        name,
        "Semantic changes from the previous consolidated plan."
        if workbook_language(plan) == "en"
        else "Modifications sémantiques depuis le précédent plan consolidé.",
        7,
    )
    headers = [
        label(plan, "action"),
        label(plan, "entity"),
        label(plan, "key"),
        label(plan, "summary"),
        label(plan, "before"),
        label(plan, "after"),
        label(plan, "event"),
    ]
    for column, header_value in enumerate(headers, 1):
        ws.cell(4, column, header_value)
    apply_header(ws, 4, 7)
    for index, change in enumerate(changes, 5):
        apply_table_row(ws, index, 7)
        key = str(change.get("key", ""))
        event_name = key.split(":", 1)[0] if ":" in key else ""
        ws.cell(index, 1, change.get("action", ""))
        ws.cell(index, 2, change.get("entity", ""))
        ws.cell(index, 3, key)
        ws.cell(index, 4, change.get("summary", ""))
        ws.cell(index, 5, compact_value(change.get("before")))
        ws.cell(index, 6, compact_value(change.get("after")))
        ws.cell(index, 7, event_name)
        ws.row_dimensions[index].height = 56
    set_widths(ws, [16, 18, 32, 54, 42, 42, 24])
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:G{max(5, 4 + len(changes))}"
    ws.sheet_view.showGridLines = False


def load_changes(path: Path | None, plan: dict[str, Any]) -> list[dict[str, Any]]:
    if path:
        value = load_json(path)
        changes = value.get("changes", [])
        return [item for item in changes if isinstance(item, dict)]
    return [item for item in plan.get("changes", []) if isinstance(item, dict)]


def embed_model(wb, plan: dict[str, Any]) -> None:
    """Embed the canonical model for lossless maintenance without exposing machinery."""
    if MODEL_SHEET in wb.sheetnames:
        del wb[MODEL_SHEET]
    ws = wb.create_sheet(MODEL_SHEET)
    ws.cell(1, 1, MODEL_MARKER)
    ws.cell(1, 2, str(plan.get("schema_version", "")))
    payload = json.dumps(plan, ensure_ascii=False, separators=(",", ":"))
    for index in range(0, len(payload), MODEL_CELL_LIMIT):
        ws.cell(2 + index // MODEL_CELL_LIMIT, 1, payload[index : index + MODEL_CELL_LIMIT])
    ws.sheet_state = "veryHidden"


def build_workbook(
    plan: dict[str, Any],
    template: Path = DEFAULT_TEMPLATE,
    changes: list[dict[str, Any]] | None = None,
    screenshot_dir: Path | None = None,
):
    if not template.exists():
        raise FileNotFoundError(f"Default workbook asset not found: {template}")
    wb = load_workbook(template)
    event_template = wb["__EVENT_TEMPLATE"]
    event_sheet_names: dict[str, str] = {}
    used = set(wb.sheetnames)
    for event in plan.get("events", []):
        event_name = str(event.get("event_name", ""))
        sheet_name = safe_sheet_title(event_name, used)
        used.add(sheet_name)
        ws = wb.copy_worksheet(event_template)
        ws.title = sheet_name
        ws.sheet_state = "visible"
        fill_event_sheet(ws, plan, event, screenshot_dir)
        event_sheet_names[event_name] = sheet_name
    wb.remove(event_template)
    fill_guide(wb, plan, event_sheet_names)
    fill_event_matrix(wb, plan, event_sheet_names)
    fill_parameter_reference(wb, plan)
    if changes:
        add_change_log(wb, plan, changes)
    if workbook_language(plan) == "fr":
        wb["Parameter Reference"].title = label(plan, "parameter_reference")
    embed_model(wb, plan)
    wb.active = wb.sheetnames.index("Guide")
    return wb


def main() -> int:
    args = parse_args()
    try:
        plan = load_json(args.plan)
        issues = validate_plan(plan)
        if issues:
            print(render_text(issues), file=sys.stderr)
        if any(item.severity == "error" for item in issues):
            return 1
        changes = load_changes(args.changes, plan)
        workbook = build_workbook(
            plan,
            template=args.template,
            changes=changes,
            screenshot_dir=args.screenshot_dir,
        )
        args.output.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(args.output)
    except (OSError, ValueError, KeyError, json.JSONDecodeError) as error:
        print(str(error), file=sys.stderr)
        return 2
    print(args.output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
