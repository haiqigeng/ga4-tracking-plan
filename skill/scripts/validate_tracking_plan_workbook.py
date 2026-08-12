from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

from import_tracking_plan_workbook import (
    MODEL_SHEET,
    PROJECTION_SHEET,
    import_workbook,
)
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string
from tracking_plan_model import (
    classification_label,
    combined_value_rule_text,
    compact_value,
    datalayer_code,
    event_journey_names,
    label,
    load_json,
    location_text,
    parameter_reference_rows,
    possible_values_or_example,
    requirement_label,
    scope_label,
    value_rule_text,
)


def _value(sheet: Any, row: int, column: int) -> str:
    value = sheet.cell(row, column).value
    return "" if value is None else str(value).strip()


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _merged_range_errors(workbook: Any) -> list[str]:
    errors: list[str] = []
    for sheet in workbook.worksheets:
        ranges = list(sheet.merged_cells.ranges)
        for index, left in enumerate(ranges):
            for right in ranges[index + 1 :]:
                overlaps = (
                    left.min_row <= right.max_row
                    and right.min_row <= left.max_row
                    and left.min_col <= right.max_col
                    and right.min_col <= left.max_col
                )
                if overlaps:
                    errors.append(
                        f'Overlapping merged ranges in "{sheet.title}": {left} and {right}.'
                    )
    return errors


def _event_matrix_value(plan: dict[str, Any], event: dict[str, Any], field: str) -> Any:
    return {
        "journey": " | ".join(event_journey_names(plan, event)),
        "event": event.get("event_name", ""),
        "classification": classification_label(plan, str(event.get("classification", ""))),
        "definition": event.get("definition", ""),
        "trigger": event.get("trigger", ""),
        "locations": location_text(event),
        "variables": "\n".join(
            f"{parameter.get('name')} ({requirement_label(plan, str(parameter.get('requirement', '')))})"
            for parameter in event.get("parameters", [])
            if isinstance(parameter, dict)
        ),
        "datalayer": datalayer_code(event, plan.get("data_layer_convention")),
        "notes": event.get("notes", ""),
    }.get(field, "")


def _reference_value(plan: dict[str, Any], row: dict[str, Any], field: str) -> Any:
    return {
        "variable": row.get("name", ""),
        "scope": scope_label(plan, str(row.get("scope", ""))),
        "type": row.get("type", ""),
        "definition": row.get("definition", ""),
        "rule": row.get("rule", ""),
        "possible_values_or_examples": row.get("possible_values_or_example", ""),
        "example": row.get("possible_values_or_example", ""),
        "values": f"{row.get('rule', '')}\n{row.get('possible_values_or_example', '')}".strip(),
        "concerned_events": " | ".join(row.get("events", [])),
    }.get(field, "")


def _parameter_value(plan: dict[str, Any], parameter: dict[str, Any], field: str) -> Any:
    source = str(parameter.get("data_layer_path", ""))
    if parameter.get("source"):
        source += f"\n{parameter.get('source')}"
    return {
        "variable": parameter.get("name", ""),
        "scope": scope_label(plan, str(parameter.get("scope", ""))),
        "type": parameter.get("type", ""),
        "requirement": requirement_label(plan, str(parameter.get("requirement", ""))),
        "condition": parameter.get("condition", ""),
        "definition": parameter.get("definition", ""),
        "rule": value_rule_text(parameter, plan),
        "possible_values_or_examples": possible_values_or_example(parameter),
        "values": combined_value_rule_text(parameter, plan),
        "example": compact_value(parameter.get("example")),
        "source_path": source,
    }.get(field, "")


def _validate_region(
    workbook: Any,
    region: dict[str, Any],
    rows: list[dict[str, Any]],
    value_for: Any,
    label_text: str,
) -> list[str]:
    if not region:
        return [f"Missing mapped {label_text} region."]
    sheet_name = str(region.get("sheet", ""))
    if sheet_name not in workbook.sheetnames:
        return [f"Mapped {label_text} sheet is missing: {sheet_name}."]
    sheet = workbook[sheet_name]
    start = int(region["data_start_row"])
    columns = {str(field): int(column) for field, column in region.get("columns", {}).items()}
    errors: list[str] = []
    for offset, item in enumerate(rows):
        row = start + offset
        for field, column in columns.items():
            expected = _text(value_for(item, field))
            observed = _value(sheet, row, column)
            if observed != expected:
                errors.append(f"Mapped {label_text} differs at {sheet_name}!{sheet.cell(row, column).coordinate} ({field}).")
    return errors


def _right_coordinate(coordinate: str) -> tuple[int, int]:
    letters, row = coordinate_from_string(coordinate)
    return int(row), column_index_from_string(letters) + 1


def _bind_mapped_event_tabs(
    workbook: Any,
    mappings: list[Any],
    event_by_name: dict[str, dict[str, Any]],
) -> dict[str, tuple[Any, dict[str, Any]]]:
    used_sheets: set[str] = set()
    bound_tabs: dict[str, tuple[Any, dict[str, Any]]] = {}
    for event_mapping in mappings:
        if not isinstance(event_mapping, dict) or not event_mapping.get("event_name_cell"):
            continue
        coordinate = str(event_mapping["event_name_cell"])
        mapped_title = str(event_mapping.get("sheet", ""))
        candidates = [workbook[mapped_title]] if mapped_title in workbook.sheetnames else []
        candidates.extend(
            sheet
            for sheet in workbook.worksheets
            if sheet.title not in used_sheets
            and sheet.title not in {MODEL_SHEET, PROJECTION_SHEET}
            and sheet.title != mapped_title
        )
        for sheet in candidates:
            event_name = _text(sheet[coordinate].value)
            if event_name not in event_by_name or event_name in bound_tabs:
                continue
            bound_tabs[event_name] = (sheet, event_mapping)
            used_sheets.add(sheet.title)
            break
    return bound_tabs


def _validate_mapped_event_tab(
    workbook: Any,
    plan: dict[str, Any],
    event: dict[str, Any],
    sheet: Any,
    event_mapping: dict[str, Any],
) -> list[str]:
    event_name = str(event.get("event_name", ""))
    errors: list[str] = []
    direct_values = {
        "event": event_name,
        "journey": " | ".join(event_journey_names(plan, event)),
        "classification": classification_label(plan, str(event.get("classification", ""))),
        "definition": event.get("definition", ""),
        "trigger": event.get("trigger", ""),
        "locations": location_text(event),
        "notes": event.get("notes", ""),
    }
    for field, coordinate in event_mapping.get("field_labels", {}).items():
        if field not in direct_values:
            continue
        row, column = _right_coordinate(str(coordinate))
        if _value(sheet, row, column) != _text(direct_values[field]):
            errors.append(f'Mapped event-tab field "{field}" differs for "{event_name}".')
    parameter_region = event_mapping.get("parameter_region") or {}
    if parameter_region:
        parameters = [item for item in event.get("parameters", []) if isinstance(item, dict)]
        errors.extend(
            _validate_region(
                workbook,
                {**parameter_region, "sheet": sheet.title},
                parameters,
                lambda parameter, field: _parameter_value(plan, parameter, field),
                f'event-tab parameters for "{event_name}"',
            )
        )
    code_cell = str(event_mapping.get("data_layer_cell", ""))
    expected_code = datalayer_code(event, plan.get("data_layer_convention")).strip()
    if code_cell and _text(sheet[code_cell].value) != expected_code:
        errors.append(f'dataLayer code differs for mapped event tab "{event_name}".')
    return errors


def _validate_mapped_workbook(workbook: Any, plan: dict[str, Any], mapping: dict[str, Any]) -> list[str]:
    errors: list[str] = []
    regions = mapping.get("regions", {}) if isinstance(mapping.get("regions"), dict) else {}
    events = [event for event in plan.get("events", []) if isinstance(event, dict)]
    event_by_name = {str(event.get("event_name", "")): event for event in events}
    event_matrix = regions.get("event_matrix") or {}
    errors.extend(
        _validate_region(
            workbook,
            event_matrix,
            events,
            lambda event, field: _event_matrix_value(plan, event, field),
            "Event Matrix",
        )
    )
    reference = regions.get("parameter_reference") or {}
    role = str(reference.get("semantic_role", "all_used_parameters"))
    try:
        reference_rows = parameter_reference_rows(plan, role)
    except ValueError as error:
        errors.append(str(error))
        reference_rows = []
    errors.extend(
        _validate_region(
            workbook,
            reference,
            reference_rows,
            lambda row, field: _reference_value(plan, row, field),
            "Parameter Reference",
        )
    )
    data_layer_table = regions.get("data_layer_table") or {}
    if data_layer_table:
        errors.extend(
            _validate_region(
                workbook,
                data_layer_table,
                events,
                lambda event, field: _event_matrix_value(plan, event, field),
                "dataLayer table",
            )
        )

    bound_tabs = _bind_mapped_event_tabs(
        workbook,
        regions.get("event_tabs", []),
        event_by_name,
    )

    if not data_layer_table:
        missing_tabs = sorted(set(event_by_name) - set(bound_tabs))
        if missing_tabs:
            errors.append("Mapped event tabs are missing canonical events: " + ", ".join(missing_tabs) + ".")
    for event_name, (sheet, event_mapping) in bound_tabs.items():
        errors.extend(
            _validate_mapped_event_tab(
                workbook,
                plan,
                event_by_name[event_name],
                sheet,
                event_mapping,
            )
        )
    return errors


def validate_workbook(
    path: Path,
    plan: dict[str, Any],
    mapping: dict[str, Any] | None = None,
) -> list[str]:
    workbook = load_workbook(path, data_only=False, read_only=False)
    errors = _merged_range_errors(workbook)
    try:
        imported = import_workbook(path)
    except (ValueError, OSError, json.JSONDecodeError) as error:
        errors.append(str(error))
        return errors
    if imported != plan:
        errors.append("Embedded canonical model does not equal the delivery plan.")
    for internal in (MODEL_SHEET, PROJECTION_SHEET):
        if internal not in workbook.sheetnames:
            errors.append(f"Missing internal maintenance sheet {internal}.")
        elif workbook[internal].sheet_state != "veryHidden":
            errors.append(f"Internal maintenance sheet {internal} must be veryHidden.")

    if mapping is not None:
        errors.extend(_validate_mapped_workbook(workbook, plan, mapping))
        return errors

    if workbook.properties.creator != "ga4-tracking-plan":
        return errors

    reference_name = label(plan, "parameter_reference")
    required_sheets = {"Guide", "Event Matrix", reference_name}
    missing = required_sheets - set(workbook.sheetnames)
    if missing:
        errors.append("Default workbook is missing: " + ", ".join(sorted(missing)))
        return errors

    matrix = workbook["Event Matrix"]
    expected_headers = [label(plan, "event"), label(plan, "definition")]
    if [_value(matrix, 4, column) for column in (1, 2)] != expected_headers:
        errors.append("Event Matrix headers do not match the lean two-column contract.")
    if any(_value(matrix, 4, column) for column in range(3, matrix.max_column + 1)):
        errors.append("Event Matrix contains an unapproved visible column after Definition.")
    events = [event for event in plan.get("events", []) if isinstance(event, dict)]
    observed_matrix = [(_value(matrix, row, 1), _value(matrix, row, 2)) for row in range(5, 5 + len(events))]
    expected_matrix = [(str(event.get("event_name", "")), str(event.get("definition", "")).strip()) for event in events]
    if observed_matrix != expected_matrix:
        errors.append("Event Matrix rows differ from the canonical events and definitions.")

    reference = workbook[reference_name]
    expected_reference_headers = [
        label(plan, "variable"),
        label(plan, "scope_label"),
        label(plan, "type"),
        label(plan, "definition"),
        label(plan, "rule"),
        label(plan, "possible_values_or_examples"),
    ]
    if [_value(reference, 4, column) for column in range(1, 7)] != expected_reference_headers:
        errors.append("Parameter Reference headers do not match the lean six-column contract.")
    if any(_value(reference, 4, column) for column in range(7, reference.max_column + 1)):
        errors.append("Parameter Reference contains an unapproved visible column after possible values or examples.")
    expected_reference = [
        (
            str(row["name"]),
            scope_label(plan, str(row["scope"])),
            str(row["type"]),
            str(row["definition"]),
            str(row["rule"]),
            str(row["possible_values_or_example"]),
        )
        for row in parameter_reference_rows(plan)
    ]
    observed_reference = [tuple(_value(reference, row, column) for column in range(1, 7)) for row in range(5, 5 + len(expected_reference))]
    if observed_reference != expected_reference:
        errors.append("Parameter Reference does not exactly project the canonical parameter semantics.")

    event_sheets = {_value(sheet, 3, 2): sheet for sheet in workbook.worksheets if _value(sheet, 3, 1) == label(plan, "event") and _value(sheet, 3, 2)}
    if set(event_sheets) != {str(event.get("event_name")) for event in events}:
        errors.append("Visible event tabs do not match the canonical event set.")
        return errors
    for event in events:
        event_name = str(event["event_name"])
        sheet = event_sheets[event_name]
        expected_fields = [
            event_name,
            classification_label(plan, str(event.get("classification", ""))),
            " | ".join(event_journey_names(plan, event)),
            str(event.get("definition", "")),
            str(event.get("trigger", "")),
            location_text(event),
            str(event.get("notes", "")),
        ]
        if [_value(sheet, row, 2) for row in range(3, 10)] != [value.strip() for value in expected_fields]:
            errors.append(f'Event-tab header fields differ for "{event_name}".')
        parameters = [parameter for parameter in event.get("parameters", []) if isinstance(parameter, dict)]
        expected_parameters = [
            (
                str(parameter.get("name", "")),
                scope_label(plan, str(parameter.get("scope", ""))),
                str(parameter.get("type", "")),
                requirement_label(plan, str(parameter.get("requirement", ""))),
                str(parameter.get("definition", "")),
                value_rule_text(parameter, plan),
                possible_values_or_example(parameter),
            )
            for parameter in parameters
        ]
        observed_parameters = [tuple(_value(sheet, row, column) for column in range(1, 8)) for row in range(12, 12 + len(parameters))]
        if observed_parameters != expected_parameters:
            errors.append(f'Parameter rows differ for "{event_name}".')
        code_row = 14 + max(1, len(parameters))
        if _value(sheet, code_row, 1) != datalayer_code(event, plan.get("data_layer_convention")).strip():
            errors.append(f'dataLayer code differs for "{event_name}".')
    return errors


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Validate the rendered human workbook against canonical semantics.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--plan", type=Path, required=True)
    parser.add_argument("--mapping", type=Path)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    try:
        errors = validate_workbook(
            args.workbook,
            load_json(args.plan),
            load_json(args.mapping) if args.mapping else None,
        )
    except (OSError, ValueError, json.JSONDecodeError) as error:
        print(str(error), file=sys.stderr)
        return 2
    if errors:
        print("\n".join(f"ERROR {error}" for error in errors), file=sys.stderr)
        return 1
    print(args.workbook)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
