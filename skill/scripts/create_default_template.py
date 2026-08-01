from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NAVY = "20343E"
TEAL = "2F6F7E"
TEAL_LIGHT = "E8F4F2"
PALE = "F6F9FA"
WHITE = "FFFFFF"
TEXT = "263238"
MUTED = "61717A"
GRID = "D8E1E5"
ACCENT = "DCEFEA"

THIN = Side(style="thin", color=GRID)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")
FORMULA_PREFIXES = ("=", "+", "-", "@")
SKILL_ROOT = Path(__file__).resolve().parents[1]


def release_version() -> str:
    release = json.loads(
        (SKILL_ROOT / "release.json").read_text(encoding="utf-8-sig")
    )
    return str(release["version"])


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build the approved default GA4 tracking-plan XLSX asset.")
    parser.add_argument("--output", "-o", type=Path, required=True)
    return parser.parse_args()


def set_cell_value(cell, value):
    """Write user-controlled text as text, never as an Excel formula."""
    cell.value = value
    if isinstance(value, str) and value.startswith(FORMULA_PREFIXES):
        cell.data_type = "s"
    return cell


def fill(cell, color: str) -> None:
    cell.fill = PatternFill("solid", fgColor=color)


def apply_title(ws, title: str, subtitle: str, columns: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columns)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=columns)
    set_cell_value(ws.cell(1, 1), title)
    set_cell_value(ws.cell(2, 1), subtitle)
    fill(ws.cell(1, 1), TEAL)
    ws.cell(1, 1).font = Font(color=WHITE, bold=True, size=18)
    ws.cell(1, 1).alignment = Alignment(vertical="center")
    ws.cell(2, 1).font = Font(color=TEAL, size=10)
    fill(ws.cell(2, 1), TEAL_LIGHT)
    ws.cell(2, 1).alignment = WRAP
    ws.row_dimensions[1].height = 34
    ws.row_dimensions[2].height = 30


def apply_header(ws, row: int, columns: int) -> None:
    for column in range(1, columns + 1):
        cell = ws.cell(row, column)
        fill(cell, NAVY)
        cell.font = Font(color=WHITE, bold=True, size=10)
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[row].height = 34


def apply_table_row(ws, row: int, columns: int) -> None:
    for column in range(1, columns + 1):
        cell = ws.cell(row, column)
        fill(cell, WHITE)
        cell.font = Font(color=TEXT, size=10)
        cell.alignment = WRAP
        cell.border = BORDER


def apply_label(cell) -> None:
    fill(cell, PALE)
    cell.font = Font(color=MUTED, bold=True, size=10)
    cell.alignment = WRAP
    cell.border = BORDER


def apply_value(cell) -> None:
    fill(cell, WHITE)
    cell.font = Font(color=TEXT, size=10)
    cell.alignment = WRAP
    cell.border = BORDER


def set_widths(ws, widths: list[int]) -> None:
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width


def build_guide(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Guide"
    apply_title(ws, "{{TITLE}}", "{{SUBTITLE}}", 7)
    for row in range(4, 13):
        apply_label(ws.cell(row, 1))
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        apply_value(ws.cell(row, 2))
        ws.row_dimensions[row].height = 28
    ws.merge_cells(start_row=14, start_column=1, end_row=14, end_column=7)
    ws.cell(14, 1, "{{JOURNEYS}}")
    fill(ws.cell(14, 1), ACCENT)
    ws.cell(14, 1).font = Font(color=TEAL, bold=True, size=12)
    ws.cell(14, 1).alignment = Alignment(vertical="center")
    ws.row_dimensions[14].height = 26
    for column in range(1, 5):
        ws.cell(15, column, "{{HEADER}}")
    apply_header(ws, 15, 4)
    apply_table_row(ws, 16, 4)
    set_widths(ws, [22, 32, 46, 46, 16, 16, 16])
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    ws.freeze_panes = "A16"


def build_event_matrix(wb: Workbook) -> None:
    ws = wb.create_sheet("Event Matrix")
    apply_title(ws, "{{EVENT_MATRIX}}", "{{EVENT_MATRIX_SUBTITLE}}", 2)
    headers = [
        "Event",
        "Definition",
    ]
    for column, header in enumerate(headers, 1):
        set_cell_value(ws.cell(4, column), header)
    apply_header(ws, 4, 2)
    apply_table_row(ws, 5, 2)
    set_widths(ws, [24, 96])
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 85
    ws.freeze_panes = "A5"


def build_parameter_reference(wb: Workbook) -> None:
    ws = wb.create_sheet("Parameter Reference")
    apply_title(ws, "{{PARAMETER_REFERENCE}}", "{{PARAMETER_REFERENCE_SUBTITLE}}", 7)
    headers = [
        "Variable",
        "Scope",
        "Type",
        "Definition",
        "Example",
        "Possible values / rule",
        "Concerned events",
    ]
    for column, header in enumerate(headers, 1):
        ws.cell(4, column, header)
    apply_header(ws, 4, 7)
    apply_table_row(ws, 5, 7)
    set_widths(ws, [22, 11, 11, 38, 22, 38, 28])
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 85
    ws.freeze_panes = "A5"


def build_event_template(wb: Workbook) -> None:
    ws = wb.create_sheet("__EVENT_TEMPLATE")
    apply_title(ws, '{{EVENT: "event_name"}}', "{{EVENT_SUBTITLE}}", 7)
    labels = [
        "Event",
        "Classification",
        "Journey",
        "Definition",
        "Trigger",
        "Pages / routes / components",
        "Notes",
    ]
    for offset, text in enumerate(labels, 3):
        set_cell_value(ws.cell(offset, 1), text)
        apply_label(ws.cell(offset, 1))
        ws.merge_cells(start_row=offset, start_column=2, end_row=offset, end_column=7)
        apply_value(ws.cell(offset, 2))
        ws.row_dimensions[offset].height = 32 if offset < 6 else 54
    headers = [
        "Variable",
        "Scope",
        "Type",
        "Requirement",
        "Definition",
        "Possible values / rule",
        "Example",
    ]
    for column, header in enumerate(headers, 1):
        set_cell_value(ws.cell(11, column), header)
    apply_header(ws, 11, 7)
    apply_table_row(ws, 12, 7)
    set_widths(ws, [20, 11, 11, 13, 42, 40, 24])
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 85
    ws.freeze_panes = "A12"
    ws.sheet_state = "hidden"


def build_template() -> Workbook:
    wb = Workbook()
    wb.properties.title = "GA4 tracking plan default template"
    wb.properties.subject = "Human-first analyst and developer tracking-plan contract"
    wb.properties.creator = "ga4-tracking-plan"
    wb.properties.description = (
        f"Default workbook asset version {release_version()}"
    )
    build_guide(wb)
    build_event_matrix(wb)
    build_parameter_reference(wb)
    build_event_template(wb)
    wb.active = 0
    return wb


def main() -> int:
    args = parse_args()
    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook = build_template()
    workbook.save(args.output)
    print(args.output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
