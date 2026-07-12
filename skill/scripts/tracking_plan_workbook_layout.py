from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, quote_sheetname

NAVY = "263238"
LIGHT_BLUE = "F6F9FC"
GREEN = "EAF3EA"
YELLOW = "FFF6D8"
GRAY = "F4F6F8"
WHITE = "FFFFFF"
RED = "FBEAEA"
DARK = "404040"
MUTED_TEXT = "6E7781"
HEADER_TEXT = "1F2933"
BLOCK_FILL = "EDF6EF"
INHERITED_FILL = "EAF4FB"
DEFAULT_FILL = "F0F7F2"
NOT_AVAILABLE_FILL = "FFF2CC"
NOT_APPLICABLE_FILL = "F5F6F7"
GRID = "E7ECF2"
GRID_DARK = "BBC8D6"
TEAL = "2F6F7E"
TEAL_LIGHT = "EAF6F4"
TEAL_SECTION = "DCEFEA"
OVERVIEW_HEADER = "F3F7F8"
OVERVIEW_ALT = "FAFCFD"

THIN = Side(style="thin", color=GRID)
MEDIUM = Side(style="medium", color=GRID_DARK)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BLOCK_BORDER = Border(top=MEDIUM, bottom=THIN, left=THIN, right=THIN)
OVERVIEW_LINE = Side(style="thin", color="DDE5EA")
OVERVIEW_BORDER = Border(bottom=OVERVIEW_LINE)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")

EVENT_SLOT_COUNT = 4
SCREENSHOT_STATUS_OPTIONS = "captured,shared_evidence,not_needed,blocked"
SCREENSHOT_ROW_HEIGHT = 216

OVERVIEW_SECTIONS = {"Document Summary", "Sheet Contents", "Version History"}
OVERVIEW_LABEL_COLUMNS = {1, 3, 5, 7}


def matrix_max_col() -> int:
    return 2 + EVENT_SLOT_COUNT


def matrix_value_columns() -> list[int]:
    return [3 + index for index in range(EVENT_SLOT_COUNT)]


def set_widths(ws, widths: list[int]) -> None:
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width


def style_cells(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = WRAP
            cell.border = BORDER
    ws.sheet_view.showGridLines = False


def _style_overview_base_row(ws, row: int, max_col: int) -> None:
    if row not in {1, 2}:
        ws.row_dimensions[row].height = 24
    for column in range(1, max_col + 1):
        cell = ws.cell(row, column)
        cell.alignment = WRAP
        cell.border = OVERVIEW_BORDER
        if cell.hyperlink:
            cell.font = Font(color=TEAL, bold=True, underline="single", size=10)
        elif row not in {1, 2}:
            cell.font = Font(color=DARK, size=10)


def _style_overview_section(ws, row: int, max_col: int) -> None:
    ws.row_dimensions[row].height = 26
    for column in range(1, max_col + 1):
        cell = ws.cell(row, column)
        cell.fill = PatternFill("solid", fgColor=TEAL_SECTION)
        cell.font = Font(color=TEAL, bold=True, size=12)
        cell.alignment = Alignment(vertical="center")
        cell.border = Border(top=Side(style="medium", color=TEAL), bottom=OVERVIEW_LINE)


def _style_overview_header(ws, row: int, max_col: int) -> None:
    ws.row_dimensions[row].height = 22
    for column in range(1, max_col + 1):
        cell = ws.cell(row, column)
        cell.fill = PatternFill("solid", fgColor=OVERVIEW_HEADER)
        cell.font = Font(color=HEADER_TEXT, bold=True, size=10)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = Border(top=OVERVIEW_LINE, bottom=OVERVIEW_LINE)


def _style_overview_data(ws, row: int, max_col: int) -> None:
    if row <= 2 or not any(ws.cell(row, column).value not in (None, "") for column in range(1, max_col + 1)):
        return
    fill = OVERVIEW_ALT if row % 2 == 0 else WHITE
    for column in range(1, max_col + 1):
        cell = ws.cell(row, column)
        cell.fill = PatternFill("solid", fgColor=fill)
        if column in OVERVIEW_LABEL_COLUMNS and cell.value not in (None, ""):
            cell.font = Font(color=MUTED_TEXT, bold=True, size=10)


def _style_overview_summary_rows(ws, max_col: int) -> None:
    for row in (4, 5):
        if row > ws.max_row:
            continue
        ws.row_dimensions[row].height = 30
        for column in range(1, max_col + 1):
            cell = ws.cell(row, column)
            cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE if column in OVERVIEW_LABEL_COLUMNS else WHITE)
            cell.border = BORDER
            if column not in OVERVIEW_LABEL_COLUMNS and cell.value not in (None, ""):
                cell.font = Font(color=HEADER_TEXT, bold=True, size=10)


def _style_overview_title(ws) -> None:
    ws.cell(1, 1).fill = PatternFill("solid", fgColor=TEAL)
    ws.cell(1, 1).font = Font(color=WHITE, bold=True, size=20)
    ws.cell(1, 1).alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 38
    ws.cell(2, 1).fill = PatternFill("solid", fgColor=TEAL_LIGHT)
    ws.cell(2, 1).font = Font(color=TEAL, size=11)
    ws.cell(2, 1).alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 30


def style_overview(ws, max_col: int) -> None:
    style_cells(ws)
    ws.freeze_panes = None
    ws.sheet_properties.tabColor = "95C8C1"
    ws.sheet_view.zoomScale = 90
    ws.sheet_view.zoomScaleNormal = 90
    for row in range(1, ws.max_row + 1):
        label = ws.cell(row, 1).value
        _style_overview_base_row(ws, row, max_col)
        if label in OVERVIEW_SECTIONS:
            _style_overview_section(ws, row, max_col)
        elif label == "#" or (label == "Version" and ws.cell(row, 2).value == "Date"):
            _style_overview_header(ws, row, max_col)
        else:
            _style_overview_data(ws, row, max_col)
    _style_overview_summary_rows(ws, max_col)
    _style_overview_title(ws)


def title(ws, text: str, subtitle: str, max_col: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    ws.cell(1, 1, text)
    ws.cell(2, 1, subtitle)
    ws.cell(1, 1).fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(1, 1).font = Font(color=WHITE, bold=True, size=15)
    ws.cell(2, 1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws.cell(2, 1).font = Font(color=DARK)
    ws.cell(1, 1).alignment = Alignment(vertical="center")
    ws.cell(2, 1).alignment = WRAP
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 34


def section(ws, row: int, label: str, max_col: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row, 1, label)
    cell.fill = PatternFill("solid", fgColor=BLOCK_FILL)
    cell.font = Font(color=HEADER_TEXT, bold=True)
    cell.alignment = Alignment(vertical="center")
    for column in range(1, max_col + 1):
        ws.cell(row, column).border = BLOCK_BORDER


def header(ws, row: int, max_col: int, fill: str = NAVY) -> None:
    for column in range(1, max_col + 1):
        cell = ws.cell(row, column)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(color=WHITE if fill == NAVY else "000000", bold=True)
        cell.alignment = CENTER
        cell.border = BORDER


def set_internal_link(cell, sheet_name: str) -> None:
    cell.hyperlink = f"#{quote_sheetname(sheet_name)}!A1"
    cell.font = Font(color=TEAL, bold=True, underline="single")


def style_matrix_value_cell(cell, availability: str) -> None:
    if cell.value in (None, ""):
        return
    styles = {
        "not_applicable": (NOT_APPLICABLE_FILL, Font(color=MUTED_TEXT, italic=True)),
        "not_available": (NOT_AVAILABLE_FILL, Font(color=DARK)),
        "event_level_used": (INHERITED_FILL, Font(color=DARK, italic=True)),
        "send_default_quantity": (DEFAULT_FILL, Font(color=DARK)),
    }
    if availability not in styles:
        return
    fill, font = styles[availability]
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = font


def _style_matrix_block_row(ws, row: int, max_col: int) -> None:
    for column in range(1, max_col + 1):
        cell = ws.cell(row, column)
        cell.fill = PatternFill("solid", fgColor=BLOCK_FILL)
        cell.border = BLOCK_BORDER
        cell.font = Font(color=HEADER_TEXT, bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = 24


def _style_matrix_data_row(ws, row: int, max_col: int, value_columns: list[int]) -> None:
    parameter = str(ws.cell(row, 1).value or "")
    for column in range(1, max_col + 1):
        ws.cell(row, column).fill = PatternFill("solid", fgColor=WHITE)
    for column in value_columns:
        cell = ws.cell(row, column)
        value = str(cell.value or "")
        availability = ""
        if value in {"not_applicable", "not_available"}:
            availability = value
        elif value.startswith("event-level "):
            availability = "event_level_used"
        elif parameter == "items[].quantity" and value == "1":
            availability = "send_default_quantity"
        style_matrix_value_cell(cell, availability)


def style_event_matrix_rows(ws) -> None:
    value_columns = matrix_value_columns()
    max_col = matrix_max_col()
    for row in range(6, ws.max_row + 1):
        if str(ws.cell(row, 1).value or "").startswith("J-"):
            _style_matrix_block_row(ws, row, max_col)
        else:
            _style_matrix_data_row(ws, row, max_col, value_columns)


def apply_workbook_settings(workbook: Workbook) -> None:
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and len(cell.value) > 80:
                    worksheet.row_dimensions[cell.row].height = max(worksheet.row_dimensions[cell.row].height or 15, 42)
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
    workbook.active = 0
