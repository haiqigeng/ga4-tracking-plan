from __future__ import annotations

import argparse
import copy
import json
import re
import sys
import unicodedata
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from tracking_plan_model import (
    BASE_SHEETS,
    CLASSIFICATION_LABELS,
    LABELS,
    REQUIREMENT_LABELS,
    SCOPE_LABELS,
    classification_label,
    datalayer_code,
    event_journey_names,
    flatten_push_paths,
    location_text,
    path_exists,
    slugify,
    workbook_projection,
)

MODEL_SHEET = "__tracking_plan_model"
MODEL_MARKER = "ga4-tracking-plan/model"
PROJECTION_SHEET = "__tracking_plan_projection"
PROJECTION_MARKER = "ga4-tracking-plan/projection"
ROOT = Path(__file__).resolve().parents[1]
CATALOG_PATH = ROOT / "references" / "library-ga4-recommended-events.json"
OFFICIAL_EVENTS_URL = "https://developers.google.com/analytics/devguides/collection/ga4/reference/events"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import a generated GA4 tracking-plan workbook into its canonical JSON model.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--output", "-o", type=Path, required=True)
    parser.add_argument(
        "--allow-visible-recovery",
        action="store_true",
        help="Recover a best-effort model from visible event tabs when no embedded model exists.",
    )
    parser.add_argument(
        "--reconcile-visible-edits",
        action="store_true",
        help=(
            "Merge supported edits from the visible event tabs into the embedded canonical "
            "model. Structural edits are rejected and must be made in canonical JSON."
        ),
    )
    return parser.parse_args()


def _normalized(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    return "".join(char for char in text if not unicodedata.combining(char)).strip().casefold()


def _reverse_labels(table: dict[str, dict[str, str]]) -> dict[str, str]:
    result: dict[str, str] = {}
    for translations in table.values():
        for key, value in translations.items():
            result[_normalized(value)] = key
    return result


CLASSIFICATION_BY_LABEL = _reverse_labels(CLASSIFICATION_LABELS)
SCOPE_BY_LABEL = _reverse_labels(SCOPE_LABELS)
REQUIREMENT_BY_LABEL = _reverse_labels(REQUIREMENT_LABELS)


def read_embedded_model(workbook) -> dict[str, Any] | None:
    if MODEL_SHEET not in workbook.sheetnames:
        return None
    sheet = workbook[MODEL_SHEET]
    if sheet.cell(1, 1).value != MODEL_MARKER:
        return None
    chunks: list[str] = []
    row = 2
    while sheet.cell(row, 1).value is not None:
        chunks.append(str(sheet.cell(row, 1).value))
        row += 1
    if not chunks:
        raise ValueError("The embedded tracking-plan model is empty.")
    value = json.loads("".join(chunks))
    if not isinstance(value, dict):
        raise ValueError("The embedded tracking-plan model is not a JSON object.")
    return value


def read_embedded_projection(workbook) -> dict[str, Any] | None:
    if PROJECTION_SHEET not in workbook.sheetnames:
        return None
    sheet = workbook[PROJECTION_SHEET]
    if sheet.cell(1, 1).value != PROJECTION_MARKER:
        return None
    chunks: list[str] = []
    row = 2
    while sheet.cell(row, 1).value is not None:
        chunks.append(str(sheet.cell(row, 1).value))
        row += 1
    if not chunks:
        raise ValueError("The embedded visible-workbook projection is empty.")
    value = json.loads("".join(chunks))
    if not isinstance(value, dict):
        raise ValueError("The embedded visible-workbook projection is not a JSON object.")
    return value


def _cell_value(sheet, row: int, column: int) -> str:
    value = sheet.cell(row, column).value
    return "" if value is None else str(value).strip()


def _language_from_workbook(workbook) -> str:
    if "Valeurs des variables" in workbook.sheetnames:
        return "fr"
    if "Parameter Reference" in workbook.sheetnames:
        return "en"
    observed = {_normalized(cell.value) for sheet in workbook.worksheets for row in sheet.iter_rows(max_row=20) for cell in row if cell.value not in (None, "")}
    scores = {"en": 0, "fr": 0}
    discriminating_keys = {
        "scope",
        "target_state",
        "analyst_entry",
        "developer_entry",
        "datalayer_convention",
        "journeys",
        "business_goal",
        "trigger",
        "locations",
        "requirement",
        "values",
        "rule",
        "possible_values_or_examples",
        "concerned_events",
    }
    for key in discriminating_keys:
        english = _normalized(LABELS["en"][key])
        french = _normalized(LABELS["fr"][key])
        if english == french:
            continue
        scores["en"] += int(english in observed)
        scores["fr"] += int(french in observed)
    return "fr" if scores["fr"] > scores["en"] else "en"


def _parse_json_example(value: Any, expected_type: str) -> Any:
    if value is None:
        return None
    if expected_type in {"array", "object", "boolean", "integer", "number"}:
        try:
            return json.loads(str(value))
        except (json.JSONDecodeError, TypeError):
            pass
    return value


def _parse_possible_values(value: Any, expected_type: str) -> list[Any]:
    text = "" if value is None else str(value).strip()
    if not text:
        return []
    return [_parse_json_example(part.strip(), expected_type) for part in text.split(" | ") if part.strip()]


def _extract_push(
    code: str,
    event_name: str,
    *,
    context: bool = False,
) -> dict[str, Any]:
    candidates = re.findall(
        r"window\.dataLayer\.push\(\s*(\{.*?\})\s*\);",
        code,
        flags=re.DOTALL,
    )
    for candidate in reversed(candidates):
        try:
            value = json.loads(candidate)
        except json.JSONDecodeError:
            continue
        if isinstance(value, dict) and (value.get("event") == event_name or "event" not in value):
            return value
    return {} if context else {"event": event_name}


def _extract_data_layer(code: str, event_name: str, *, context: bool) -> dict[str, Any]:
    """Parse the generated JSON-only dataLayer block without executing JavaScript."""
    candidates = re.findall(
        r"window\.dataLayer\.push\(\s*(\{.*?\})\s*\);",
        code,
        flags=re.DOTALL,
    )
    values: list[dict[str, Any]] = []
    for candidate in candidates:
        try:
            value = json.loads(candidate)
        except json.JSONDecodeError:
            continue
        if isinstance(value, dict):
            values.append(value)
    selected_index: int | None = None
    for index in range(len(values) - 1, -1, -1):
        value = values[index]
        if value.get("event") == event_name or (context and "event" not in value):
            selected_index = index
            break
    push = values[selected_index] if selected_index is not None else ({} if context else {"event": event_name})
    clear = [str(key) for index, value in enumerate(values) if index != selected_index and len(value) == 1 for key, item in value.items() if item is None]
    result: dict[str, Any] = {"push": push}
    if clear:
        result["clear"] = list(dict.fromkeys(clear))
    return result


def _normalized_projection(value: dict[str, Any]) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def _projection_differences(
    expected: dict[str, Any],
    observed: dict[str, Any],
    *,
    limit: int = 12,
) -> list[str]:
    differences: list[str] = []
    expected_sheets = {str(item.get("title")): item for item in expected.get("sheets", []) if isinstance(item, dict)}
    observed_sheets = {str(item.get("title")): item for item in observed.get("sheets", []) if isinstance(item, dict)}
    for title in sorted(set(expected_sheets) | set(observed_sheets)):
        if len(differences) >= limit:
            break
        if title not in expected_sheets:
            differences.append(f"added sheet {title!r}")
            continue
        if title not in observed_sheets:
            differences.append(f"removed sheet {title!r}")
            continue
        before = expected_sheets[title]
        after = observed_sheets[title]
        if before.get("state") != after.get("state"):
            differences.append(f"sheet state changed for {title!r}")
        if before.get("merged_cells") != after.get("merged_cells"):
            differences.append(f"merged cells changed for {title!r}")
        before_cells = {str(item[0]): item[1:] for item in before.get("cells", []) if isinstance(item, list) and item}
        after_cells = {str(item[0]): item[1:] for item in after.get("cells", []) if isinstance(item, list) and item}
        for coordinate in sorted(set(before_cells) | set(after_cells)):
            if before_cells.get(coordinate) != after_cells.get(coordinate):
                differences.append(f"{title}!{coordinate}")
                if len(differences) >= limit:
                    break
    return differences


def _projection_changes(expected: dict[str, Any], observed: dict[str, Any]) -> list[dict[str, str]]:
    changes: list[dict[str, str]] = []
    expected_sheets = {str(item.get("title")): item for item in expected.get("sheets", []) if isinstance(item, dict)}
    observed_sheets = {str(item.get("title")): item for item in observed.get("sheets", []) if isinstance(item, dict)}
    for title in sorted(set(expected_sheets) | set(observed_sheets)):
        if title not in expected_sheets:
            changes.append({"kind": "sheet_added", "sheet": title, "coordinate": ""})
            continue
        if title not in observed_sheets:
            changes.append({"kind": "sheet_removed", "sheet": title, "coordinate": ""})
            continue
        before = expected_sheets[title]
        after = observed_sheets[title]
        if before.get("state") != after.get("state"):
            changes.append({"kind": "sheet_state", "sheet": title, "coordinate": ""})
        if before.get("merged_cells") != after.get("merged_cells"):
            changes.append({"kind": "merged_cells", "sheet": title, "coordinate": ""})
        before_cells = {str(item[0]): item[1:] for item in before.get("cells", []) if isinstance(item, list) and item}
        after_cells = {str(item[0]): item[1:] for item in after.get("cells", []) if isinstance(item, list) and item}
        for coordinate in sorted(set(before_cells) | set(after_cells)):
            if before_cells.get(coordinate) != after_cells.get(coordinate):
                changes.append({"kind": "cell", "sheet": title, "coordinate": coordinate})
    return changes


def _load_catalog() -> dict[str, dict[str, Any]]:
    records = json.loads(CATALOG_PATH.read_text(encoding="utf-8-sig"))
    return {str(record.get("event")): record for record in records if isinstance(record, dict) and record.get("event")}


def _catalog_parameter(
    record: dict[str, Any] | None,
    name: str,
    scope: str,
) -> dict[str, Any] | None:
    if not record:
        return None
    for parameter in record.get("parameters", []):
        if isinstance(parameter, dict) and str(parameter.get("name")) == name and str(parameter.get("scope", "event")) == scope:
            return parameter
    return None


def _official_source(
    event_name: str,
    section: str,
    wording_origin: str,
    official_text: str,
) -> dict[str, str]:
    return {
        "url": f"{OFFICIAL_EVENTS_URL}#{event_name}",
        "section": section,
        "wording_origin": wording_origin,
        "official_text": official_text,
        "checked_date": date.today().isoformat(),
    }


def _wording_origin(visible: Any, official: Any) -> str:
    return "exact" if " ".join(str(visible or "").split()).casefold() == " ".join(str(official or "").split()).casefold() else "faithful_translation"


def _canonical_type(value: Any) -> str:
    text = str(value or "").strip().casefold()
    if text.startswith("array"):
        return "array"
    if text.startswith("string"):
        return "string"
    if text in {"float", "double"}:
        return "number"
    return (
        text
        if text
        in {
            "string",
            "integer",
            "number",
            "boolean",
            "array",
            "object",
        }
        else "string"
    )


def _split_value_rule(value: str, language: str) -> tuple[str, list[str] | None]:
    marker = "Valeurs possibles" if language == "fr" else "Allowed values"
    lines = value.splitlines()
    allowed: list[str] | None = None
    retained: list[str] = []
    for line in lines:
        match = re.match(rf"^\s*{re.escape(marker)}\s*:\s*(.+)$", line, re.I)
        if match:
            allowed = [item.strip() for item in match.group(1).split("|") if item.strip()]
        else:
            retained.append(line)
    rule = "\n".join(retained).strip()
    return rule, allowed


def _infer_path(
    name: str,
    scope: str,
    push: dict[str, Any],
    event_classification: str,
) -> str:
    matches = sorted(path for path in flatten_push_paths(push) if path.rsplit(".", 1)[-1].replace("[]", "") == name)
    if matches:
        if scope == "item":
            item_matches = [path for path in matches if "[]" in path]
            if item_matches:
                return item_matches[0]
        if scope == "user":
            user_matches = [path for path in matches if path.startswith("user.")]
            if user_matches:
                return user_matches[0]
        return matches[0]
    if scope == "item":
        return f"ecommerce.items[].{name}"
    if scope == "user":
        return f"user.{name}"
    if event_classification == "official_ecommerce":
        return f"ecommerce.{name}"
    if event_classification == "context":
        return f"page.{name}"
    return f"event_data.{name}"


def _set_path_value(root: dict[str, Any], path: str, value: Any) -> None:
    current: dict[str, Any] = root
    parts = [part for part in path.split(".") if part]
    for index, part in enumerate(parts):
        is_array = part.endswith("[]")
        key = part[:-2] if is_array else part
        final = index == len(parts) - 1
        if final:
            if key not in current:
                current[key] = value
            return
        if is_array:
            children = current.setdefault(key, [{}])
            if not isinstance(children, list) or not children:
                current[key] = [{}]
                children = current[key]
            if not isinstance(children[0], dict):
                children[0] = {}
            current = children[0]
        else:
            child = current.setdefault(key, {})
            if not isinstance(child, dict):
                current[key] = {}
            current = current[key]


def _event_tabs(workbook, language: str) -> list[Any]:
    event_label = _normalized(LABELS[language]["event"])
    result: list[Any] = []
    for sheet in workbook.worksheets:
        if sheet.title in BASE_SHEETS or sheet.title in {"Valeurs des variables", "Journal des modifications"}:
            continue
        if sheet.sheet_state != "visible":
            continue
        if _normalized(sheet.cell(3, 1).value) == event_label and _cell_value(sheet, 3, 2):
            result.append(sheet)
    return result


def _recover_visible_model(workbook, source: Path) -> dict[str, Any]:
    language = _language_from_workbook(workbook)
    sheets = _event_tabs(workbook, language)
    if not sheets:
        raise ValueError(
            "No embedded model or recognizable event tabs were found. Use inspect_tracking_plan_template.py and analyst-led semantic mapping for this workbook."
        )

    catalog = _load_catalog()
    journey_by_name: dict[str, str] = {}
    events: list[dict[str, Any]] = []
    for sheet in sheets:
        event_name = _cell_value(sheet, 3, 2)
        classification = CLASSIFICATION_BY_LABEL.get(
            _normalized(_cell_value(sheet, 4, 2)),
            "custom",
        )
        catalog_record = catalog.get(event_name)
        if catalog_record and classification != "context":
            classification = "official_ecommerce" if str(catalog_record.get("group", "")).casefold() == "online sales" else "official"

        journey_names = [item.strip() for item in _cell_value(sheet, 5, 2).split("|") if item.strip()] or ["Imported journey"]
        for journey_name in journey_names:
            journey_by_name.setdefault(journey_name, slugify(journey_name))

        locations = [{"url_pattern": line.strip()} for line in _cell_value(sheet, 8, 2).splitlines() if line.strip()] or [
            {"state": "Imported workbook; location requires review."}
        ]

        code = ""
        for candidate_row in range(12, sheet.max_row + 1):
            candidate = sheet.cell(candidate_row, 1).value
            if isinstance(candidate, str) and "window.dataLayer" in candidate:
                code = candidate
                break
        push = _extract_push(
            code,
            event_name,
            context=classification == "context",
        )

        headers = {_normalized(sheet.cell(11, column).value): column for column in range(1, sheet.max_column + 1) if sheet.cell(11, column).value}
        parameter_column = headers.get(
            _normalized(LABELS[language]["variable"]),
            1,
        )
        scope_column = headers.get(
            _normalized(LABELS[language]["scope_label"]),
            2,
        )
        type_column = headers.get(_normalized(LABELS[language]["type"]), 3)
        requirement_column = headers.get(
            _normalized(LABELS[language]["requirement"]),
            4,
        )
        condition_column = headers.get(_normalized(LABELS[language]["condition"]))
        definition_column = headers.get(
            _normalized(LABELS[language]["definition"]),
            5,
        )
        values_column = headers.get(
            _normalized(LABELS[language]["rule"]),
            headers.get(_normalized(LABELS[language]["values"]), 6),
        )
        possible_column = headers.get(
            _normalized(LABELS[language]["possible_values_or_examples"]),
            headers.get(_normalized(LABELS[language]["example"]), 7),
        )
        example_column = headers.get(
            _normalized(LABELS[language]["example"]),
            possible_column,
        )
        legacy_values_column = headers.get(
            _normalized(LABELS[language]["values"]),
            6,
        )
        source_column = headers.get(_normalized(LABELS[language]["source_path"]))

        parameters: list[dict[str, Any]] = []
        row = 12
        while row <= sheet.max_row:
            name = _cell_value(sheet, row, parameter_column)
            if not name or "datalayer" in _normalized(name):
                break
            scope = SCOPE_BY_LABEL.get(
                _normalized(_cell_value(sheet, row, scope_column)),
                "event",
            )
            requirement = REQUIREMENT_BY_LABEL.get(
                _normalized(_cell_value(sheet, row, requirement_column)),
                "optional",
            )
            visible_definition = _cell_value(sheet, row, definition_column)
            visible_values = _cell_value(sheet, row, values_column)
            if values_column == legacy_values_column and values_column != headers.get(_normalized(LABELS[language]["rule"])):
                value_rule, allowed_values = _split_value_rule(
                    visible_values,
                    language,
                )
            else:
                value_rule = visible_values
                allowed_values = None
            source_lines = _cell_value(sheet, row, source_column).splitlines() if source_column else []
            path = source_lines[0].strip() if source_lines and source_lines[0].strip() else _infer_path(name, scope, push, classification)
            official_parameter = _catalog_parameter(
                catalog_record,
                name,
                scope,
            )
            parameter_classification = (
                "implementation"
                if name == "user_id" or (classification == "context" and scope == "implementation")
                else ("official" if official_parameter else "custom")
            )
            parameter_type = _canonical_type(_cell_value(sheet, row, type_column) or (official_parameter or {}).get("type") or "string")
            example = _parse_json_example(
                sheet.cell(row, example_column).value,
                parameter_type,
            )
            possible_values = _parse_possible_values(
                sheet.cell(row, possible_column).value,
                parameter_type,
            )
            if len(possible_values) > 1:
                allowed_values = possible_values
                example = possible_values[0]
            if name == "user_id":
                destination = "ga4_user_id"
            elif scope == "item":
                destination = "ga4_item_parameter"
            elif scope == "user":
                destination = "ga4_user_property"
            elif parameter_classification == "implementation":
                destination = "implementation_only"
            else:
                destination = "ga4_event_parameter"

            parameter: dict[str, Any] = {
                "name": name,
                "data_layer_path": path,
                "classification": parameter_classification,
                "scope": scope,
                "type": parameter_type,
                "requirement": requirement,
                "definition": visible_definition or f"Identifies the recovered {name} value used by {event_name}.",
                "value_rule": value_rule or (f"Retain the normalized value represented by {name} in the source workbook."),
                "example": example,
                "source": "\n".join(source_lines[1:]).strip(),
                "value_mode": (
                    "structured"
                    if parameter_type in {"array", "object"}
                    else "numeric"
                    if parameter_type in {"integer", "number"}
                    else "boolean"
                    if parameter_type == "boolean"
                    else "controlled_semantic"
                    if allowed_values
                    else "authoritative_raw"
                ),
                "value_evidence_refs": [slugify(f"{event_name}_{name}_{scope}_values", "value_domain")],
                "destination": destination,
            }
            if parameter["value_mode"] == "controlled_semantic":
                parameter["value_language"] = language
            if allowed_values:
                parameter["allowed_values"] = allowed_values
            condition = _cell_value(sheet, row, condition_column) if condition_column else ""
            if requirement == "conditional":
                parameter["condition"] = condition or (
                    f"Send {name} only when its source value is available for {event_name}; review this recovered condition before delivery."
                )
            if parameter_classification == "official" and official_parameter:
                parameter["official_source"] = _official_source(
                    event_name,
                    f"{event_name} parameters",
                    _wording_origin(
                        parameter["definition"],
                        official_parameter.get("description"),
                    ),
                    str(official_parameter.get("description", "")),
                )
            if name == "user_id":
                parameter["official_source"] = {
                    "url": ("https://developers.google.com/analytics/devguides/collection/ga4/user-id"),
                    "section": "Send user IDs",
                    "wording_origin": "faithful_translation",
                    "official_text": ("The user_id parameter is a configuration parameter, not a custom user property or standard event parameter."),
                    "checked_date": date.today().isoformat(),
                }
            if parameter_classification == "custom":
                parameter["custom_decision"] = {
                    "business_need": (f"Preserve the distinct {name} analysis dimension represented by the imported {event_name} specification."),
                    "official_candidate": (
                        f"The current official {event_name} parameter table was checked for a {name} equivalent."
                        if catalog_record
                        else ("No event-specific official parameter table is available for this custom event.")
                    ),
                    "why_not_fit": (f"No official {scope}-scope parameter named {name} is prescribed for {event_name}."),
                }
            parameters.append(parameter)
            if not path_exists(push, path):
                _set_path_value(push, path, example)
            row += 1

        visible_event_definition = _cell_value(sheet, 6, 2) or f"Identifies the recovered {event_name} measurement."
        event: dict[str, Any] = {
            "event_name": event_name,
            "classification": classification,
            "journey_ids": [journey_by_name[name] for name in journey_names],
            "definition": visible_event_definition,
            "trigger": _cell_value(sheet, 7, 2) or (f"Push {event_name} at the implementation moment represented by the recovered event tab."),
            "locations": locations,
            "parameters": parameters,
            "data_layer": {"push": push},
            "notes": ("Recovered from visible workbook content. Review inferred conditions, sources, destinations, and custom decisions before delivery."),
        }
        if classification != "context":
            opportunity_id = slugify(
                f"recovered_{event_name}_opportunity",
                "opportunity",
            )
            event["business_question"] = f"Which business decision is supported by the recovered {event_name} action?"
            event["measurement_opportunity_ids"] = [opportunity_id]
        if classification in {"official", "official_ecommerce"}:
            event["official_source"] = _official_source(
                event_name,
                event_name,
                _wording_origin(
                    visible_event_definition,
                    (catalog_record or {}).get("description"),
                ),
                str((catalog_record or {}).get("description") or visible_event_definition),
            )
        if classification == "custom":
            event["custom_decision"] = {
                "business_need": (f"Preserve the distinct business action represented by the imported {event_name} trigger."),
                "official_candidate": ("The current recommended-event catalog contains no event with this imported name."),
                "why_not_fit": (f"The source workbook defines {event_name} as a distinct trigger and does not identify an equivalent recommended event."),
            }
        events.append(event)

    journeys = [
        {
            "journey_id": journey_id,
            "name": name,
            "scope": "Recovered from event-tab associations.",
            "status": "partial",
            "business_goal": "Confirm the business goal during maintenance review.",
        }
        for name, journey_id in journey_by_name.items()
    ]
    return {
        "schema_version": "5.0.0",
        "document": {
            "title": source.stem,
            "version": "imported",
            "date": date.today().isoformat(),
            "language": language,
            "scope": "Recovered from a workbook without an embedded canonical model.",
            "target_state": "hybrid",
            "notes": "Best-effort visible-content recovery; review before regeneration.",
        },
        "data_layer_convention": {
            "name": "Recovered workbook convention",
            "origin": "existing",
            "event_key": "event",
            "wrappers": {
                "page": "page",
                "event": "event_data",
                "ecommerce": "ecommerce",
                "user": "user",
            },
        },
        "journeys": journeys,
        "events": events,
    }


def _standard_event_tabs_by_name(workbook, plan: dict[str, Any]) -> dict[str, Any]:
    language = str(plan.get("document", {}).get("language", "en")).casefold()
    language = "fr" if language.startswith("fr") else "en"
    return {_cell_value(sheet, 3, 2): sheet for sheet in _event_tabs(workbook, language) if _cell_value(sheet, 3, 2)}


def _reconcile_standard_visible_edits(
    workbook,
    embedded: dict[str, Any],
    baseline: dict[str, Any],
) -> dict[str, Any]:
    """Merge only explicitly supported event-tab edits into the canonical model."""
    plan = copy.deepcopy(embedded)
    tabs = _standard_event_tabs_by_name(workbook, plan)
    event_by_name = {str(event.get("event_name")): event for event in plan.get("events", []) if isinstance(event, dict)}
    if set(tabs) != set(event_by_name):
        raise ValueError(
            "Visible edits cannot be reconciled because event tabs were added, removed, or renamed. Make structural event changes in canonical JSON."
        )

    allowed_cells: set[tuple[str, str]] = set()
    parsed_rows: dict[str, tuple[list[tuple[int, dict[str, Any]]], int]] = {}
    for event_name, event in event_by_name.items():
        sheet = tabs[event_name]
        allowed_cells.update({(sheet.title, "B6"), (sheet.title, "B7"), (sheet.title, "B9")})
        code_row = next(
            (row for row in range(12, sheet.max_row + 1) if isinstance(sheet.cell(row, 1).value, str) and "window.dataLayer" in str(sheet.cell(row, 1).value)),
            0,
        )
        if not code_row:
            raise ValueError(f'Visible edits cannot be reconciled because event tab "{sheet.title}" has no recognizable JSON-only dataLayer block.')
        allowed_cells.add((sheet.title, f"A{code_row}"))
        rows: list[tuple[int, dict[str, Any]]] = []
        for row in range(12, code_row):
            name = _cell_value(sheet, row, 1)
            if not name or _normalized(name) == _normalized(
                LABELS["fr" if str(plan["document"]["language"]).casefold().startswith("fr") else "en"]["datalayer"]
            ):
                continue
            scope = SCOPE_BY_LABEL.get(_normalized(_cell_value(sheet, row, 2)))
            if scope is None:
                raise ValueError(f'Unknown visible scope for parameter "{name}" on tab "{sheet.title}".')
            rows.append((row, {"name": name, "scope": scope}))
            for column in "CDEFG":
                allowed_cells.add((sheet.title, f"{column}{row}"))
        parsed_rows[event_name] = (rows, code_row)

    unsupported = [
        change
        for change in _projection_changes(baseline, workbook_projection(workbook))
        if change.get("kind") != "cell" or (str(change.get("sheet")), str(change.get("coordinate"))) not in allowed_cells
    ]
    if unsupported:
        details = ", ".join(
            (f"{item['sheet']}!{item['coordinate']}" if item.get("coordinate") else f"{item['kind']}:{item['sheet']}") for item in unsupported[:12]
        )
        raise ValueError(
            "Visible edits include unsupported structural or duplicated-display changes "
            f"({details}). Edit canonical JSON, or limit workbook maintenance to the "
            "editable definition, trigger, notes, parameter C:G, and dataLayer-code cells "
            "on event tabs."
        )

    for event_name, event in event_by_name.items():
        sheet = tabs[event_name]
        expected_structure = {
            "classification": classification_label(plan, str(event.get("classification", ""))),
            "journey": " | ".join(event_journey_names(plan, event)),
            "locations": location_text(event),
        }
        observed_structure = {
            "classification": _cell_value(sheet, 4, 2),
            "journey": _cell_value(sheet, 5, 2),
            "locations": _cell_value(sheet, 8, 2),
        }
        if observed_structure != expected_structure:
            raise ValueError(
                f'Structural semantics changed on event tab "{sheet.title}". '
                "Classification, journey membership, and locations must be edited in canonical JSON."
            )
        event["definition"] = _cell_value(sheet, 6, 2)
        event["trigger"] = _cell_value(sheet, 7, 2)
        notes = _cell_value(sheet, 9, 2)
        if notes:
            event["notes"] = notes
        else:
            event.pop("notes", None)

        rows, code_row = parsed_rows[event_name]
        visible_keys = [(item["name"], item["scope"]) for _, item in rows]
        parameter_by_key = {
            (str(parameter.get("name")), str(parameter.get("scope"))): parameter for parameter in event.get("parameters", []) if isinstance(parameter, dict)
        }
        if set(visible_keys) != set(parameter_by_key) or len(visible_keys) != len(parameter_by_key):
            raise ValueError(
                f'Parameters were added, removed, duplicated, or renamed on tab "{sheet.title}". Make structural parameter changes in canonical JSON.'
            )
        for row, visible in rows:
            parameter = parameter_by_key[(visible["name"], visible["scope"])]
            parameter_type = _canonical_type(_cell_value(sheet, row, 3))
            requirement = REQUIREMENT_BY_LABEL.get(_normalized(_cell_value(sheet, row, 4)))
            if requirement is None:
                raise ValueError(f'Unknown requirement for parameter "{visible["name"]}" on tab "{sheet.title}".')
            parameter["type"] = parameter_type
            parameter["requirement"] = requirement
            parameter["definition"] = _cell_value(sheet, row, 5)
            parameter["value_rule"] = _cell_value(sheet, row, 6)
            if parameter.get("allowed_values"):
                allowed_values = _parse_possible_values(
                    sheet.cell(row, 7).value,
                    parameter_type,
                )
                parameter["allowed_values"] = allowed_values
                if parameter.get("example") not in allowed_values and allowed_values:
                    parameter["example"] = allowed_values[0]
            else:
                parameter.pop("allowed_values", None)
                parameter["example"] = _parse_json_example(
                    sheet.cell(row, 7).value,
                    parameter_type,
                )

        code = _cell_value(sheet, code_row, 1)
        if code != datalayer_code(event):
            parsed = _extract_data_layer(
                code,
                event_name,
                context=str(event.get("classification")) == "context",
            )
            old_data_layer = event.get("data_layer", {})
            if set(parsed.get("clear", [])) != set(old_data_layer.get("clear", [])):
                raise ValueError(f'dataLayer clear keys changed on tab "{sheet.title}". Change wrapper structure in canonical JSON.')
            if flatten_push_paths(parsed.get("push", {})) != flatten_push_paths(old_data_layer.get("push", {})):
                raise ValueError(f'dataLayer keys changed on tab "{sheet.title}". Change parameter paths and wrapper structure in canonical JSON.')
            event["data_layer"] = parsed

    from validate_tracking_plan import render_text, validate_plan

    issues = validate_plan(plan)
    if issues:
        raise ValueError("Visible edits do not produce a delivery-valid canonical model:\n" + render_text(issues))
    return plan


def import_workbook(
    path: Path,
    allow_visible_recovery: bool = False,
    reconcile_visible_edits: bool = False,
) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=False, read_only=False)
    embedded = read_embedded_model(workbook)
    if embedded is not None:
        baseline = read_embedded_projection(workbook)
        if baseline is not None and _normalized_projection(baseline) != _normalized_projection(workbook_projection(workbook)):
            if reconcile_visible_edits:
                return _reconcile_standard_visible_edits(workbook, embedded, baseline)
            details = ", ".join(_projection_differences(baseline, workbook_projection(workbook)))
            raise ValueError(
                "The visible workbook has changed since the canonical model was embedded"
                + (f" ({details})" if details else "")
                + ". Re-run with --reconcile-visible-edits for supported event-tab edits; "
                "structural changes must be made in canonical JSON."
            )
        return embedded
    if not allow_visible_recovery:
        raise ValueError(
            "This workbook has no embedded canonical model. Re-run with --allow-visible-recovery "
            "for a best-effort import, then review every recovered semantic."
        )
    return _recover_visible_model(workbook, path)


def main() -> int:
    args = parse_args()
    try:
        plan = import_workbook(
            args.workbook,
            args.allow_visible_recovery,
            args.reconcile_visible_edits,
        )
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(
            json.dumps(plan, indent=2, ensure_ascii=False) + "\n",
            encoding="utf-8",
        )
    except (OSError, ValueError, json.JSONDecodeError) as error:
        print(str(error), file=sys.stderr)
        return 2
    print(args.output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
