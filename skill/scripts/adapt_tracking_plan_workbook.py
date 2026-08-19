from __future__ import annotations

import argparse
import copy
import json
import re
import sys
from collections import OrderedDict
from pathlib import Path
from typing import Any, Callable

from create_default_template import set_cell_value
from generate_tracking_plan_workbook import embed_model
from inspect_tracking_plan_template import normalize_template_mapping, sha256, validate_template_mapping
from openpyxl import load_workbook
from openpyxl.formatting.formatting import ConditionalFormatting
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter, quote_sheetname, range_boundaries
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string
from openpyxl.worksheet.cell_range import MultiCellRange
from template_fidelity import (
    add_package_fidelity,
    authorize_added_sheet,
    authorized_template_changes,
    compare_template_fidelity,
    workbook_fidelity_snapshot,
)
from template_preflight import ensure_writer_supported
from tracking_plan_model import (
    classification_label,
    combined_value_rule_text,
    compact_value,
    datalayer_code,
    event_journey_names,
    load_json,
    location_text,
    parameter_reference_rows,
    possible_values_or_example,
    requirement_label,
    safe_sheet_title,
    scope_label,
    value_rule_text,
)
from validate_tracking_plan import render_text, validate_plan


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Semantically adapt a validated GA4 tracking plan into a supplied workbook.")
    parser.add_argument("plan", type=Path)
    parser.add_argument("template", type=Path)
    parser.add_argument("--mapping", type=Path, required=True)
    parser.add_argument("--output", "-o", type=Path, required=True)
    parser.add_argument(
        "--fidelity-report",
        type=Path,
        help="Optional JSON output for the mandatory supplied-template fidelity gate.",
    )
    return parser.parse_args()


def _copy_row_style(
    sheet,
    source_row: int,
    target_row: int,
    columns: set[int],
    *,
    copy_formulas: bool = False,
) -> None:
    for column in columns:
        source = sheet.cell(source_row, column)
        target = sheet.cell(target_row, column)
        if source.has_style:
            target._style = copy.copy(source._style)
        if source.comment:
            target.comment = copy.copy(source.comment)
        target.number_format = source.number_format
        target.protection = copy.copy(source.protection)
        target.alignment = copy.copy(source.alignment)
        if copy_formulas and isinstance(source.value, str) and source.value.startswith("="):
            try:
                target.value = Translator(source.value, origin=source.coordinate).translate_formula(target.coordinate)
            except (TypeError, ValueError):
                target.value = source.value
    if source_row in sheet.row_dimensions:
        dimension = copy.copy(sheet.row_dimensions[source_row])
        dimension.index = target_row
        sheet.row_dimensions[target_row] = dimension


def _expanded_multi_range(value: Any, old_end: int, new_end: int, min_column: int, max_column: int) -> MultiCellRange:
    ranges: list[str] = []
    for cell_range in MultiCellRange(str(value)).ranges:
        current = copy.copy(cell_range)
        overlaps_columns = current.min_col <= max_column and current.max_col >= min_column
        if overlaps_columns and current.min_row <= old_end <= current.max_row:
            current.max_row = max(current.max_row, new_end)
        ranges.append(str(current))
    return MultiCellRange(" ".join(ranges))


def _extend_validations_and_formatting(sheet: Any, old_end: int, new_end: int, min_column: int, max_column: int) -> None:
    for validation in sheet.data_validations.dataValidation:
        validation.sqref = _expanded_multi_range(validation.sqref, old_end, new_end, min_column, max_column)
    rebuilt: OrderedDict[Any, list[Any]] = OrderedDict()
    for conditional, rules in sheet.conditional_formatting._cf_rules.items():
        expanded = _expanded_multi_range(conditional.sqref, old_end, new_end, min_column, max_column)
        key = ConditionalFormatting(sqref=expanded, pivot=conditional.pivot, extLst=getattr(conditional, "extLst", None))
        rebuilt[key] = rules
    sheet.conditional_formatting._cf_rules = rebuilt


def _extend_defined_names(workbook: Any, sheet_name: str, old_end: int, new_end: int) -> None:
    escaped = re.escape(sheet_name.replace("'", "''"))
    pattern = re.compile(rf"(?P<prefix>'{escaped}'|{re.escape(sheet_name)})!(?P<start>\$?[A-Z]+\$?\d+):(?P<column>\$?[A-Z]+\$?){old_end}(?!\d)")
    for defined_name in workbook.defined_names.values():
        text = str(defined_name.attr_text or "")
        defined_name.attr_text = pattern.sub(lambda match: f"{match.group('prefix')}!{match.group('start')}:{match.group('column')}{new_end}", text)


def _extend_totals_formula(value: Any, old_end: int, new_end: int) -> Any:
    if not (isinstance(value, str) and value.startswith("=")):
        return value
    range_end = re.compile(
        rf"(?<=:)(?P<column>\$?[A-Z]{{1,3}})(?P<absolute>\$?){old_end}(?!\d)",
        re.I,
    )
    return range_end.sub(
        lambda match: (
            f"{match.group('column')}{match.group('absolute')}{new_end}"
        ),
        value,
    )


def _prepare_region_capacity(workbook: Any, region: dict[str, Any], row_count: int) -> None:
    capacity = int(region.get("existing_capacity", 0))
    if row_count <= capacity:
        return
    policy = str(region.get("row_growth_policy", "fixed_capacity"))
    if policy == "fixed_capacity":
        raise ValueError(
            f"Mapped region {region.get('semantic_role')} at {region.get('sheet')} has capacity {capacity} for {row_count} rows and no approved growth policy."
        )
    prototype = region.get("prototype_data_row")
    if not prototype:
        raise ValueError(f"Mapped region {region.get('semantic_role')} requires growth but has no approved prototype data row.")
    sheet = workbook[str(region["sheet"])]
    start = int(region["data_start_row"])
    table_binding = region.get("table") or {}
    if policy == "excel_table":
        table_name = str(table_binding.get("name", ""))
        if table_name not in sheet.tables:
            raise ValueError(f"Mapped Excel table {table_name or '<missing>'} is not present on {sheet.title}.")
        table = sheet.tables[table_name]
        min_col, header_row, max_col, table_end = range_boundaries(str(table.ref))
        totals = bool(table.totalsRowShown)
        old_data_end = table_end - (1 if totals else 0)
        new_data_end = start + row_count - 1
        if totals and new_data_end > old_data_end:
            sheet.insert_rows(old_data_end + 1, new_data_end - old_data_end)
        for row in range(old_data_end + 1, new_data_end + 1):
            _copy_row_style(sheet, int(prototype), row, set(range(min_col, max_col + 1)), copy_formulas=True)
        new_table_end = new_data_end + (1 if totals else 0)
        if totals:
            for column in range(min_col, max_col + 1):
                total_cell = sheet.cell(new_table_end, column)
                total_cell.value = _extend_totals_formula(
                    total_cell.value,
                    old_data_end,
                    new_data_end,
                )
        table.ref = f"{get_column_letter(min_col)}{header_row}:{get_column_letter(max_col)}{new_table_end}"
        if table.autoFilter is not None:
            table.autoFilter.ref = table.ref
        _extend_validations_and_formatting(sheet, old_data_end, new_data_end, min_col, max_col)
        _extend_defined_names(workbook, sheet.title, old_data_end, new_data_end)
        return
    if policy == "prototype_row":
        columns = set(range(1, max(sheet.max_column, max(int(value) for value in region["columns"].values())) + 1))
        for row in range(start + capacity, start + row_count):
            _copy_row_style(sheet, int(prototype), row, columns, copy_formulas=True)
        old_end = start + max(capacity, 1) - 1
        new_end = start + row_count - 1
        min_column = min(int(value) for value in region["columns"].values())
        max_column = max(int(value) for value in region["columns"].values())
        permitted = region.get("permitted_changes", {})
        if permitted.get("data_validation_ranges") or permitted.get("conditional_formatting_ranges"):
            _extend_validations_and_formatting(sheet, old_end, new_end, min_column, max_column)
        if permitted.get("defined_names"):
            _extend_defined_names(workbook, sheet.title, old_end, new_end)
        if permitted.get("auto_filter") and sheet.auto_filter.ref:
            filter_min_col, filter_min_row, filter_max_col, filter_max_row = range_boundaries(str(sheet.auto_filter.ref))
            if filter_max_row == old_end:
                sheet.auto_filter.ref = (
                    f"{get_column_letter(filter_min_col)}{filter_min_row}:{get_column_letter(filter_max_col)}{new_end}"
                )
        return
    raise ValueError(f"Unsupported approved row-growth policy: {policy}")


def _fill_region(
    workbook,
    region: dict[str, Any],
    rows: list[dict[str, Any]],
    value_for: Callable[[dict[str, Any], str], Any],
) -> None:
    if not region:
        return
    _prepare_region_capacity(workbook, region, len(rows))
    sheet = workbook[str(region["sheet"])]
    start_row = int(region["data_start_row"])
    columns = {str(key): int(value) for key, value in region["columns"].items()}
    mapped_columns = set(columns.values())
    end_row = start_row + max(int(region.get("existing_capacity", 0)), len(rows), 1) - 1
    protected = set(region.get("protected_formula_cells", []))
    for row in range(start_row, end_row + 1):
        for field, column in columns.items():
            cell = sheet.cell(row, column)
            if cell.coordinate in protected or (isinstance(cell.value, str) and cell.value.startswith("=")):
                raise ValueError(f"Mapped value target {sheet.title}!{cell.coordinate} is formula-protected and cannot be overwritten.")
            cell.value = None
            if field == "event" and region.get("permitted_changes", {}).get("hyperlinks"):
                cell.hyperlink = None
    for offset, item in enumerate(rows):
        row = start_row + offset
        for field, column in columns.items():
            cell = sheet.cell(row, column)
            if cell.coordinate in protected or (isinstance(cell.value, str) and cell.value.startswith("=")):
                raise ValueError(f"Mapped value target {sheet.title}!{cell.coordinate} is formula-protected and cannot be overwritten.")
            set_cell_value(cell, value_for(item, field))
    if sheet.auto_filter.ref:
        start_column = min(mapped_columns)
        end_column = max(mapped_columns)
        header_row = int(region["header_row"])
        sheet.auto_filter.ref = (
            f"{sheet.cell(header_row, start_column).coordinate}:{sheet.cell(max(start_row, start_row + len(rows) - 1), end_column).coordinate}"
        )


def _event_matrix_value(plan: dict[str, Any], event: dict[str, Any], field: str) -> Any:
    values = {
        "journey": " | ".join(event_journey_names(plan, event)),
        "event": event.get("event_name", ""),
        "classification": classification_label(plan, str(event.get("classification", ""))),
        "definition": event.get("definition", ""),
        "trigger": event.get("trigger", ""),
        "locations": location_text(event),
        "variables": "\n".join(
            f"{parameter.get('name')} ({requirement_label(plan, str(parameter.get('requirement', '')))})"
            for parameter in event.get("parameters", [])
            if isinstance(parameter, dict)
        ),
        "datalayer": datalayer_code(event, plan.get("data_layer_convention")),
        "notes": event.get("notes", ""),
    }
    return values.get(field, "")


def _reference_value(plan: dict[str, Any], row: dict[str, Any], field: str) -> Any:
    values = {
        "variable": row.get("name", ""),
        "scope": scope_label(plan, str(row.get("scope", ""))),
        "type": row.get("type", ""),
        "definition": row.get("definition", ""),
        "rule": row.get("rule", ""),
        "possible_values_or_examples": row.get("possible_values_or_example", ""),
        "example": row.get("possible_values_or_example", ""),
        "values": (f"{row.get('rule', '')}\n{row.get('possible_values_or_example', '')}".strip()),
        "concerned_events": " | ".join(row.get("events", [])),
    }
    return values.get(field, "")


def _coordinate_right(coordinate: str) -> str:
    letters, row = coordinate_from_string(coordinate)
    column = column_index_from_string(letters)
    from openpyxl.utils import get_column_letter

    return f"{get_column_letter(column + 1)}{row}"


def _fill_event_tab(
    workbook,
    plan: dict[str, Any],
    mapping: dict[str, Any],
    event: dict[str, Any],
) -> None:
    sheet = workbook[str(mapping["sheet"])]
    field_labels = mapping.get("field_labels", {})
    direct_values = {
        "event": event.get("event_name", ""),
        "journey": " | ".join(event_journey_names(plan, event)),
        "classification": classification_label(plan, str(event.get("classification", ""))),
        "definition": event.get("definition", ""),
        "trigger": event.get("trigger", ""),
        "locations": location_text(event),
        "notes": event.get("notes", ""),
    }
    for field, value in direct_values.items():
        if field in field_labels:
            set_cell_value(
                sheet[_coordinate_right(str(field_labels[field]))],
                value,
            )

    parameter_region = mapping.get("parameter_region") or {}
    if parameter_region:
        rows = [item for item in event.get("parameters", []) if isinstance(item, dict)]

        def parameter_value(parameter: dict[str, Any], field: str) -> Any:
            source = str(parameter.get("data_layer_path", ""))
            if parameter.get("source"):
                source += f"\n{parameter.get('source')}"
            values = {
                "variable": parameter.get("name", ""),
                "scope": scope_label(plan, str(parameter.get("scope", ""))),
                "type": parameter.get("type", ""),
                "requirement": requirement_label(plan, str(parameter.get("requirement", ""))),
                "condition": parameter.get("condition", ""),
                "definition": parameter.get("definition", ""),
                "rule": value_rule_text(parameter, plan),
                "possible_values_or_examples": possible_values_or_example(parameter),
                "values": combined_value_rule_text(parameter, plan),
                "example": compact_value(parameter.get("example")),
                "source_path": source,
            }
            return values.get(field, "")

        _fill_region(workbook, parameter_region, rows, parameter_value)
    if mapping.get("data_layer_cell"):
        set_cell_value(
            sheet[str(mapping["data_layer_cell"])],
            datalayer_code(event, plan.get("data_layer_convention")),
        )


def _event_tab_assignments(
    workbook,
    mappings: list[dict[str, Any]],
    events: list[dict[str, Any]],
) -> tuple[
    list[tuple[dict[str, Any], dict[str, Any], str]],
    list[str],
    list[tuple[dict[str, Any], str]],
]:
    by_name = {str(event.get("event_name", "")): event for event in events}
    assignments: list[tuple[dict[str, Any], dict[str, Any], str]] = []
    assigned: set[str] = set()
    reusable_mappings: list[tuple[dict[str, Any], str]] = []
    non_reusable_mappings: list[tuple[dict[str, Any], str]] = []
    for mapping in mappings:
        sheet = workbook[str(mapping["sheet"])]
        existing = str(sheet[str(mapping["event_name_cell"])].value or "").strip()
        if existing in by_name:
            assignments.append((mapping, by_name[existing], existing))
            assigned.add(existing)
        elif not existing or mapping.get("reusable") is True:
            reusable_mappings.append((mapping, existing))
        else:
            non_reusable_mappings.append((mapping, existing))
    remaining = [event for event in events if str(event.get("event_name", "")) not in assigned]
    reused_count = 0
    for (mapping, existing), event in zip(reusable_mappings, remaining, strict=False):
        assignments.append((mapping, event, existing))
        assigned.add(str(event.get("event_name", "")))
        reused_count += 1
    missing = [str(event.get("event_name", "")) for event in events if str(event.get("event_name", "")) not in assigned]
    return (
        assignments,
        missing,
        [
            *reusable_mappings[reused_count:],
            *non_reusable_mappings,
        ],
    )


def _clone_event_tabs_for_missing(
    workbook: Any,
    prototype: dict[str, Any],
    missing_events: list[dict[str, Any]],
    fidelity_before: dict[str, Any],
    fidelity_authorized: dict[str, Any],
) -> list[dict[str, Any]]:
    if not missing_events:
        return []
    if not prototype or prototype.get("cloning_allowed") is not True:
        return []
    prototype_sheet_name = str(prototype.get("sheet", ""))
    if prototype_sheet_name not in workbook.sheetnames:
        raise ValueError("The approved event-tab prototype sheet is missing from the supplied template.")
    prototype_mapping = prototype.get("mapping") or {}
    before_by_title = {str(item["title"]): item for item in fidelity_before.get("sheets", [])}
    prototype_snapshot = before_by_title.get(prototype_sheet_name)
    if prototype_snapshot is None:
        raise ValueError("The approved event-tab prototype could not be included in the fidelity baseline.")
    created: list[dict[str, Any]] = []
    source_sheet = workbook[prototype_sheet_name]
    for event in missing_events:
        event_name = str(event.get("event_name", ""))
        target = workbook.copy_worksheet(source_sheet)
        target.freeze_panes = source_sheet.freeze_panes
        target.views = copy.copy(source_sheet.views)
        target.sheet_properties = copy.copy(source_sheet.sheet_properties)
        target.sheet_format = copy.copy(source_sheet.sheet_format)
        target.page_margins = copy.copy(source_sheet.page_margins)
        target.page_setup = copy.copy(source_sheet.page_setup)
        target.print_options = copy.copy(source_sheet.print_options)
        target.protection = copy.copy(source_sheet.protection)
        target.auto_filter = copy.copy(source_sheet.auto_filter)
        target.print_area = source_sheet.print_area
        target.print_title_rows = source_sheet.print_title_rows
        target.print_title_cols = source_sheet.print_title_cols
        target.title = safe_sheet_title(event_name, workbook.sheetnames)
        target.sheet_state = "visible"
        cloned_mapping = copy.deepcopy(prototype_mapping)
        cloned_mapping["sheet"] = target.title
        cloned_mapping["existing_event_name"] = ""
        cloned_mapping["reusable"] = True
        cloned_mapping["_template_source_sheet"] = prototype_sheet_name
        if cloned_mapping.get("parameter_region"):
            cloned_mapping["parameter_region"]["sheet"] = target.title
        authorize_added_sheet(fidelity_authorized, target.title, prototype_snapshot)
        created.append(cloned_mapping)
    return created


def _clear_event_tab(workbook, mapping: dict[str, Any]) -> None:
    sheet = workbook[str(mapping["sheet"])]
    field_labels = mapping.get("field_labels", {})
    for field in (
        "event",
        "journey",
        "classification",
        "definition",
        "trigger",
        "locations",
        "notes",
    ):
        coordinate = field_labels.get(field)
        if not coordinate:
            continue
        value_cell = sheet[_coordinate_right(str(coordinate))]
        value_cell.value = None
        value_cell.hyperlink = None
    parameter_region = mapping.get("parameter_region") or {}
    if parameter_region:
        _fill_region(workbook, parameter_region, [], lambda _item, _field: "")
    if mapping.get("data_layer_cell"):
        cell = sheet[str(mapping["data_layer_cell"])]
        cell.value = None
        cell.hyperlink = None
    sheet.sheet_state = "hidden"


def _clear_hyperlinks_to_sheets(workbook, sheet_names: set[str]) -> None:
    if not sheet_names:
        return
    prefixes = {prefix for name in sheet_names for prefix in (f"#{quote_sheetname(name)}!", f"#{name}!")}
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                hyperlink = cell.hyperlink
                target = str(getattr(hyperlink, "target", "") or "") if hyperlink else ""
                if target and any(target.startswith(prefix) for prefix in prefixes):
                    cell.hyperlink = None


def _link_event_matrix(
    workbook,
    region: dict[str, Any],
    events: list[dict[str, Any]],
    event_sheets: dict[str, str],
) -> None:
    event_column = (region.get("columns") or {}).get("event")
    if not event_column:
        return
    sheet = workbook[str(region["sheet"])]
    start_row = int(region["data_start_row"])
    for offset, event in enumerate(events):
        event_name = str(event.get("event_name", ""))
        target = event_sheets.get(event_name)
        if target:
            sheet.cell(start_row + offset, int(event_column)).hyperlink = f"#{quote_sheetname(target)}!A1"


def adapt(
    plan: dict[str, Any],
    template: Path,
    mapping: dict[str, Any],
    *,
    enforce_preflight: bool = True,
):
    mapping = normalize_template_mapping(mapping, template)
    mapping_errors = validate_template_mapping(mapping)
    if mapping_errors:
        raise ValueError("Template mapping contract failed:\n- " + "\n- ".join(mapping_errors))
    expected_hash = str(mapping.get("template", {}).get("sha256", ""))
    if expected_hash and sha256(template) != expected_hash:
        raise ValueError("The supplied workbook no longer matches the inspected template hash. Inspect the current file again before adaptation.")
    if mapping.get("review_required"):
        raise ValueError("Template mapping requires review before adaptation:\n- " + "\n- ".join(str(item) for item in mapping["review_required"]))
    selected_writer = ensure_writer_supported(mapping.get("preflight", {}), str(mapping.get("policy", {}).get("writer", "auto")))
    if enforce_preflight and selected_writer != "openpyxl":
        raise ValueError("This supplied template requires the native Excel preservation adapter; use save_adapted_workbook().")
    workbook = load_workbook(
        template,
        data_only=False,
        read_only=False,
        keep_links=True,
        keep_vba=template.suffix.lower() == ".xlsm",
    )
    fidelity_before = workbook_fidelity_snapshot(workbook)
    fidelity_authorized = authorized_template_changes(workbook, mapping, plan)
    regions = mapping.get("regions", {})
    event_matrix = regions.get("event_matrix") or {}
    parameter_reference = regions.get("parameter_reference") or {}
    data_layer_table = regions.get("data_layer_table") or {}
    event_tabs = copy.deepcopy(regions.get("event_tabs") or [])
    event_tab_prototype = copy.deepcopy(regions.get("event_tab_prototype") or {})
    if event_tab_prototype:
        prototype_sheet = str(event_tab_prototype.get("sheet", ""))
        event_tabs = [item for item in event_tabs if str(item.get("sheet", "")) != prototype_sheet]
    if not event_matrix:
        raise ValueError("The mapping has no Event Matrix region for analyst review.")
    if not parameter_reference:
        raise ValueError("The mapping has no Parameter Reference region.")
    if not event_tabs and not data_layer_table and not event_tab_prototype:
        raise ValueError(
            "The mapping has no legitimate location for complete dataLayer examples. Approve and map a suitable template region before adaptation."
        )
    tab_contracts = [*event_tabs, *([event_tab_prototype.get("mapping", {})] if event_tab_prototype else [])]
    if tab_contracts and not data_layer_table and any(not item.get("data_layer_cell") for item in tab_contracts):
        raise ValueError("At least one mapped event tab has no dataLayer example cell. Map a legitimate code region before adaptation.")

    events = [item for item in plan.get("events", []) if isinstance(item, dict)]
    _fill_region(
        workbook,
        event_matrix,
        events,
        lambda item, field: _event_matrix_value(plan, item, field),
    )
    _fill_region(
        workbook,
        parameter_reference,
        parameter_reference_rows(
            plan,
            str(parameter_reference.get("semantic_role", "all_used_parameters")),
        ),
        lambda item, field: _reference_value(plan, item, field),
    )
    if data_layer_table:
        _fill_region(
            workbook,
            data_layer_table,
            events,
            lambda item, field: _event_matrix_value(plan, item, field),
        )

    assignments, missing, unused = _event_tab_assignments(
        workbook,
        event_tabs,
        events,
    )
    if missing and event_tab_prototype:
        by_name = {str(event.get("event_name", "")): event for event in events}
        event_tabs.extend(
            _clone_event_tabs_for_missing(
                workbook,
                event_tab_prototype,
                [by_name[name] for name in missing],
                fidelity_before,
                fidelity_authorized,
            )
        )
        assignments, missing, unused = _event_tab_assignments(workbook, event_tabs, events)
    obsolete_sheet_names: set[str] = set()
    event_sheets: dict[str, str] = {}
    for event_mapping, event, previous_event_name in assignments:
        original_sheet_name = str(event_mapping["sheet"])
        event_mapping["_template_source_sheet"] = str(event_mapping.get("_template_source_sheet") or original_sheet_name)
        _fill_event_tab(workbook, plan, event_mapping, event)
        sheet = workbook[original_sheet_name]
        sheet.sheet_state = "visible"
        event_name = str(event.get("event_name", ""))
        if previous_event_name and previous_event_name != event_name and sheet.title.casefold() == previous_event_name.casefold():
            used = [name for name in workbook.sheetnames if name != sheet.title]
            new_title = safe_sheet_title(event_name, used)
            obsolete_sheet_names.add(sheet.title)
            sheet.title = new_title
            event_mapping["sheet"] = new_title
            if event_mapping.get("parameter_region"):
                event_mapping["parameter_region"]["sheet"] = new_title
        event_sheets[event_name] = sheet.title
    for event_mapping, previous_event_name in unused:
        if not previous_event_name:
            continue
        obsolete_sheet_names.add(str(event_mapping["sheet"]))
        _clear_event_tab(workbook, event_mapping)
    if missing and not data_layer_table:
        raise ValueError("The template has no mapped event tab for: " + ", ".join(missing) + ". Do not add sheets without explicit template approval.")
    _clear_hyperlinks_to_sheets(workbook, obsolete_sheet_names)
    _link_event_matrix(workbook, event_matrix, events, event_sheets)
    embed_model(workbook, plan)
    fidelity_report = compare_template_fidelity(
        fidelity_before,
        workbook_fidelity_snapshot(workbook),
        fidelity_authorized,
    )
    if fidelity_report["violations"]:
        first = ", ".join(
            f"{item.get('kind')}:{item.get('sheet', item.get('sheet_index', ''))}" + (f"!{item.get('coordinate')}" if item.get("coordinate") else "")
            for item in fidelity_report["violations"][:12]
        )
        raise ValueError("Supplied-template fidelity gate failed for unmapped content: " + first)
    workbook._ga4_template_fidelity_report = fidelity_report
    workbook._ga4_template_fidelity_before = fidelity_before
    workbook._ga4_template_fidelity_authorized = fidelity_authorized
    effective_mapping = copy.deepcopy(mapping)
    effective_mapping["regions"]["event_tabs"] = event_tabs
    workbook._ga4_effective_mapping = effective_mapping
    return workbook


def save_adapted_workbook(
    plan: dict[str, Any],
    template: Path,
    mapping: dict[str, Any],
    output: Path,
) -> dict[str, Any]:
    mapping = normalize_template_mapping(mapping, template)
    if template.suffix.lower() == ".xlsm" and output.suffix.lower() != ".xlsm":
        raise ValueError("An XLSM supplied template must be delivered as XLSM.")
    selected_writer = ensure_writer_supported(mapping.get("preflight", {}), str(mapping.get("policy", {}).get("writer", "auto")))
    if selected_writer == "openpyxl":
        workbook = adapt(plan, template, mapping)
        output.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output)
        before = workbook._ga4_template_fidelity_before
        authorized = workbook._ga4_template_fidelity_authorized
        effective_mapping = workbook._ga4_effective_mapping
    else:
        from native_excel_adapter import save_with_native_excel

        result = save_with_native_excel(plan, template, mapping, output)
        before = result["before"]
        authorized = result["authorized"]
        effective_mapping = result["effective_mapping"]
    reopened = load_workbook(
        output,
        data_only=False,
        read_only=False,
        keep_links=True,
        keep_vba=output.suffix.lower() == ".xlsm",
    )
    final_fidelity = compare_template_fidelity(before, workbook_fidelity_snapshot(reopened), authorized)
    final_fidelity = add_package_fidelity(final_fidelity, template, output)
    final_fidelity["writer"] = selected_writer
    if final_fidelity["violations"]:
        raise ValueError(
            "Saved supplied-template fidelity gate failed: "
            + ", ".join(str(item.get("kind")) for item in final_fidelity["violations"][:12])
        )
    return {
        "writer": selected_writer,
        "fidelity": final_fidelity,
        "effective_mapping": effective_mapping,
    }


def main() -> int:
    args = parse_args()
    try:
        plan = load_json(args.plan)
        issues = validate_plan(plan)
        if issues:
            print(render_text(issues), file=sys.stderr)
        if issues:
            return 1
        mapping = load_json(args.mapping)
        result = save_adapted_workbook(plan, args.template, mapping, args.output)
        if args.fidelity_report:
            args.fidelity_report.parent.mkdir(parents=True, exist_ok=True)
            args.fidelity_report.write_text(
                json.dumps(
                    result["fidelity"],
                    indent=2,
                    ensure_ascii=False,
                )
                + "\n",
                encoding="utf-8",
            )
    except (OSError, ValueError, KeyError, json.JSONDecodeError) as error:
        print(str(error), file=sys.stderr)
        return 2
    print(args.output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
