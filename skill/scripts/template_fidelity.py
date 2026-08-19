from __future__ import annotations

import copy
import hashlib
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any
from xml.etree.ElementTree import tostring
from zipfile import BadZipFile, ZipFile

from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string
from openpyxl.worksheet.cell_range import MultiCellRange

INTERNAL_PREFIX = "__tracking_plan_"
VALUE_PROPERTIES = {"value", "data_type"}


def _value(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def _color(value: Any) -> dict[str, Any] | None:
    if value is None:
        return None
    return {
        "type": str(value.type or ""),
        "rgb": str(value.rgb or ""),
        "indexed": value.indexed,
        "theme": value.theme,
        "tint": float(value.tint or 0),
        "auto": value.auto,
    }


def _side(value: Any) -> dict[str, Any] | None:
    if value is None:
        return None
    return {"style": str(value.style or ""), "color": _color(value.color)}


def _style_snapshot(cell: Any) -> dict[str, Any]:
    font = cell.font
    fill = cell.fill
    border = cell.border
    alignment = cell.alignment
    protection = cell.protection
    return {
        "font": {
            "name": str(font.name or ""),
            "size": float(font.sz) if font.sz is not None else None,
            "bold": bool(font.b),
            "italic": bool(font.i),
            "underline": str(font.u or ""),
            "strike": bool(font.strike),
            "color": _color(font.color),
            "vert_align": str(font.vertAlign or ""),
            "scheme": str(font.scheme or ""),
            "family": font.family,
            "charset": font.charset,
            "outline": bool(font.outline),
            "shadow": bool(font.shadow),
            "condense": bool(font.condense),
            "extend": bool(font.extend),
        },
        "fill": {
            "fill_type": str(fill.fill_type or ""),
            "pattern_type": str(getattr(fill, "patternType", "") or ""),
            "fg_color": _color(getattr(fill, "fgColor", None)),
            "bg_color": _color(getattr(fill, "bgColor", None)),
        },
        "border": {
            "left": _side(border.left),
            "right": _side(border.right),
            "top": _side(border.top),
            "bottom": _side(border.bottom),
            "diagonal": _side(border.diagonal),
            "vertical": _side(border.vertical),
            "horizontal": _side(border.horizontal),
            "diagonal_up": bool(border.diagonalUp),
            "diagonal_down": bool(border.diagonalDown),
            "outline": bool(border.outline),
        },
        "alignment": {
            "horizontal": str(alignment.horizontal or ""),
            "vertical": str(alignment.vertical or ""),
            "text_rotation": int(alignment.textRotation or 0),
            "wrap_text": bool(alignment.wrapText),
            "shrink_to_fit": bool(alignment.shrinkToFit),
            "indent": float(alignment.indent or 0),
            "relative_indent": float(alignment.relativeIndent or 0),
            "justify_last_line": bool(alignment.justifyLastLine),
            "reading_order": float(alignment.readingOrder or 0),
        },
        "number_format": str(cell.number_format),
        "protection": {"locked": bool(protection.locked), "hidden": bool(protection.hidden)},
        "quote_prefix": bool(cell.quotePrefix),
        "pivot_button": bool(cell.pivotButton),
    }


def _cell_snapshot(cell: Any) -> dict[str, Any]:
    hyperlink = None
    if cell.hyperlink is not None:
        hyperlink = {
            "target": str(cell.hyperlink.target or ""),
            "location": str(cell.hyperlink.location or ""),
            "tooltip": str(cell.hyperlink.tooltip or ""),
            "display": str(cell.hyperlink.display or ""),
        }
    comment = None
    if cell.comment:
        comment = {
            "text": str(cell.comment.text),
            "author": str(cell.comment.author or ""),
            "width": str(cell.comment.width or ""),
            "height": str(cell.comment.height or ""),
        }
    return {
        "value": _value(cell.value),
        "data_type": str(cell.data_type),
        "is_formula": bool(isinstance(cell.value, str) and cell.value.startswith("=")),
        "hyperlink": hyperlink,
        "style": _style_snapshot(cell),
        "comment": comment,
    }


def _serialised_model(value: Any) -> str:
    if value is None:
        return ""
    try:
        return tostring(value.to_tree(), encoding="unicode")
    except (AttributeError, TypeError, ValueError):
        return str(value)


def _serialised_model_without(value: Any, *attributes: str) -> str:
    if value is None:
        return ""
    clone = copy.copy(value)
    for attribute in attributes:
        if hasattr(clone, attribute):
            setattr(clone, attribute, None)
    return _serialised_model(clone)


def drawing_anchor_snapshot(anchor: Any) -> dict[str, Any]:
    if isinstance(anchor, str):
        return {"kind": "coordinate", "coordinate": anchor}
    result: dict[str, Any] = {"kind": type(anchor).__name__}
    for field in ("_from", "to"):
        marker = getattr(anchor, field, None)
        if marker is None:
            continue
        result[field] = {
            name: int(getattr(marker, name, 0) or 0)
            for name in ("col", "colOff", "row", "rowOff")
        }
    extent = getattr(anchor, "ext", None)
    if extent is not None:
        result["extent"] = {
            "cx": int(getattr(extent, "cx", 0) or 0),
            "cy": int(getattr(extent, "cy", 0) or 0),
        }
    position = getattr(anchor, "pos", None)
    if position is not None:
        result["position"] = {
            "x": int(getattr(position, "x", 0) or 0),
            "y": int(getattr(position, "y", 0) or 0),
        }
    return result


def drawing_object_snapshot(value: Any, *, include_model: bool) -> dict[str, Any]:
    snapshot = {
        "type": type(value).__name__,
        "anchor": drawing_anchor_snapshot(getattr(value, "anchor", "")),
        "width": float(getattr(value, "width", 0) or 0),
        "height": float(getattr(value, "height", 0) or 0),
    }
    if include_model:
        snapshot["model"] = _serialised_model(value)
    return snapshot


def _table_snapshot(sheet: Any) -> list[dict[str, Any]]:
    result = []
    for table in sheet.tables.values():
        columns = []
        for column in table.tableColumns:
            columns.append(
                {
                    "id": column.id,
                    "name": str(column.name),
                    "calculated_formula": str(getattr(column.calculatedColumnFormula, "text", "") or ""),
                    "totals_function": str(column.totalsRowFunction or ""),
                    "totals_label": str(column.totalsRowLabel or ""),
                }
            )
        result.append(
            {
                "name": str(table.name),
                "display_name": str(table.displayName),
                "ref": str(table.ref),
                "totals_row_shown": bool(table.totalsRowShown),
                "style": _serialised_model(table.tableStyleInfo),
                "auto_filter_ref": str(getattr(table.autoFilter, "ref", "") or ""),
                "auto_filter_definition": _serialised_model_without(table.autoFilter, "ref"),
                "columns": columns,
            }
        )
    return sorted(result, key=lambda item: item["name"])


def _validation_snapshot(sheet: Any) -> list[dict[str, Any]]:
    definition_fields = (
        "type",
        "formula1",
        "formula2",
        "allowBlank",
        "operator",
        "errorStyle",
        "errorTitle",
        "error",
        "promptTitle",
        "prompt",
        "showErrorMessage",
        "showInputMessage",
        "showDropDown",
        "imeMode",
    )
    return sorted(
        [
            {
                "sqref": str(item.sqref),
                "definition": {
                    field: _value(getattr(item, field, None))
                    for field in definition_fields
                },
            }
            for item in sheet.data_validations.dataValidation
        ],
        key=lambda item: (repr(item["definition"]), item["sqref"]),
    )


def _conditional_format_snapshot(sheet: Any) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for conditional, rules in sheet.conditional_formatting._cf_rules.items():
        result.append({"sqref": str(conditional.sqref), "rules": [_serialised_model(rule) for rule in rules]})
    return sorted(result, key=lambda item: item["sqref"])


def _dimension_snapshot(dimensions: Any) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for key, item in dimensions.items():
        result[str(key)] = {
            "hidden": bool(item.hidden),
            "outline_level": int(item.outlineLevel or 0),
            "collapsed": bool(item.collapsed),
            "style": int(item.style_id or 0),
            "height": float(item.height) if getattr(item, "height", None) is not None else None,
            "width": float(item.width) if getattr(item, "width", None) is not None else None,
            "best_fit": bool(getattr(item, "bestFit", False)),
            "min": getattr(item, "min", None),
            "max": getattr(item, "max", None),
        }
    return result


def _defined_name_snapshot(workbook: Any) -> list[dict[str, Any]]:
    return sorted(
        [
            {
                "name": str(item.name),
                "attr_text": str(item.attr_text or ""),
                "local_sheet_id": item.localSheetId,
                "hidden": bool(item.hidden),
            }
            for item in workbook.defined_names.values()
        ],
        key=lambda item: (item["name"], item["local_sheet_id"] if item["local_sheet_id"] is not None else -1),
    )


def _workbook_properties_snapshot(properties: Any) -> dict[str, Any]:
    return {
        field: _value(getattr(properties, field, None))
        for field in (
            "creator",
            "lastModifiedBy",
            "title",
            "subject",
            "description",
            "identifier",
            "language",
            "created",
            "contentStatus",
            "category",
            "keywords",
            "lastPrinted",
            "revision",
            "version",
        )
    }


def workbook_fidelity_snapshot(workbook: Any) -> dict[str, Any]:
    sheets: list[dict[str, Any]] = []
    for index, sheet in enumerate(workbook.worksheets):
        if sheet.title.startswith(INTERNAL_PREFIX):
            continue
        cells = {
            cell.coordinate: _cell_snapshot(cell)
            for cell in sheet._cells.values()
            if cell.value not in (None, "") or cell.hyperlink is not None or cell.comment is not None or cell.has_style
        }
        sheets.append(
            {
                "index": index,
                "title": sheet.title,
                "state": sheet.sheet_state,
                "cells": cells,
                "merged_cells": sorted(str(item) for item in sheet.merged_cells.ranges),
                "images": [
                    drawing_object_snapshot(item, include_model=False)
                    for item in sheet._images
                ],
                "charts": [
                    drawing_object_snapshot(item, include_model=True)
                    for item in sheet._charts
                ],
                "tables": _table_snapshot(sheet),
                "data_validations": _validation_snapshot(sheet),
                "conditional_formatting": _conditional_format_snapshot(sheet),
                "freeze_panes": str(sheet.freeze_panes or ""),
                "auto_filter": str(sheet.auto_filter.ref or ""),
                "print_area": str(sheet.print_area or ""),
                "print_title_rows": str(sheet.print_title_rows or ""),
                "print_title_cols": str(sheet.print_title_cols or ""),
                "row_dimensions": _dimension_snapshot(sheet.row_dimensions),
                "column_dimensions": _dimension_snapshot(sheet.column_dimensions),
                "sheet_properties": _serialised_model(sheet.sheet_properties),
                "sheet_format": _serialised_model(sheet.sheet_format),
                "sheet_views": _serialised_model(sheet.views),
                "page_margins": _serialised_model(sheet.page_margins),
                "page_setup": _serialised_model(sheet.page_setup),
                "sheet_protection": _serialised_model(sheet.protection),
            }
        )
    return {
        "snapshot_version": "2.0.0",
        "sheets": sheets,
        "defined_names": _defined_name_snapshot(workbook),
        "workbook_properties": _workbook_properties_snapshot(workbook.properties),
        "workbook_security": _serialised_model(workbook.security),
        "calculation": _serialised_model(workbook.calculation),
        "workbook_views": [_serialised_model(item) for item in workbook.views],
    }


def _right(coordinate: str) -> str:
    letters, row = coordinate_from_string(coordinate)
    return f"{get_column_letter(column_index_from_string(letters) + 1)}{row}"


def _property_authorisation() -> dict[str, Any]:
    return {
        "cells": {},
        "sheet_properties": {},
        "structure_properties": {},
        "obsolete_targets": set(),
        "added_sheets": {},
        "new_cell_prototypes": {},
        "new_row_prototypes": {},
        "table_refs": {},
        "range_expansions": {},
        "auto_filter_refs": {},
        "defined_name_values": {},
    }


def _allow_cell(authorized: dict[str, Any], index: int, coordinate: str, *properties: str) -> None:
    authorized["cells"].setdefault((index, coordinate), set()).update(properties)


def _allow_sheet(authorized: dict[str, Any], index: int, *properties: str) -> None:
    authorized["sheet_properties"].setdefault(index, set()).update(properties)


def _allow_structure(authorized: dict[str, Any], index: int, *properties: str) -> None:
    authorized["structure_properties"].setdefault(index, set()).update(properties)


def _expanded_multi_range_text(
    value: str,
    *,
    old_end: int,
    new_end: int,
    min_column: int,
    max_column: int,
) -> str:
    ranges: list[str] = []
    for cell_range in MultiCellRange(value).ranges:
        current = copy.copy(cell_range)
        overlaps_columns = current.min_col <= max_column and current.max_col >= min_column
        if overlaps_columns and current.min_row <= old_end <= current.max_row:
            current.max_row = max(current.max_row, new_end)
        ranges.append(str(current))
    return " ".join(ranges)


def _expanded_defined_name_text(
    value: str,
    *,
    sheet_name: str,
    old_end: int,
    new_end: int,
) -> str:
    escaped = re.escape(sheet_name.replace("'", "''"))
    pattern = re.compile(
        rf"(?P<prefix>'{escaped}'|{re.escape(sheet_name)})!"
        rf"(?P<start>\$?[A-Z]+\$?\d+):(?P<column>\$?[A-Z]+\$?){old_end}(?!\d)"
    )
    return pattern.sub(
        lambda match: (
            f"{match.group('prefix')}!{match.group('start')}:"
            f"{match.group('column')}{new_end}"
        ),
        value,
    )


def authorized_template_changes(workbook: Any, mapping: dict[str, Any], plan: dict[str, Any]) -> dict[str, Any]:
    title_to_index = {sheet.title: index for index, sheet in enumerate(workbook.worksheets)}
    authorized = _property_authorisation()

    def add_region(region: dict[str, Any], row_count: int) -> None:
        if not region:
            return
        title = str(region["sheet"])
        index = title_to_index[title]
        sheet = workbook[title]
        start = int(region["data_start_row"])
        columns = {str(field): int(value) for field, value in region.get("columns", {}).items()}
        existing_capacity = int(region.get("existing_capacity", 0))
        end = start + max(existing_capacity, row_count, 1) - 1
        protected = set(region.get("protected_formula_cells", []))
        for row in range(start, end + 1):
            for field, column in columns.items():
                coordinate = sheet.cell(row, column).coordinate
                if coordinate in protected:
                    continue
                _allow_cell(authorized, index, coordinate, *VALUE_PROPERTIES)
                if field == "event" and region.get("permitted_changes", {}).get("hyperlinks"):
                    _allow_cell(authorized, index, coordinate, "hyperlink")
        permitted = region.get("permitted_changes", {})
        if row_count > existing_capacity:
            policy = str(region.get("row_growth_policy", "fixed_capacity"))
            if policy in {"excel_table", "prototype_row"}:
                old_end = start + max(existing_capacity, 1) - 1
                new_end = start + row_count - 1
                min_column = min(columns.values())
                max_column = max(columns.values())
                if permitted.get("table_ref"):
                    table = region.get("table") or {}
                    table_name = str(table.get("name", ""))
                    table_ref = str(table.get("ref", ""))
                    if table_name and table_ref:
                        min_col, header_row, max_col, _table_end = range_boundaries(table_ref)
                        totals = bool(table.get("totals_row_shown"))
                        expected_end = new_end + (1 if totals else 0)
                        expected_ref = (
                            f"{get_column_letter(min_col)}{header_row}:"
                            f"{get_column_letter(max_col)}{expected_end}"
                        )
                        authorized["table_refs"][(index, table_name)] = expected_ref
                        if totals:
                            old_total = old_end + 1
                            new_total = new_end + 1
                            authorized["new_row_prototypes"][(index, str(new_total))] = str(
                                old_total
                            )
                            for column in range(min_col, max_col + 1):
                                authorized["new_cell_prototypes"][
                                    (index, sheet.cell(new_total, column).coordinate)
                                ] = sheet.cell(old_total, column).coordinate
                expansion = {
                    "old_end": old_end,
                    "new_end": new_end,
                    "min_column": min_column,
                    "max_column": max_column,
                }
                if permitted.get("data_validation_ranges") or permitted.get("conditional_formatting_ranges"):
                    authorized["range_expansions"].setdefault(index, []).append(expansion)
                if permitted.get("defined_names"):
                    for name in workbook.defined_names.values():
                        identity = (str(name.name), name.localSheetId)
                        current = authorized["defined_name_values"].get(
                            identity,
                            str(name.attr_text or ""),
                        )
                        expanded = _expanded_defined_name_text(
                            current,
                            sheet_name=title,
                            old_end=old_end,
                            new_end=new_end,
                        )
                        if expanded != current:
                            authorized["defined_name_values"][identity] = expanded
                prototype_row = region.get("prototype_data_row")
                if prototype_row:
                    for row in range(start + existing_capacity, start + row_count):
                        authorized["new_row_prototypes"][(index, str(row))] = str(
                            prototype_row
                        )
                        for column in range(1, sheet.max_column + 1):
                            authorized["new_cell_prototypes"][(index, sheet.cell(row, column).coordinate)] = sheet.cell(
                                int(prototype_row), column
                            ).coordinate
        if permitted.get("auto_filter"):
            old_filter = str(sheet.auto_filter.ref or "")
            if old_filter:
                min_col, min_row, max_col, max_row = range_boundaries(old_filter)
                desired_end = start + max(row_count, 1) - 1
                authorized["auto_filter_refs"][index] = (
                    f"{get_column_letter(min_col)}{min_row}:"
                    f"{get_column_letter(max_col)}{desired_end}"
                )

    regions = mapping.get("regions", {})
    events = [item for item in plan.get("events", []) if isinstance(item, dict)]
    add_region(regions.get("event_matrix") or {}, len(events))
    add_region(
        regions.get("parameter_reference") or {},
        len(
            {
                (str(parameter.get("name")), str(parameter.get("scope")))
                for event in events
                for parameter in event.get("parameters", [])
                if isinstance(parameter, dict)
            }
        ),
    )
    add_region(regions.get("data_layer_table") or {}, len(events))

    event_names = {str(event.get("event_name", "")) for event in events}
    events_by_name = {str(event.get("event_name", "")): event for event in events}
    assigned_event_names: set[str] = set()
    reusable_event_mappings: list[tuple[dict[str, Any], str]] = []
    event_assignments: dict[int, dict[str, Any]] = {}
    event_tab_mappings = regions.get("event_tabs") or []
    for event_mapping in event_tab_mappings:
        title = str(event_mapping["sheet"])
        existing = str(workbook[title][str(event_mapping["event_name_cell"])].value or "").strip()
        if existing in events_by_name:
            event_assignments[id(event_mapping)] = events_by_name[existing]
            assigned_event_names.add(existing)
        elif not existing or event_mapping.get("reusable") is True:
            reusable_event_mappings.append((event_mapping, existing))
    remaining_events = [
        event
        for event in events
        if str(event.get("event_name", "")) not in assigned_event_names
    ]
    for (event_mapping, _existing), event in zip(
        reusable_event_mappings,
        remaining_events,
        strict=False,
    ):
        event_assignments[id(event_mapping)] = event

    for event_mapping in event_tab_mappings:
        title = str(event_mapping["sheet"])
        index = title_to_index[title]
        permitted = event_mapping.get("permitted_changes", {})
        if permitted.get("sheet_title"):
            _allow_sheet(authorized, index, "title")
        if permitted.get("sheet_visibility"):
            _allow_sheet(authorized, index, "state")
        existing = str(workbook[title][str(event_mapping["event_name_cell"])].value or "").strip()
        if existing and existing not in event_names:
            authorized["obsolete_targets"].add(title)
        for coordinate in event_mapping.get("writable_value_cells", []):
            _allow_cell(authorized, index, str(coordinate), *VALUE_PROPERTIES)
        assigned_event = event_assignments.get(id(event_mapping))
        parameter_count = len(assigned_event.get("parameters", [])) if assigned_event else 0
        add_region(event_mapping.get("parameter_region") or {}, parameter_count)
        if event_mapping.get("data_layer_cell"):
            _allow_cell(authorized, index, str(event_mapping["data_layer_cell"]), *VALUE_PROPERTIES)
    return authorized


def authorize_added_sheet(authorized: dict[str, Any], title: str, prototype_snapshot: dict[str, Any]) -> None:
    authorized["added_sheets"][title] = prototype_snapshot


def _cell_differences(old: dict[str, Any], new: dict[str, Any], allowed: set[str]) -> list[str]:
    differences: list[str] = []
    for key in sorted(set(old) | set(new)):
        if old.get(key) == new.get(key):
            continue
        if key in allowed:
            if key == "data_type" and old.get("is_formula") != new.get("is_formula"):
                differences.append("formula_type")
            continue
        differences.append(key)
    return differences


def _clone_presentation_violations(prototype: dict[str, Any], clone: dict[str, Any]) -> list[dict[str, Any]]:
    violations: list[dict[str, Any]] = []
    for field in (
        "merged_cells",
        "images",
        "charts",
        "tables",
        "data_validations",
        "conditional_formatting",
        "freeze_panes",
        "auto_filter",
        "print_area",
        "print_title_rows",
        "print_title_cols",
        "column_dimensions",
        "sheet_properties",
        "sheet_format",
        "sheet_views",
        "page_margins",
        "page_setup",
        "sheet_protection",
    ):
        if prototype.get(field) != clone.get(field):
            violations.append({"kind": "cloned_sheet_structure", "sheet": clone["title"], "property": field})
    prototype_cells = prototype.get("cells", {})
    clone_cells = clone.get("cells", {})
    for coordinate, old_cell in prototype_cells.items():
        new_cell = clone_cells.get(coordinate)
        if new_cell is None:
            violations.append({"kind": "cloned_sheet_cell_missing", "sheet": clone["title"], "coordinate": coordinate})
            continue
        differences = _cell_differences(old_cell, new_cell, VALUE_PROPERTIES | {"hyperlink"})
        if differences:
            violations.append({"kind": "cloned_sheet_presentation", "sheet": clone["title"], "coordinate": coordinate, "properties": differences})
    return violations


def _expected_tables(
    tables: list[dict[str, Any]],
    authorized: dict[str, Any],
    index: int,
) -> list[dict[str, Any]]:
    expected = copy.deepcopy(tables)
    for table in expected:
        ref = authorized.get("table_refs", {}).get((index, str(table.get("name", ""))))
        if ref:
            table["ref"] = ref
            table["auto_filter_ref"] = ref
    return expected


def _expected_ranged_collection(
    values: list[dict[str, Any]],
    authorized: dict[str, Any],
    index: int,
) -> list[dict[str, Any]]:
    expected = copy.deepcopy(values)
    for item in expected:
        sqref = str(item.get("sqref", ""))
        for expansion in authorized.get("range_expansions", {}).get(index, []):
            sqref = _expanded_multi_range_text(sqref, **expansion)
        item["sqref"] = sqref
    discriminator = "definition" if any("definition" in item for item in expected) else "rules"
    return sorted(
        expected,
        key=lambda item: (repr(item.get(discriminator, "")), item.get("sqref", "")),
    )


def _expected_row_dimensions(
    dimensions: dict[str, dict[str, Any]],
    authorized: dict[str, Any],
    index: int,
) -> dict[str, dict[str, Any]]:
    expected = copy.deepcopy(dimensions)
    for (sheet_index, row), prototype in authorized.get("new_row_prototypes", {}).items():
        if sheet_index == index and prototype in dimensions:
            expected[row] = copy.deepcopy(dimensions[prototype])
    return expected


def _defined_names_match(
    before: list[dict[str, Any]],
    after: list[dict[str, Any]],
    authorized: dict[str, Any],
) -> bool:
    expected = copy.deepcopy(before)
    allowed = authorized.get("defined_name_values", {})
    for item in expected:
        identity = (str(item.get("name", "")), item.get("local_sheet_id"))
        if identity in allowed:
            item["attr_text"] = allowed[identity]
    return expected == after


def compare_template_fidelity(before: dict[str, Any], after: dict[str, Any], authorized: dict[str, Any]) -> dict[str, Any]:
    violations: list[dict[str, Any]] = []
    before_sheets = {int(item["index"]): item for item in before["sheets"]}
    after_sheets = {int(item["index"]): item for item in after["sheets"]}
    for index in sorted(set(after_sheets) - set(before_sheets)):
        new = after_sheets[index]
        prototype = authorized.get("added_sheets", {}).get(new["title"])
        if prototype is None:
            violations.append({"kind": "sheet_added", "sheet_index": index, "sheet": new["title"]})
        else:
            violations.extend(_clone_presentation_violations(prototype, new))
    for index in sorted(set(before_sheets) - set(after_sheets)):
        violations.append({"kind": "sheet_removed", "sheet_index": index, "sheet": before_sheets[index]["title"]})

    for index in sorted(set(before_sheets) & set(after_sheets)):
        old = before_sheets[index]
        new = after_sheets[index]
        sheet_allowed = authorized.get("sheet_properties", {}).get(index, set())
        for field in ("title", "state"):
            if old[field] != new[field] and field not in sheet_allowed:
                violations.append({"kind": f"sheet_{field}", "sheet": old["title"], "before": old[field], "after": new[field]})
        structure_allowed = authorized.get("structure_properties", {}).get(index, set())
        for field in (
            "merged_cells",
            "images",
            "charts",
            "tables",
            "data_validations",
            "conditional_formatting",
            "freeze_panes",
            "auto_filter",
            "print_area",
            "print_title_rows",
            "print_title_cols",
            "row_dimensions",
            "column_dimensions",
            "sheet_properties",
            "sheet_format",
            "sheet_views",
            "page_margins",
            "page_setup",
            "sheet_protection",
        ):
            if old[field] == new[field]:
                continue
            expected = None
            if field == "tables" and authorized.get("table_refs"):
                expected = _expected_tables(old[field], authorized, index)
            elif field in {"data_validations", "conditional_formatting"} and authorized.get(
                "range_expansions", {}
            ).get(index):
                expected = _expected_ranged_collection(old[field], authorized, index)
                actual = _expected_ranged_collection(new[field], {"range_expansions": {}}, index)
                if expected == actual:
                    continue
            elif field == "row_dimensions" and authorized.get("new_row_prototypes"):
                expected = _expected_row_dimensions(old[field], authorized, index)
            elif field == "auto_filter" and index in authorized.get("auto_filter_refs", {}):
                expected = authorized["auto_filter_refs"][index]
            if expected is not None and expected == new[field]:
                continue
            if field not in structure_allowed:
                violations.append({"kind": field, "sheet": old["title"], "before": old[field], "after": new[field]})
        old_cells = old["cells"]
        new_cells = new["cells"]
        for coordinate in sorted(set(old_cells) | set(new_cells)):
            if old_cells.get(coordinate) == new_cells.get(coordinate):
                continue
            allowed = authorized.get("cells", {}).get((index, coordinate), set())
            old_cell = old_cells.get(coordinate)
            new_cell = new_cells.get(coordinate)
            if (index, coordinate) in authorized.get("new_cell_prototypes", {}):
                prototype_coordinate = authorized["new_cell_prototypes"][(index, coordinate)]
                prototype = old_cells.get(prototype_coordinate, {})
                differences = _cell_differences(prototype, new_cell or {}, VALUE_PROPERTIES | {"hyperlink"})
                if differences:
                    violations.append(
                        {"kind": "new_cell_not_from_prototype", "sheet": old["title"], "coordinate": coordinate, "properties": differences}
                    )
                continue
            differences = _cell_differences(old_cell or {}, new_cell or {}, allowed)
            if not differences:
                continue
            old_target = str((old_cell or {}).get("hyperlink") or "")
            obsolete_link_change = differences == ["hyperlink"] and any(
                target in old_target for target in authorized.get("obsolete_targets", set())
            )
            if not obsolete_link_change:
                violations.append(
                    {"kind": "unmapped_cell_property", "sheet": old["title"], "coordinate": coordinate, "properties": differences}
                )

    for field in ("workbook_properties", "workbook_security", "calculation", "workbook_views"):
        if before.get(field) != after.get(field):
            violations.append({"kind": field, "before": before.get(field), "after": after.get(field)})
    if not _defined_names_match(
        before.get("defined_names", []),
        after.get("defined_names", []),
        authorized,
    ):
        violations.append({"kind": "defined_names"})
    return {
        "status": "passed" if not violations else "failed",
        "snapshot_version": "2.0.0",
        "checked_value_only_authorization": True,
        "checked_complete_structural_fidelity": True,
        "checked_formulas_styles_comments_and_hyperlinks": True,
        "violations": violations,
    }


def _package_inventory(path: Path) -> dict[str, str]:
    try:
        with ZipFile(path) as archive:
            return {name: hashlib.sha256(archive.read(name)).hexdigest() for name in archive.namelist() if not name.endswith("/")}
    except (BadZipFile, OSError):
        return {}


def _protected_part(name: str) -> bool:
    lowered = name.casefold()
    return (
        lowered.endswith((".bin", ".png", ".jpg", ".jpeg", ".gif", ".bmp", ".emf", ".wmf"))
        or name.startswith("customXml/")
        or name.startswith("xl/embeddings/")
        or name.startswith("xl/activeX/")
    )


def add_package_fidelity(report: dict[str, Any], template: Path, output: Path) -> dict[str, Any]:
    result = {**report, "violations": list(report.get("violations", []))}
    result["template_extension"] = template.suffix.lower()
    result["output_extension"] = output.suffix.lower()
    before = _package_inventory(template)
    after = _package_inventory(output)
    removed = sorted(set(before) - set(after) - {"xl/calcChain.xml"})
    result["package_parts_before"] = len(before)
    result["package_parts_after"] = len(after)
    result["removed_package_parts"] = removed
    result["added_package_parts"] = sorted(set(after) - set(before))
    for name in removed:
        result["violations"].append({"kind": "package_part_removed", "part": name})
    altered_protected = sorted(name for name in set(before) & set(after) if _protected_part(name) and before[name] != after[name])
    result["altered_protected_parts"] = altered_protected
    for name in altered_protected:
        result["violations"].append({"kind": "protected_package_part_changed", "part": name})
    if template.suffix.lower() == ".xlsm" and output.suffix.lower() != ".xlsm":
        result["violations"].append({"kind": "macro_extension", "message": "An XLSM template must remain XLSM."})
    result["checked_package_part_presence"] = True
    result["checked_protected_binary_hashes"] = True
    result["rendered_layout_check"] = {
        "status": "not_run",
        "reason": "No deterministic workbook renderer was configured; structural and package gates remain mandatory.",
    }
    result["status"] = "passed" if not result["violations"] else "failed"
    return result
