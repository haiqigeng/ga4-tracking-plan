from __future__ import annotations

import copy
import json
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from jsonschema import Draft202012Validator
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

import generate_tracking_plan_workbook as workbook_generator
from adapt_tracking_plan_workbook import adapt
from build_analysis_context_seed import build_analysis_context_seed
from capture_interactive_journey import load_and_validate_spec
from check_official_sources import fetch_cached, semantic_errors
from delivery_artifacts import event_push_schema, expected_events_contract
from diff_tracking_plans import compare, load_plan
from discover_site_journeys import same_host
from discover_site_journeys_playwright import (
    build_auto_interaction_recipes,
    candidate_priority,
    discovery_round_stop_reason,
    material_unvisited_candidates,
    measurement_opportunity_hints,
    summarize_languages,
    summarize_measurement_evidence,
)
from discovery_contract import load_discovery_report, validate_discovery_bindings
from generate_tracking_plan_workbook import build_workbook
from import_tracking_plan_workbook import import_workbook
from inspect_tracking_plan_template import inspect
from maintenance_analysis import analyze_change_impact, detect_context_drift
from official_ga4_catalog import parse_catalog_html
from tracking_plan_model import load_json, parameter_reference_rows
from validate_analysis_context import validate_analysis_context
from validate_tracking_plan import validate_plan
from validate_tracking_plan_workbook import validate_workbook

EXAMPLE = ROOT / "references" / "example-tracking-plan.json"
ASSET = ROOT / "assets" / "default-tracking-plan.xlsx"


class TrackingPlanSkillTests(unittest.TestCase):
    def setUp(self) -> None:
        self.plan = load_json(EXAMPLE)

    def error_codes(self, plan: dict) -> set[str]:
        return {issue.code for issue in validate_plan(plan) if issue.severity == "error"}

    def warning_codes(self, plan: dict) -> set[str]:
        return {issue.code for issue in validate_plan(plan) if issue.severity == "warning"}

    def purchase_plan(self) -> dict:
        plan = copy.deepcopy(self.plan)
        catalog = json.loads((ROOT / "references" / "library-ga4-recommended-events.json").read_text(encoding="utf-8-sig"))
        record = next(item for item in catalog if item["event"] == "purchase")
        official = {(item["name"], item["scope"]): item for item in record["parameters"]}
        source = {
            "url": ("https://developers.google.com/analytics/devguides/collection/ga4/reference/events#purchase"),
            "section": "purchase",
            "wording_origin": "exact",
            "official_text": record["description"],
            "checked_date": "2026-07-31",
        }

        def parameter(
            name: str,
            scope: str,
            path: str,
            value,
            requirement: str,
            *,
            condition: str = "",
            allowed_values: list[str] | None = None,
        ) -> dict:
            row = official[(name, scope)]
            raw_type = str(row["type"]).casefold()
            parameter_type = "array" if raw_type.startswith("array") else ("string" if raw_type.startswith("string") else raw_type)
            result = {
                "name": name,
                "data_layer_path": path,
                "classification": "official",
                "scope": scope,
                "type": parameter_type,
                "requirement": requirement,
                "definition": row["description"],
                "value_rule": (
                    "Set value to the sum of price * quantity for all items; exclude shipping and tax."
                    if name == "value"
                    else f"Use the confirmed purchase {name}."
                ),
                "value_mode": (
                    "official_enum"
                    if name == "customer_type"
                    else (
                        "structured"
                        if parameter_type in {"array", "object"}
                        else (
                            "numeric"
                            if parameter_type in {"number", "integer"}
                            else ("technical_identifier" if name in {"currency", "transaction_id", "item_id"} else "authoritative_raw")
                        )
                    )
                ),
                "value_evidence_refs": [f"test_{scope}_{name}_domain"],
                "example": value,
                "source": "Confirmed order.",
                "destination": ("ga4_item_parameter" if scope == "item" else "ga4_event_parameter"),
                "official_source": {
                    **source,
                    "section": "purchase parameters",
                    "official_text": row["description"],
                },
            }
            if condition:
                result["condition"] = condition
            if allowed_values is not None:
                result["allowed_values"] = allowed_values
            return result

        items = [{"item_id": "SKU-123", "item_name": "Example item"}]
        purchase = {
            "event_name": "purchase",
            "classification": "official_ecommerce",
            "journey_ids": ["product_discovery"],
            "business_question": ("Which confirmed orders generate revenue, and are they from new or returning customers?"),
            "measurement_opportunity_ids": ["purchase_confirmation"],
            "definition": record["description"],
            "trigger": "Push once after the backend confirms and identifies the completed order.",
            "locations": [{"state": "Order confirmation"}],
            "parameters": [
                parameter("currency", "event", "ecommerce.currency", "EUR", "required"),
                parameter("value", "event", "ecommerce.value", 129.9, "required"),
                parameter(
                    "customer_type",
                    "event",
                    "ecommerce.customer_type",
                    "new",
                    "conditional",
                    condition=("Send when the confirmed order can be classified reliably as new or returning."),
                    allowed_values=["new", "returning"],
                ),
                parameter(
                    "transaction_id",
                    "event",
                    "ecommerce.transaction_id",
                    "T-123",
                    "required",
                ),
                parameter("items", "event", "ecommerce.items", items, "required"),
                parameter(
                    "item_id",
                    "item",
                    "ecommerce.items[].item_id",
                    "SKU-123",
                    "required",
                ),
                parameter(
                    "item_name",
                    "item",
                    "ecommerce.items[].item_name",
                    "Example item",
                    "optional",
                ),
            ],
            "data_layer": {
                "clear": ["ecommerce"],
                "push": {
                    "event": "purchase",
                    "ecommerce": {
                        "currency": "EUR",
                        "value": 129.9,
                        "customer_type": "new",
                        "transaction_id": "T-123",
                        "items": items,
                    },
                },
            },
            "official_source": source,
        }
        plan["events"] = [plan["events"][0], purchase]
        return plan

    def test_example_is_strictly_valid(self) -> None:
        self.assertEqual(validate_plan(self.plan), [])

    def test_non_context_event_without_business_question_is_flagged(self) -> None:
        plan = copy.deepcopy(self.plan)
        plan["events"][1].pop("business_question")
        self.assertIn("EVENT_PURPOSE_MISSING", self.error_codes(plan))

    def test_context_push_does_not_require_a_business_question(self) -> None:
        plan = copy.deepcopy(self.plan)
        plan["events"] = [plan["events"][0]]
        self.assertNotIn("EVENT_PURPOSE_MISSING", self.warning_codes(plan))

    def test_exact_trigger_overlap_in_a_shared_journey_is_flagged(self) -> None:
        plan = copy.deepcopy(self.plan)
        duplicate = copy.deepcopy(plan["events"][2])
        duplicate["event_name"] = "quote_start_duplicate"
        duplicate["business_question"] = "Which visitors reach a second quote-start signal?"
        duplicate["definition"] = "Indicates that a second quote-start signal was recorded."
        duplicate["custom_decision"] = {
            "business_need": "Distinguish a second quote-start signal.",
            "official_candidate": "No official event represents this signal.",
            "why_not_fit": "The existing quote start expresses a different purpose.",
        }
        duplicate["data_layer"]["push"]["event"] = "quote_start_duplicate"
        plan["events"].append(duplicate)
        self.assertIn(
            "POTENTIAL_DUPLICATE_EVENT_TRIGGER",
            self.warning_codes(plan),
        )

    def test_official_catalog_parser_preserves_event_and_item_scope(self) -> None:
        html = """
        <h2 id="sales">Sales</h2>
        <h3 id="sample_event"><code>sample_event</code></h3>
        <p>Official event wording.</p>
        <table><tr><th>Name</th><th>Type</th><th>Required</th><th>Example</th><th>Description</th></tr>
        <tr><td>value</td><td>number</td><td>No</td><td>1</td><td>The value.</td></tr></table>
        <h4>Item parameters</h4>
        <table><tr><th>Name</th><th>Type</th><th>Required</th><th>Example</th><th>Description</th></tr>
        <tr><td>item_id</td><td>string</td><td>Yes*</td><td>SKU</td><td>The item ID.</td></tr></table>
        """
        records = parse_catalog_html(html)
        self.assertEqual(records[0]["description"], "Official event wording.")
        self.assertEqual(
            [(item["name"], item["scope"]) for item in records[0]["parameters"]],
            [("value", "event"), ("item_id", "item")],
        )

    def test_faithful_translation_still_preserves_exact_official_source_text(self) -> None:
        plan = copy.deepcopy(self.plan)
        event = plan["events"][1]
        event["official_source"]["wording_origin"] = "faithful_translation"
        event["definition"] = "Cet événement indique qu'un contenu a été présenté à l'utilisateur."
        self.assertNotIn("OFFICIAL_EVENT_SOURCE_TEXT", self.error_codes(plan))
        event["official_source"]["official_text"] = "Invented source wording."
        self.assertIn("OFFICIAL_EVENT_SOURCE_TEXT", self.error_codes(plan))

    def test_ecommerce_index_rule_must_be_zero_based(self) -> None:
        plan = copy.deepcopy(self.plan)
        event = plan["events"][1]
        catalog = json.loads((ROOT / "references" / "library-ga4-recommended-events.json").read_text(encoding="utf-8-sig"))
        record = next(item for item in catalog if item["event"] == "view_item")
        index_row = next(item for item in record["parameters"] if item["name"] == "index" and item["scope"] == "item")
        event["data_layer"]["push"]["ecommerce"]["items"][0]["index"] = 0
        event["parameters"].append(
            {
                "name": "index",
                "data_layer_path": "ecommerce.items[].index",
                "classification": "official",
                "scope": "item",
                "type": "integer",
                "requirement": "optional",
                "definition": index_row["description"],
                "value_rule": "Use a zero-based position; the first item is index 0.",
                "value_mode": "numeric",
                "value_evidence_refs": ["item_indexes"],
                "example": 0,
                "source": "Rendered item order.",
                "destination": "ga4_item_parameter",
                "official_source": {
                    "url": ("https://developers.google.com/analytics/devguides/collection/ga4/reference/events#view_item"),
                    "section": "items parameter",
                    "wording_origin": "exact",
                    "official_text": index_row["description"],
                    "checked_date": "2026-08-01",
                },
            }
        )
        self.assertNotIn("INDEX_ZERO_BASE_MISSING", self.error_codes(plan))
        event["parameters"][-1]["value_rule"] = "Start the position at 1 for the first item."
        codes = self.error_codes(plan)
        self.assertIn("INDEX_ZERO_BASE_MISSING", codes)
        self.assertIn("INDEX_ONE_BASED", codes)

    def test_ecommerce_value_rule_accepts_precise_french_semantics(self) -> None:
        plan = copy.deepcopy(self.plan)
        value = next(parameter for parameter in plan["events"][1]["parameters"] if parameter["name"] == "value")
        value["value_rule"] = "Utiliser la somme de prix * quantite pour les articles, hors livraison et taxe."
        self.assertNotIn("ECOMMERCE_VALUE_RULE_INCOMPLETE", self.error_codes(plan))

    def test_controlled_values_follow_workbook_language_and_ascii_snake_case(self) -> None:
        plan = copy.deepcopy(self.plan)
        parameter = plan["events"][0]["parameters"][0]
        parameter["allowed_values"] = ["fiche_produit"]
        parameter["example"] = "fiche_produit"
        parameter["value_language"] = "fr"
        self.assertIn("CONTROLLED_VALUE_LANGUAGE", self.error_codes(plan))
        parameter["value_language"] = "en"
        parameter["allowed_values"] = ["Product Detail"]
        parameter["example"] = "Product Detail"
        self.assertIn("CONTROLLED_VALUE_FORMAT", self.error_codes(plan))

    def test_manual_only_schema_rejects_enhanced_measurement_classification(self) -> None:
        plan = copy.deepcopy(self.plan)
        plan["events"][0]["classification"] = "enhanced_measurement"
        self.assertIn("SCHEMA", self.error_codes(plan))

    def test_finite_value_domains_stop_at_fifty(self) -> None:
        plan = copy.deepcopy(self.plan)
        plan["events"][0]["parameters"][0]["allowed_values"] = [f"value_{index}" for index in range(51)]
        self.assertIn("SCHEMA", self.error_codes(plan))

    def test_custom_event_cannot_relabel_an_official_event(self) -> None:
        plan = copy.deepcopy(self.plan)
        plan["events"][1]["classification"] = "custom"
        plan["events"][1]["custom_decision"] = {
            "business_need": "Measure product detail views.",
            "official_candidate": "view_item",
            "why_not_fit": "The official event would fit.",
        }
        self.assertIn("CUSTOM_EVENT_IS_OFFICIAL", self.error_codes(plan))

    def test_every_pushed_field_requires_an_event_binding(self) -> None:
        plan = copy.deepcopy(self.plan)
        plan["events"][2]["data_layer"]["push"]["event_data"]["unbound_field"] = "x"
        self.assertIn("UNBOUND_DATALAYER_FIELDS", self.error_codes(plan))

    def test_workbook_is_lean_quoted_and_maintenance_ready(self) -> None:
        workbook = build_workbook(self.plan)
        visible = [sheet.title for sheet in workbook.worksheets if sheet.sheet_state == "visible"]
        self.assertEqual(
            visible,
            [
                "Guide",
                "Event Matrix",
                "Parameter Reference",
                "core_data",
                "view_item",
                "begin_quote",
            ],
        )
        self.assertEqual(
            workbook["__tracking_plan_model"].sheet_state,
            "veryHidden",
        )
        matrix_headers = [str(workbook["Event Matrix"].cell(4, column).value or "") for column in range(1, 3)]
        parameter_headers = {str(workbook["Parameter Reference"].cell(4, column).value or "") for column in range(1, 8)}
        forbidden = {
            "Availability",
            "Data owner",
            "Registered in GA4",
            "Privacy",
            "Display name",
            "Agent reasoning",
        }
        self.assertEqual(matrix_headers, ["Event", "Definition"])
        self.assertTrue(parameter_headers.isdisjoint(forbidden))
        guide_headers = {str(workbook["Guide"].cell(15, column).value or "") for column in range(1, 5)}
        self.assertNotIn("Status", guide_headers)
        self.assertIsNone(workbook["Event Matrix"]["C4"].value)
        event_headers = [str(workbook["core_data"].cell(11, column).value or "") for column in range(1, 8)]
        self.assertEqual(
            event_headers,
            [
                "Variable",
                "Scope",
                "Type",
                "Requirement",
                "Definition",
                "Rule",
                "Possible values or examples",
            ],
        )
        self.assertNotIn("Condition", event_headers)
        self.assertNotIn("dataLayer path / source", event_headers)
        self.assertTrue(workbook["core_data"].row_dimensions[4].hidden)
        code = str(workbook["view_item"]["A21"].value)
        self.assertIn('"event": "view_item"', code)
        self.assertIn('"item_color": "white"', code)
        core_code = str(workbook["core_data"]["A17"].value)
        self.assertIn('"page": {', core_code)
        self.assertIn('"user": {', core_code)
        self.assertIn('"user_id": "customer_12345"', core_code)
        visible_text = "\n".join(
            str(cell.value or "") for sheet in workbook.worksheets if sheet.sheet_state == "visible" for row in sheet.iter_rows() for cell in row
        )
        self.assertNotIn(self.plan["events"][1]["business_question"], visible_text)
        self.assertNotIn(self.plan["events"][2]["business_question"], visible_text)

    def test_french_workbook_localizes_human_labels_and_requirements(self) -> None:
        plan = copy.deepcopy(self.plan)
        plan["document"]["language"] = "fr"
        workbook = build_workbook(plan)
        self.assertIn("Valeurs des variables", workbook.sheetnames)
        self.assertEqual(
            [workbook["Event Matrix"].cell(4, column).value for column in range(1, 3)],
            ["Événement", "Définition"],
        )
        self.assertEqual(workbook["view_item"]["D14"].value, "obligatoire")
        self.assertEqual(workbook["view_item"]["D12"].value, "conditionnel")

    def test_generated_workbook_round_trips_losslessly(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "plan.xlsx"
            build_workbook(self.plan).save(path)
            imported = import_workbook(path)
        self.assertEqual(imported, self.plan)

    def test_formula_like_text_stays_text_and_round_trips(self) -> None:
        plan = copy.deepcopy(self.plan)
        plan["events"][2]["definition"] = '=HYPERLINK("https://invalid.example","x")'
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "formula-safe.xlsx"
            build_workbook(plan).save(path)
            reopened = load_workbook(path, data_only=False)
            imported = import_workbook(path)
        self.assertEqual(reopened["begin_quote"]["B6"].data_type, "s")
        self.assertEqual(
            reopened["begin_quote"]["B6"].value,
            '=HYPERLINK("https://invalid.example","x")',
        )
        self.assertEqual(imported, plan)

    def test_embedded_model_chunk_starting_with_equals_stays_text(self) -> None:
        plan = copy.deepcopy(self.plan)
        plan["document"]["notes"] = "=model-chunk-marker"
        payload = json.dumps(plan, ensure_ascii=False, separators=(",", ":"))
        marker_index = payload.index("=")
        original_limit = workbook_generator.MODEL_CELL_LIMIT
        try:
            workbook_generator.MODEL_CELL_LIMIT = marker_index
            with tempfile.TemporaryDirectory() as directory:
                path = Path(directory) / "chunk-safe.xlsx"
                build_workbook(plan).save(path)
                reopened = load_workbook(path, data_only=False)
                imported = import_workbook(path)
        finally:
            workbook_generator.MODEL_CELL_LIMIT = original_limit
        self.assertTrue(str(reopened["__tracking_plan_model"]["A3"].value).startswith("="))
        self.assertEqual(reopened["__tracking_plan_model"]["A3"].data_type, "s")
        self.assertEqual(imported, plan)

    def test_visible_recovery_works_for_english_and_revalidates(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "visible-only.xlsx"
            workbook = build_workbook(self.plan)
            del workbook["__tracking_plan_model"]
            workbook.save(path)
            recovered = import_workbook(path, allow_visible_recovery=True)
        errors = [item for item in validate_plan(recovered) if item.severity == "error"]
        self.assertEqual(recovered["document"]["language"], "en")
        self.assertNotIn("EVENT_PURPOSE_MISSING", {item.code for item in errors})
        view_item = next(item for item in recovered["events"] if item["event_name"] == "view_item")
        self.assertIn("official_source", view_item)
        self.assertTrue(view_item["business_question"])
        self.assertTrue(view_item["measurement_opportunity_ids"])
        self.assertTrue(all("official_source" in parameter for parameter in view_item["parameters"] if parameter["classification"] == "official"))

    def test_diff_can_opt_into_visible_workbook_recovery(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "visible-only.xlsx"
            workbook = build_workbook(self.plan)
            del workbook["__tracking_plan_model"]
            workbook.save(path)
            recovered = load_plan(path, allow_visible_recovery=True)
        self.assertEqual(
            [event["event_name"] for event in recovered["events"]],
            ["core_data", "view_item", "begin_quote"],
        )

    def test_shrinking_adaptation_hides_obsolete_tabs_and_clears_links(self) -> None:
        updated = copy.deepcopy(self.plan)
        updated["events"] = updated["events"][1:]
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "source.xlsx"
            build_workbook(self.plan).save(source)
            mapping = inspect(source)
            workbook = adapt(updated, source, mapping)
        self.assertEqual(workbook["core_data"].sheet_state, "hidden")
        self.assertIsNone(workbook["Guide"]["B11"].hyperlink)
        self.assertIsNone(workbook["Event Matrix"]["A7"].value)
        self.assertIsNone(workbook["Event Matrix"]["A7"].hyperlink)
        self.assertEqual(workbook["Event Matrix"]["A5"].value, "view_item")
        self.assertIn(
            "view_item",
            str(workbook["Event Matrix"]["A5"].hyperlink.target),
        )

    def test_legitimate_contextual_wording_is_not_rejected_as_filler(self) -> None:
        plan = copy.deepcopy(self.plan)
        plan["events"][2]["parameters"][0]["definition"] = "Identifies the value associated with the selected loyalty tier."
        plan["events"][2]["trigger"] = "Push after the booking succeeds, when applicable shipping options have already been resolved."
        self.assertNotIn("PARAMETER_DEFINITION_GENERIC", self.error_codes(plan))
        self.assertNotIn("TRIGGER_GENERIC", self.error_codes(plan))

    def test_exact_generic_filler_is_still_rejected(self) -> None:
        plan = copy.deepcopy(self.plan)
        plan["events"][2]["parameters"][0]["definition"] = "When applicable."
        self.assertIn("PARAMETER_DEFINITION_GENERIC", self.error_codes(plan))

    def test_unsupported_workbook_language_is_explicitly_rejected(self) -> None:
        plan = copy.deepcopy(self.plan)
        plan["document"]["language"] = "de"
        self.assertIn("UNSUPPORTED_WORKBOOK_LANGUAGE", self.error_codes(plan))

    def test_page_and_user_context_cannot_be_split(self) -> None:
        plan = copy.deepcopy(self.plan)
        core = plan["events"][0]
        user_parameters = [parameter for parameter in core["parameters"] if str(parameter["data_layer_path"]).startswith("user.")]
        core["parameters"] = [parameter for parameter in core["parameters"] if not str(parameter["data_layer_path"]).startswith("user.")]
        core["data_layer"]["push"].pop("user")
        plan["events"].insert(
            1,
            {
                "event_name": "user_context",
                "classification": "context",
                "journey_ids": ["product_discovery", "quote_request"],
                "definition": "Provides reusable authenticated user state.",
                "trigger": "Push after user state is available on every page and route.",
                "locations": [{"state": "All pages"}],
                "parameters": user_parameters,
                "data_layer": {
                    "push": {
                        "user": {
                            "login_status": "logged_in",
                            "user_id": "customer_12345",
                        }
                    }
                },
            },
        )
        self.assertIn("CORE_CONTEXT_SPLIT", self.error_codes(plan))

    def test_user_id_must_use_the_official_configuration_destination(self) -> None:
        plan = copy.deepcopy(self.plan)
        user_id = next(parameter for parameter in plan["events"][0]["parameters"] if parameter["name"] == "user_id")
        user_id["destination"] = "ga4_user_property"
        self.assertIn("USER_ID_DESTINATION", self.error_codes(plan))

    def test_authenticated_plan_requires_user_id_in_core_context(self) -> None:
        plan = copy.deepcopy(self.plan)
        core = plan["events"][0]
        core["parameters"] = [parameter for parameter in core["parameters"] if parameter["name"] != "user_id"]
        core["data_layer"]["push"]["user"].pop("user_id")
        catalog = json.loads((ROOT / "references" / "library-ga4-recommended-events.json").read_text(encoding="utf-8-sig"))
        login = next(item for item in catalog if item["event"] == "login")
        plan["events"].append(
            {
                "event_name": "login",
                "classification": "official",
                "journey_ids": ["quote_request"],
                "definition": login["description"],
                "trigger": "Push once after the authentication backend confirms a successful login.",
                "locations": [{"state": "Successful login"}],
                "parameters": [],
                "data_layer": {"push": {"event": "login"}},
                "official_source": {
                    "url": ("https://developers.google.com/analytics/devguides/collection/ga4/reference/events#login"),
                    "section": "login",
                    "wording_origin": "exact",
                    "checked_date": "2026-07-31",
                },
            }
        )
        self.assertIn("AUTHENTICATION_USER_ID_MISSING", self.error_codes(plan))

    def test_purchase_requires_conditional_customer_type_and_values(self) -> None:
        plan = self.purchase_plan()
        self.assertEqual(validate_plan(plan), [])
        plan["events"][1]["parameters"] = [parameter for parameter in plan["events"][1]["parameters"] if parameter["name"] != "customer_type"]
        plan["events"][1]["data_layer"]["push"]["ecommerce"].pop("customer_type")
        self.assertIn("PURCHASE_CUSTOMER_TYPE_MISSING", self.error_codes(plan))

    def test_event_parameter_collection_limit_is_enforced(self) -> None:
        plan = copy.deepcopy(self.plan)
        event = plan["events"][2]
        for index in range(24):
            name = f"extra_{index}"
            event["parameters"].append(
                {
                    "name": name,
                    "data_layer_path": f"event_data.{name}",
                    "classification": "custom",
                    "scope": "event",
                    "type": "string",
                    "requirement": "optional",
                    "definition": f"Identifies the normalized extra field {index}.",
                    "value_rule": f"Use the normalized extra value {index}.",
                    "example": f"value_{index}",
                    "source": "Synthetic test source.",
                    "destination": "ga4_event_parameter",
                    "custom_decision": {
                        "business_need": f"Compare the extra business dimension {index}.",
                        "official_candidate": "No prescribed begin_quote parameter represents this dimension.",
                        "why_not_fit": f"The dimension {index} is business-specific.",
                    },
                }
            )
            event["data_layer"]["push"]["event_data"][name] = f"value_{index}"
        self.assertIn("EVENT_PARAMETER_COLLECTION_LIMIT", self.error_codes(plan))

    def test_reserved_automatic_web_event_is_rejected(self) -> None:
        plan = copy.deepcopy(self.plan)
        event = plan["events"][2]
        event["event_name"] = "page_view"
        event["data_layer"]["push"]["event"] = "page_view"
        self.assertIn("RESERVED_OR_AUTOMATIC_EVENT", self.error_codes(plan))

    def test_scope_and_destination_must_agree(self) -> None:
        plan = copy.deepcopy(self.plan)
        plan["events"][2]["parameters"][0]["destination"] = "ga4_item_parameter"
        self.assertIn("SCOPE_DESTINATION_MISMATCH", self.error_codes(plan))

    def test_zero_event_catalog_parse_is_reported_as_parser_failure(self) -> None:
        errors = semantic_errors(
            self.plan,
            {("https://developers.google.com/analytics/devguides/collection/ga4/reference/events"): {"content": "<html><body>No event tables</body></html>"}},
        )
        self.assertTrue(any("parsed zero events" in error for error in errors))

    def test_www_and_bare_domain_are_same_site(self) -> None:
        self.assertTrue(same_host("https://www.example.com/a", "https://example.com/"))

    def test_measurement_evidence_summary_deduplicates_identifiers(self) -> None:
        summary = summarize_measurement_evidence(
            [
                {
                    "measurement_evidence": {
                        "data_layer_present": True,
                        "data_layer_push_count": 3,
                        "gtm_container_ids": ["GTM-ABC", "GTM-ABC"],
                        "google_tag_ids": ["GT-ONE"],
                        "ga4_measurement_ids": ["G-123456"],
                    }
                },
                {
                    "measurement_evidence": {
                        "data_layer_present": False,
                        "data_layer_push_count": 0,
                        "gtm_container_ids": ["GTM-XYZ"],
                        "google_tag_ids": [],
                        "ga4_measurement_ids": ["G-123456"],
                    }
                },
            ]
        )
        self.assertEqual(summary["pages_with_data_layer"], 1)
        self.assertEqual(summary["observed_data_layer_push_count"], 3)
        self.assertEqual(summary["gtm_container_ids"], ["GTM-ABC", "GTM-XYZ"])

    def test_runtime_dependencies_are_declared(self) -> None:
        requirements = (ROOT / "requirements.txt").read_text(encoding="utf-8")
        for package in ("jsonschema", "openpyxl", "Pillow", "playwright"):
            self.assertIn(package, requirements)

    def test_semantic_diff_reports_trigger_and_values(self) -> None:
        updated = copy.deepcopy(self.plan)
        updated["document"]["version"] = "1.1"
        updated["events"][2]["trigger"] += " Fire only once per form instance."
        updated["events"][2]["parameters"][1]["allowed_values"].append("solar")
        result = compare(self.plan, updated)
        entities = {(item["entity"], item["key"]) for item in result["changes"]}
        self.assertIn(("trigger", "begin_quote:trigger"), entities)
        self.assertIn(
            (
                "value_domain",
                "begin_quote:project_type|event|event_data.project_type",
            ),
            entities,
        )

    def test_default_asset_has_semantic_regions(self) -> None:
        result = inspect(ASSET)
        self.assertTrue(result["regions"]["event_matrix"])
        self.assertTrue(result["regions"]["parameter_reference"])
        self.assertTrue(result["regions"]["event_tabs"])
        release = json.loads((ROOT / "release.json").read_text(encoding="utf-8-sig"))
        workbook = load_workbook(ASSET, read_only=True)
        self.assertEqual(
            workbook.properties.description,
            f"Default workbook asset version {release['version']}",
        )
        workbook.close()

    def test_supplied_template_adaptation_uses_semantic_regions(self) -> None:
        updated = copy.deepcopy(self.plan)
        updated["events"][2]["trigger"] += " Fire only once per form instance."
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "source.xlsx"
            output = Path(directory) / "adapted.xlsx"
            build_workbook(self.plan).save(source)
            mapping = inspect(source)
            workbook = adapt(updated, source, mapping)
            workbook.save(output)
            reopened = load_workbook(output, data_only=False)
            imported = import_workbook(output)
        self.assertIn(
            "Fire only once per form instance.",
            str(reopened["begin_quote"]["B7"].value),
        )
        self.assertIn(
            '"event": "begin_quote"',
            str(reopened["begin_quote"]["A16"].value),
        )
        self.assertEqual(imported, updated)

    def test_skill_has_one_adaptive_workflow_not_scope_tiers(self) -> None:
        text = (ROOT / "SKILL.md").read_text(encoding="utf-8")
        self.assertIn("Use one adaptive workflow and one quality standard.", text)
        self.assertNotIn("## Scope Tiers", text)
        self.assertNotIn('Tier 1 — "Quick Plan"', text)
        self.assertNotIn("event-count-based execution mode", text)

    def test_official_ecommerce_rejects_event_wrapper(self) -> None:
        plan = copy.deepcopy(self.plan)
        event = plan["events"][1]
        event["data_layer"]["push"]["event_data"] = event["data_layer"]["push"].pop("ecommerce")
        for parameter in event["parameters"]:
            parameter["data_layer_path"] = str(parameter["data_layer_path"]).replace("ecommerce.", "event_data.")
        codes = self.error_codes(plan)
        self.assertIn("ECOMMERCE_WRAPPER_MISSING", codes)
        self.assertIn("DATALAYER_WRAPPER_MISMATCH", codes)

    def test_adapted_datalayer_convention_is_valid(self) -> None:
        plan = copy.deepcopy(self.plan)
        plan["data_layer_convention"]["origin"] = "adapted"
        self.assertEqual(validate_plan(plan), [])

    def test_parameter_reference_is_truly_deduplicated_by_name_and_scope(self) -> None:
        plan = copy.deepcopy(self.plan)
        duplicate = copy.deepcopy(plan["events"][0]["parameters"][0])
        duplicate["data_layer_path"] = "event_data.page_template"
        plan["events"][2]["parameters"].append(duplicate)
        rows = [row for row in parameter_reference_rows(plan) if row["name"] == "page_template" and row["scope"] == "implementation"]
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["events"], ["core_data", "begin_quote"])

    def test_analysis_context_is_a_real_delivery_gate(self) -> None:
        context = load_json(ROOT / "references" / "example-analysis-context.json")
        self.assertEqual(
            validate_analysis_context(context, self.plan, delivery=True),
            [],
        )
        context["value_domains"][0]["values"] = ["home"]
        codes = {item.code for item in validate_analysis_context(context, self.plan, delivery=True)}
        self.assertIn("FINITE_VALUE_DOMAIN_MISMATCH", codes)

    def test_material_interaction_opportunity_cannot_disappear_at_delivery(self) -> None:
        context = load_json(ROOT / "references" / "example-analysis-context.json")
        context["measurement_opportunities"][0]["decision"] = "unresolved"
        context["measurement_opportunities"][0]["event_names"] = []
        codes = {item.code for item in validate_analysis_context(context, self.plan, delivery=True)}
        self.assertIn("MATERIAL_OPPORTUNITY_UNRESOLVED", codes)

    def test_every_non_context_event_must_map_back_to_an_opportunity(self) -> None:
        context = load_json(ROOT / "references" / "example-analysis-context.json")
        opportunity = next(item for item in context["measurement_opportunities"] if "begin_quote" in item["event_names"])
        opportunity["decision"] = "exclude"
        opportunity["decision_reason"] = "Synthetic exclusion used to test backlink closure."
        opportunity["event_names"] = []
        codes = {item.code for item in validate_analysis_context(context, self.plan, delivery=True)}
        self.assertIn("EVENT_WITHOUT_MEASUREMENT_OPPORTUNITY", codes)

    def test_visible_workbook_edits_are_not_silently_ignored(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "maintained.xlsx"
            build_workbook(self.plan).save(path)
            workbook = load_workbook(path)
            workbook["begin_quote"]["B7"] = "Push once after the visitor opens the quote form."
            workbook.save(path)
            with self.assertRaisesRegex(ValueError, "visible workbook has changed"):
                import_workbook(path)
            reconciled = import_workbook(path, reconcile_visible_edits=True)
        event = next(item for item in reconciled["events"] if item["event_name"] == "begin_quote")
        self.assertEqual(
            event["trigger"],
            "Push once after the visitor opens the quote form.",
        )

    def test_structural_visible_edits_require_canonical_json(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "structural.xlsx"
            build_workbook(self.plan).save(path)
            workbook = load_workbook(path)
            workbook["begin_quote"]["B5"] = "Different journey"
            workbook.save(path)
            with self.assertRaisesRegex(ValueError, "unsupported structural"):
                import_workbook(path, reconcile_visible_edits=True)

    def test_discovery_priority_surfaces_unseen_material_journeys(self) -> None:
        root = "https://example.com/"
        quote = {"url": "https://example.com/mon-projet/devis", "text": "Devis", "source": "sitemap"}
        article = {"url": "https://example.com/blog/article", "text": "Article", "source": "sitemap"}
        self.assertGreater(
            candidate_priority(quote, root, {"homepage"}),
            candidate_priority(article, root, {"homepage"}),
        )
        gaps = material_unvisited_candidates([quote, article], root, {"homepage"})
        self.assertIn(quote["url"], {item["url"] for item in gaps})

    def test_repeated_product_pages_do_not_outrank_an_unseen_journey(self) -> None:
        root = "https://example.com/"
        product = {
            "url": "https://example.com/products/repeated-product",
            "text": "Repeated product",
            "source": "sitemap",
        }
        catalogue = {
            "url": "https://example.com/request-catalogue",
            "text": "Request a catalogue",
            "source": "sitemap",
        }
        observed_templates = {"homepage": 1, "product_detail": 4}
        observed_families = {"homepage", "product_detail"}
        self.assertGreater(
            candidate_priority(
                catalogue,
                root,
                observed_templates,
                observed_families,
            ),
            candidate_priority(
                product,
                root,
                observed_templates,
                observed_families,
            ),
        )

    def test_rendered_interactions_seed_measurement_opportunity_review(self) -> None:
        hints = measurement_opportunity_hints(
            [
                {
                    "url": "https://example.com/products",
                    "template": "listing",
                    "buttons": ["Filter", "Sort"],
                    "links": [],
                    "interactive_controls": [
                        {
                            "label": "Size filter",
                            "name": "size",
                            "option_values": ["small", "large"],
                        }
                    ],
                    "rendered_structure_sha256": "1" * 64,
                }
            ]
        )
        hint_keys = {item["hint_key"] for item in hints}
        self.assertIn("item_list_discovery", hint_keys)
        self.assertIn("filter_and_sort_usage", hint_keys)
        self.assertTrue(all(item["requires_interactive_review"] for item in hints))
        self.assertEqual(
            next(item["materiality"] for item in hints if item["hint_key"] == "item_list_discovery"),
            "material",
        )
        self.assertEqual(
            next(item["materiality"] for item in hints if item["hint_key"] == "filter_and_sort_usage"),
            "candidate",
        )

    def test_targeted_discovery_continues_until_material_families_close(self) -> None:
        self.assertEqual(
            discovery_round_stop_reason(3, 1, 3, 20),
            "continue_targeted_discovery",
        )
        self.assertEqual(
            discovery_round_stop_reason(0, 2, 3, 20),
            "material_coverage_complete",
        )
        self.assertEqual(
            discovery_round_stop_reason(2, 3, 3, 20),
            "max_rounds_reached",
        )

    def test_safe_interaction_recipes_are_generated_but_payment_is_excluded(self) -> None:
        pages = [
            {
                "url": "https://example.com/request-a-quote",
                "template": "lead_form",
                "forms": [
                    {
                        "selector": "#quote",
                        "action": "https://example.com/request-a-quote",
                        "fields": [
                            {
                                "selector": "input[name=email]",
                                "name": "email",
                                "type": "email",
                            }
                        ],
                        "submit_controls": [{"selector": "button", "label": "Send request"}],
                    }
                ],
            },
            {
                "url": "https://example.com/payment",
                "template": "lead_form",
                "forms": [
                    {
                        "selector": "#payment",
                        "action": "https://example.com/pay",
                        "fields": [],
                        "submit_controls": [{"selector": "button", "label": "Pay now"}],
                    }
                ],
            },
        ]
        recipes = build_auto_interaction_recipes(pages)
        self.assertEqual([item["start_url"] for item in recipes], [pages[0]["url"]])
        self.assertEqual(recipes[0]["fields"][0]["kind"], "email")

    def test_website_language_summary_is_evidence_backed(self) -> None:
        summary = summarize_languages(
            [
                {"url": "https://example.fr/", "language": "fr-FR"},
                {"url": "https://example.fr/devis", "language": "fr"},
            ],
            "https://example.fr/",
        )
        self.assertEqual(summary["primary_language"], "fr")
        self.assertEqual(summary["observed_languages"], ["fr", "fr-FR"])

    def test_interactive_spec_requires_explicit_non_transactional_submission(self) -> None:
        invalid = {
            "journey_id": "quote",
            "root_url": "https://example.com/",
            "actions": [{"action": "submit", "selector": "button[type=submit]"}],
        }
        valid = {
            **invalid,
            "allow_form_submission": True,
            "submission_kind": "lead",
        }
        with tempfile.TemporaryDirectory() as directory:
            invalid_path = Path(directory) / "invalid.json"
            valid_path = Path(directory) / "valid.json"
            invalid_path.write_text(json.dumps(invalid), encoding="utf-8")
            valid_path.write_text(json.dumps(valid), encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "specification is invalid"):
                load_and_validate_spec(invalid_path)
            self.assertEqual(load_and_validate_spec(valid_path), valid)

    def test_event_push_schemas_reject_wrong_wrappers(self) -> None:
        event = self.plan["events"][1]
        schema = event_push_schema(event)
        self.assertEqual(
            list(Draft202012Validator(schema).iter_errors(event["data_layer"]["push"])),
            [],
        )
        wrong = copy.deepcopy(event["data_layer"]["push"])
        wrong["event_data"] = wrong.pop("ecommerce")
        self.assertTrue(list(Draft202012Validator(schema).iter_errors(wrong)))

    def test_expected_events_contract_has_one_runtime_schema_per_event(self) -> None:
        contract = expected_events_contract(self.plan)
        self.assertEqual(len(contract["events"]), len(self.plan["events"]))
        self.assertEqual(
            {item["push_schema"] for item in contract["events"]},
            {f"schemas/{event['event_name']}.schema.json" for event in self.plan["events"]},
        )

    def test_rendered_workbook_gate_matches_canonical_plan(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "plan.xlsx"
            build_workbook(self.plan).save(path)
            self.assertEqual(validate_workbook(path, self.plan), [])

    def test_semantic_diff_covers_convention_purpose_and_path(self) -> None:
        updated = copy.deepcopy(self.plan)
        updated["data_layer_convention"]["origin"] = "adapted"
        updated["events"][2]["business_question"] += " By project type?"
        parameter = updated["events"][2]["parameters"][0]
        old_path = parameter["data_layer_path"]
        parameter["data_layer_path"] = "event_data.form_identifier"
        updated["events"][2]["data_layer"]["push"]["event_data"]["form_identifier"] = updated["events"][2]["data_layer"]["push"]["event_data"].pop(
            old_path.rsplit(".", 1)[-1]
        )
        result = compare(self.plan, updated)
        entities = {item["entity"] for item in result["changes"]}
        self.assertIn("data_layer_convention", entities)
        self.assertIn("business_question", entities)
        self.assertTrue(
            any(item["entity"] == "parameter" and item.get("after", {}).get("data_layer_path") == "event_data.form_identifier" for item in result["changes"])
        )

    def test_official_source_cache_avoids_duplicate_network_fetches(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            cache = Path(directory)
            with patch(
                "check_official_sources.fetch_with_retry",
                return_value=(200, "https://developers.google.com/test", "<html>ok</html>"),
            ) as mocked:
                first = fetch_cached(
                    "https://developers.google.com/test",
                    cache,
                    24,
                )
                second = fetch_cached(
                    "https://developers.google.com/test",
                    cache,
                    24,
                )
        self.assertEqual(mocked.call_count, 1)
        self.assertEqual(first["response_source"], "live")
        self.assertEqual(second["response_source"], "cache")

    def test_template_adaptation_preserves_unmapped_human_content(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "source.xlsx"
            workbook = build_workbook(self.plan)
            workbook["Guide"]["G30"] = "Human template instruction"
            workbook.save(source)
            mapping = inspect(source)
            adapted = adapt(self.plan, source, mapping)
        self.assertEqual(adapted["Guide"]["G30"].value, "Human template instruction")
        self.assertEqual(adapted._ga4_template_fidelity_report["status"], "passed")

    def test_template_event_tabs_are_not_arbitrarily_repurposed(self) -> None:
        updated = copy.deepcopy(self.plan)
        updated["events"][0]["event_name"] = "replacement_context"
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "source.xlsx"
            build_workbook(self.plan).save(source)
            mapping = inspect(source)
            with self.assertRaisesRegex(ValueError, "no mapped event tab"):
                adapt(updated, source, mapping)

    def test_context_drift_targets_affected_events_without_mutating_plan(self) -> None:
        before = load_json(ROOT / "references" / "example-analysis-context.json")
        after = copy.deepcopy(before)
        domain = next(item for item in after["value_domains"] if item["domain_id"] == "project_types")
        domain["values"].append("solar")
        report = detect_context_drift(before, after, self.plan)
        self.assertEqual(report["status"], "review_required")
        self.assertEqual(report["affected_events"], ["begin_quote"])
        self.assertNotIn("solar", self.plan["events"][2]["parameters"][1]["allowed_values"])

    def test_discovery_context_plan_closure_is_hash_bound(self) -> None:
        report_path = ROOT / "references" / "example-discovery-report.json"
        report = load_discovery_report(report_path)
        context = load_json(ROOT / "references" / "example-analysis-context.json")
        self.assertEqual(
            validate_discovery_bindings(
                context,
                [report_path],
                require_live_report=True,
            ),
            [],
        )
        seed = build_analysis_context_seed(report, report_path)
        seed_issues = validate_analysis_context(seed, delivery=True)
        self.assertIn(
            "MATERIAL_OPPORTUNITY_UNRESOLVED",
            {item.code for item in seed_issues},
        )
        tampered = copy.deepcopy(context)
        tampered["discovery_reports"][0]["sha256"] = "0" * 64
        self.assertTrue(
            any(
                "hash mismatch" in message
                for message in validate_discovery_bindings(
                    tampered,
                    [report_path],
                    require_live_report=True,
                )
            )
        )

    def test_rendered_drift_includes_hint_and_opportunity_impact(self) -> None:
        context = load_json(ROOT / "references" / "example-analysis-context.json")
        before = load_discovery_report(ROOT / "references" / "example-discovery-report.json")
        after = copy.deepcopy(before)
        after["measurement_opportunity_hints"][0]["reason"] = "Changed evidence"
        report = detect_context_drift(context, context, self.plan, before, after)
        self.assertEqual(report["status"], "review_required")
        self.assertIn("product_item_consideration", report["affected_opportunities"])
        self.assertIn("view_item", report["affected_events"])

    def test_business_change_impact_targets_contract_and_recette(self) -> None:
        context = load_json(ROOT / "references" / "example-analysis-context.json")
        request = {
            "request_version": "1.0.0",
            "change_id": "add_project_type",
            "description": "Add a solar quote project type.",
            "change_type": "value_domain",
            "selectors": {"value_domain_ids": ["project_types"]},
        }
        report = analyze_change_impact(self.plan, request, context)
        self.assertEqual(
            [item["event_name"] for item in report["affected_events"]],
            ["begin_quote"],
        )
        self.assertIn("schemas/begin_quote.schema.json", report["artifacts_to_regenerate"])
        self.assertEqual(report["unresolved_selectors"], [])

    def test_business_change_description_can_infer_payment_impact(self) -> None:
        plan = copy.deepcopy(self.plan)
        payment = copy.deepcopy(plan["events"][2])
        payment["event_name"] = "add_payment_info"
        payment["journey_ids"] = ["quote_request"]
        payment["business_question"] = "Which payment method is selected?"
        payment["measurement_opportunity_ids"] = ["checkout_payment_selection"]
        payment["parameters"][1]["name"] = "payment_type"
        payment["parameters"][1]["data_layer_path"] = "event_data.payment_type"
        plan["events"].append(payment)
        context = load_json(ROOT / "references" / "example-analysis-context.json")
        context["measurement_opportunities"].append(
            {
                "opportunity_id": "checkout_payment_selection",
                "journey_id": "quote_request",
                "name": "Payment method selection",
                "category": "progression",
                "material": True,
                "evidence_status": "confirmed",
                "evidence_refs": ["business_brief"],
                "business_question": "Which payment method is selected?",
                "official_candidate": "add_payment_info",
                "official_fit": "fit",
                "decision": "measure",
                "decision_reason": "Official checkout semantic.",
                "event_names": ["add_payment_info"],
                "discovery_hint_ids": [],
            }
        )
        report = analyze_change_impact(
            plan,
            {
                "request_version": "1.0.0",
                "change_id": "add_paypal",
                "description": "Add PayPal as a new payment method.",
                "change_type": "other",
                "selectors": {},
            },
            context,
        )
        self.assertTrue(report["inference"]["used"])
        self.assertIn("add_payment_info", {item["event_name"] for item in report["affected_events"]})
        self.assertEqual(report["unresolved_selectors"], [])

    def test_official_parameter_website_values_still_need_evidence(self) -> None:
        plan = copy.deepcopy(self.plan)
        currency = next(parameter for parameter in plan["events"][1]["parameters"] if parameter["name"] == "currency")
        currency.pop("value_evidence_refs")
        self.assertIn("FINITE_VALUE_EVIDENCE_MISSING", self.error_codes(plan))


if __name__ == "__main__":
    unittest.main()
