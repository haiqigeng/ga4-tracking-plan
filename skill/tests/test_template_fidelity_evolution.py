from __future__ import annotations

import copy
import sys
import tempfile
import unittest
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

from adapt_tracking_plan_workbook import _fill_region, _reference_value, adapt, save_adapted_workbook
from create_default_template import apply_label, apply_table_row, apply_value, build_template, set_cell_value
from generate_tracking_plan_workbook import build_workbook
from inspect_tracking_plan_template import event_tab_candidate, inspect, validate_template_mapping
from native_excel_adapter import _event_parameter_count
from template_fidelity import compare_template_fidelity, workbook_fidelity_snapshot
from template_preflight import inspect_template_richness
from tracking_plan_model import load_json, parameter_reference_rows
from validate_tracking_plan_workbook import validate_workbook


class SuppliedTemplateEvolutionTests(unittest.TestCase):
    def setUp(self) -> None:
        self.plan = load_json(ROOT / "references" / "example-tracking-plan.json")

    def _expanded_plan(self) -> dict:
        plan = copy.deepcopy(self.plan)
        event = next(item for item in plan["events"] if item["event_name"] == "begin_quote")
        parameter = copy.deepcopy(next(item for item in event["parameters"] if item["name"] == "project_type"))
        parameter.update(
            {
                "name": "project_phase",
                "data_layer_path": "event_data.project_phase",
                "definition": "Identifies the normalized maturity phase selected for the quote request.",
                "value_rule": "Use one of the exhaustive website project-maturity phases.",
                "allowed_values": ["planning", "ready"],
                "example": "planning",
                "source": "Synthetic regression fixture project phase options.",
                "custom_decision": {
                    "business_need": "Compare quote starts by project maturity.",
                    "official_candidate": "No prescribed begin_quote parameter represents project maturity.",
                    "why_not_fit": "Project maturity is business-specific.",
                },
            }
        )
        event["parameters"].append(parameter)
        event["data_layer"]["push"]["event_data"]["project_phase"] = "planning"
        return plan

    def test_v28_single_sheet_event_candidate_call_remains_supported(self) -> None:
        workbook = build_workbook(self.plan)
        candidate, review = event_tab_candidate(workbook["view_item"])
        self.assertEqual(review, [])
        self.assertEqual(candidate["sheet"], "view_item")

    def test_native_event_capacity_uses_its_event_not_the_global_maximum(self) -> None:
        workbook = build_workbook(self.plan)
        with tempfile.TemporaryDirectory() as directory:
            template = Path(directory) / "template.xlsx"
            workbook.save(template)
            mapping = inspect(template)["regions"]["event_tabs"]
        counts = {
            str(item["sheet"]): _event_parameter_count(self.plan, workbook, item)
            for item in mapping
        }
        self.assertEqual(counts["core_data"], len(self.plan["events"][0]["parameters"]))
        self.assertEqual(counts["view_item"], len(self.plan["events"][1]["parameters"]))
        self.assertNotEqual(counts["core_data"], counts["view_item"])

    def test_generated_mapping_is_formal_hash_bound_and_unambiguous(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "source.xlsx"
            build_workbook(self.plan).save(source)
            mapping = inspect(source)
        self.assertEqual(mapping["mapping_version"], "2.0.0")
        self.assertEqual(validate_template_mapping(mapping), [])
        self.assertEqual(mapping["review_required"], [])
        self.assertEqual(mapping["preflight"]["recommended_writer"], "openpyxl")
        self.assertTrue(mapping["regions"]["event_matrix"]["writable_value_cells"])
        self.assertIn("row_growth_policy", mapping["regions"]["parameter_reference"])

    def test_v28_mapping_is_reinspected_and_upgraded_without_losing_compatibility(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "source.xlsx"
            build_workbook(self.plan).save(source)
            current = inspect(source)
            legacy = {
                "mapping_version": "1.1",
                "template": current["template"],
                "regions": {},
                "policy": {"preserve_unmapped_content": True, "embed_internal_model": True},
                "review_required": [],
            }
            for name in ("event_matrix", "parameter_reference", "data_layer_table"):
                region = current["regions"][name]
                if not region:
                    legacy["regions"][name] = {}
                    continue
                legacy["regions"][name] = {
                    key: copy.deepcopy(region[key])
                    for key in ("sheet", "header_row", "data_start_row", "columns")
                }
                if name == "parameter_reference":
                    legacy["regions"][name]["semantic_role"] = "all_used_parameters"
            legacy["regions"]["event_tabs"] = [
                {
                    key: copy.deepcopy(item[key])
                    for key in ("sheet", "event_name_cell", "existing_event_name", "reusable", "field_labels", "data_layer_cell")
                }
                | {
                    "parameter_region": {
                        key: copy.deepcopy(item["parameter_region"][key])
                        for key in ("sheet", "header_row", "data_start_row", "columns")
                    }
                }
                for item in current["regions"]["event_tabs"]
            ]
            workbook = adapt(self.plan, source, legacy)
        self.assertEqual(workbook._ga4_template_fidelity_report["status"], "passed")
        self.assertEqual(workbook._ga4_effective_mapping["mapping_version"], "2.0.0")

    def test_ambiguous_regions_require_review_instead_of_first_match(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "ambiguous.xlsx"
            workbook = Workbook()
            for index, title in enumerate(("Events A", "Events B")):
                sheet = workbook.active if index == 0 else workbook.create_sheet()
                sheet.title = title
                sheet.append(["Event", "Definition"])
                sheet.append(["example", "Example definition"])
            workbook.save(source)
            mapping = inspect(source)
        self.assertTrue(any("Ambiguous event_matrix" in item for item in mapping["review_required"]))
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "ambiguous.xlsx"
            workbook.save(source)
            mapping = inspect(source)
            with self.assertRaisesRegex(ValueError, "requires review"):
                adapt(self.plan, source, mapping)

    def test_formula_bearing_mapped_value_cell_is_never_overwritten(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "formula.xlsx"
            workbook = build_workbook(self.plan)
            workbook["Event Matrix"]["B5"] = '=IF(A5="","",A5)'
            workbook.save(source)
            mapping = inspect(source)
            self.assertTrue(any("formula" in item.casefold() for item in mapping["review_required"]))
            with self.assertRaisesRegex(ValueError, "requires review"):
                adapt(self.plan, source, mapping)

    def test_plan_growth_stops_at_a_fixed_event_tab_capacity(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "fixed-capacity.xlsx"
            build_workbook(self.plan).save(source)
            mapping = inspect(source)
            begin_quote = next(item for item in mapping["regions"]["event_tabs"] if item["existing_event_name"] == "begin_quote")
            self.assertEqual(begin_quote["parameter_region"]["row_growth_policy"], "fixed_capacity")
            with self.assertRaisesRegex(ValueError, "no approved growth policy"):
                adapt(self._expanded_plan(), source, mapping)

    def test_table_growth_extends_formula_validation_conditional_format_and_name(self) -> None:
        updated = copy.deepcopy(self.plan)
        source_plan = copy.deepcopy(self.plan)
        source_event = next(item for item in source_plan["events"] if item["event_name"] == "begin_quote")
        source_event["parameters"][1] = copy.deepcopy(source_event["parameters"][0])
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "table-template.xlsx"
            output = Path(directory) / "adapted.xlsx"
            workbook = build_workbook(source_plan)
            sheet = workbook["Parameter Reference"]
            existing_rows = len(parameter_reference_rows(source_plan))
            old_end = 4 + existing_rows
            sheet["G4"] = "Audit formula"
            for row in range(5, old_end + 1):
                sheet.cell(row, 7, f"=LEN(A{row})")
                sheet.cell(row, 7)._style = copy.copy(sheet.cell(row, 6)._style)
            table = Table(displayName="ParameterRegistry", ref=f"A4:G{old_end}")
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
            sheet.add_table(table)
            validation = DataValidation(type="list", formula1='"implementation,event,item"')
            validation.add(f"B5:B{old_end}")
            sheet.add_data_validation(validation)
            sheet.conditional_formatting.add(
                f"A5:A{old_end}",
                CellIsRule(operator="notEqual", formula=['""'], fill=PatternFill("solid", fgColor="FFF2CC")),
            )
            workbook.defined_names.add(
                DefinedName("parameter_values", attr_text=f"'Parameter Reference'!$A$5:$A${old_end}")
            )
            workbook.save(source)

            mapping = inspect(source)
            self.assertEqual(mapping["preflight"]["decision"], "supported")
            self.assertEqual(mapping["regions"]["parameter_reference"]["row_growth_policy"], "excel_table")
            result = save_adapted_workbook(updated, source, mapping, output)
            reopened = load_workbook(output, data_only=False)
            new_end = old_end + 1
            self.assertEqual(reopened["Parameter Reference"].tables["ParameterRegistry"].ref, f"A4:G{new_end}")
            self.assertEqual(reopened["Parameter Reference"].cell(new_end, 7).value, f"=LEN(A{new_end})")
            self.assertIn(str(new_end), str(reopened["Parameter Reference"].data_validations.dataValidation[0].sqref))
            self.assertIn(str(new_end), str(next(iter(reopened["Parameter Reference"].conditional_formatting._cf_rules)).sqref))
            self.assertIn(str(new_end), str(reopened.defined_names["parameter_values"].attr_text))
            self.assertEqual(result["fidelity"]["status"], "passed")
            self.assertEqual(validate_workbook(output, updated, result["effective_mapping"]), [])

    def test_table_growth_preserves_totals_and_extends_only_its_data_range(self) -> None:
        updated = self._expanded_plan()
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "table-with-totals.xlsx"
            output = Path(directory) / "adapted.xlsx"
            workbook = build_workbook(self.plan)
            sheet = workbook["Parameter Reference"]
            old_data_end = 4 + len(parameter_reference_rows(self.plan))
            old_total_row = old_data_end + 1
            sheet["G4"] = "Audit formula"
            for row in range(5, old_data_end + 1):
                sheet.cell(row, 7, f"=LEN(A{row})")
                sheet.cell(row, 7)._style = copy.copy(sheet.cell(row, 6)._style)
            sheet.cell(old_total_row, 1, "Totals")
            sheet.cell(old_total_row, 7, f"=SUBTOTAL(103,A5:A{old_data_end})")
            table = Table(displayName="ParameterRegistryTotals", ref=f"A4:G{old_total_row}")
            table.totalsRowShown = True
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
            sheet.add_table(table)
            workbook.save(source)

            mapping = inspect(source)
            reopened_source = load_workbook(source, data_only=False)
            before = workbook_fidelity_snapshot(reopened_source)
            authorized = __import__("template_fidelity").authorized_template_changes(
                reopened_source,
                mapping,
                updated,
            )
            region = mapping["regions"]["parameter_reference"]
            _fill_region(
                reopened_source,
                region,
                parameter_reference_rows(updated),
                lambda item, field: _reference_value(updated, item, field),
            )
            reopened_source.save(output)
            reopened = load_workbook(output, data_only=False)
            new_data_end = old_data_end + 1
            new_total_row = new_data_end + 1
            reopened_table = reopened["Parameter Reference"].tables["ParameterRegistryTotals"]
            self.assertTrue(reopened_table.totalsRowShown)
            self.assertEqual(reopened_table.ref, f"A4:G{new_total_row}")
            self.assertEqual(
                reopened["Parameter Reference"].cell(new_total_row, 7).value,
                f"=SUBTOTAL(103,A5:A{new_data_end})",
            )
            fidelity = compare_template_fidelity(
                before,
                workbook_fidelity_snapshot(reopened),
                authorized,
            )
            self.assertEqual(fidelity["status"], "passed", fidelity["violations"])

    def test_range_growth_does_not_authorise_validation_definition_changes(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "validation.xlsx"
            workbook = build_workbook(self.plan)
            sheet = workbook["Parameter Reference"]
            end = 4 + len(parameter_reference_rows(self.plan))
            table = Table(displayName="ParameterValidation", ref=f"A4:F{end}")
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
            sheet.add_table(table)
            validation = DataValidation(type="list", formula1='"implementation,event,item"')
            validation.add(f"B5:B{end}")
            sheet.add_data_validation(validation)
            workbook.save(source)
            mapping = inspect(source)
            reopened = load_workbook(source, data_only=False)
            before = workbook_fidelity_snapshot(reopened)
            authorized = __import__("template_fidelity").authorized_template_changes(
                reopened,
                mapping,
                self._expanded_plan(),
            )
            reopened["Parameter Reference"].data_validations.dataValidation[0].formula1 = '"changed"'
            report = compare_template_fidelity(
                before,
                workbook_fidelity_snapshot(reopened),
                authorized,
            )
        self.assertEqual(report["status"], "failed")
        self.assertTrue(any(item.get("kind") == "data_validations" for item in report["violations"]))

    def test_value_authorisation_does_not_authorise_style_or_comment_changes(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "source.xlsx"
            build_workbook(self.plan).save(source)
            mapping = inspect(source)
            workbook = load_workbook(source)
            before = workbook_fidelity_snapshot(workbook)
            authorized = __import__("template_fidelity").authorized_template_changes(workbook, mapping, self.plan)
            changed_font = copy.copy(workbook["Event Matrix"]["A5"].font)
            changed_font.italic = not bool(changed_font.italic)
            workbook["Event Matrix"]["A5"].font = changed_font
            report = compare_template_fidelity(before, workbook_fidelity_snapshot(workbook), authorized)
        self.assertEqual(report["status"], "failed")
        self.assertTrue(any("style" in item.get("properties", []) for item in report["violations"]))

    def test_rich_features_require_native_writer_or_exact_refusal(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "chart.xlsx"
            workbook = build_workbook(self.plan)
            sheet = workbook["Guide"]
            for row, value in enumerate((1, 2, 3), 20):
                sheet.cell(row, 1, value)
            chart = BarChart()
            chart.add_data(Reference(sheet, min_col=1, min_row=20, max_row=22))
            sheet.add_chart(chart, "C20")
            workbook.save(source)
            preflight = inspect_template_richness(source)
        features = {item["feature"] for item in preflight["native_required_features"]}
        self.assertTrue({"chart", "drawing"} & features)
        self.assertIn(preflight["recommended_writer"], {"native_excel", "blocked"})
        if preflight["recommended_writer"] == "blocked":
            self.assertTrue(any(item["feature"] == "native_excel_unavailable" for item in preflight["blocking_features"]))

    def test_unmodelled_excel_connection_routes_to_native_or_exact_refusal(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "connection.xlsx"
            build_workbook(self.plan).save(source)
            with ZipFile(source, "a", compression=ZIP_DEFLATED) as archive:
                archive.writestr(
                    "xl/connections.xml",
                    '<?xml version="1.0" encoding="UTF-8"?><connections xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="0"/>',
                )
            preflight = inspect_template_richness(source)
        self.assertTrue(any(item["feature"] == "connection" for item in preflight["native_required_features"]))
        self.assertIn(preflight["recommended_writer"], {"native_excel", "blocked"})

    def test_chart_anchor_changes_are_not_hidden_by_equal_chart_counts(self) -> None:
        workbook = build_workbook(self.plan)
        sheet = workbook["Guide"]
        for row, value in enumerate((1, 2, 3), 20):
            sheet.cell(row, 1, value)
        chart = BarChart()
        chart.add_data(Reference(sheet, min_col=1, min_row=20, max_row=22))
        sheet.add_chart(chart, "C20")
        before = workbook_fidelity_snapshot(workbook)
        chart.anchor = "D20"
        report = compare_template_fidelity(
            before,
            workbook_fidelity_snapshot(workbook),
            __import__("template_fidelity")._property_authorisation(),
        )
        self.assertEqual(report["status"], "failed")
        self.assertTrue(any(item.get("kind") == "charts" for item in report["violations"]))

    def test_explicit_event_prototype_can_be_cloned_without_redesign(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "prototype.xlsx"
            output = Path(directory) / "adapted.xlsx"
            workbook = build_template()
            prototype = workbook["__EVENT_TEMPLATE"]
            set_cell_value(prototype["A39"], "dataLayer specification")
            apply_label(prototype["A39"])
            prototype.merge_cells("A39:G39")
            prototype.merge_cells("A40:G40")
            set_cell_value(
                prototype["A40"],
                'window.dataLayer = window.dataLayer || [];\nwindow.dataLayer.push({"event":"event_name"});',
            )
            apply_value(prototype["A40"])
            for row in range(13, 20):
                apply_table_row(prototype, row, 7)
            workbook.save(source)
            mapping = inspect(source)
            self.assertEqual(mapping["review_required"], [])
            result = save_adapted_workbook(self.plan, source, mapping, output)
            reopened = load_workbook(output, data_only=False)
            event_names = {item["event_name"] for item in self.plan["events"]}
            self.assertTrue(event_names <= set(reopened.sheetnames))
            self.assertEqual(result["fidelity"]["status"], "passed")
            self.assertEqual(validate_workbook(output, self.plan, result["effective_mapping"]), [])


if __name__ == "__main__":
    unittest.main()
