from __future__ import annotations

import copy
import json
import re
import subprocess
import sys
import tempfile
import tomllib
import zipfile
from datetime import date
from itertools import chain
from operator import attrgetter
from pathlib import Path
from xml.etree import ElementTree

from jsonschema import Draft202012Validator
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SKILL = ROOT / "skill"
RULES = SKILL / "references" / "03-rules"
COMMANDS = SKILL / "references" / "02-commands"
MAINTENANCE = ROOT / "maintenance"
SCHEMA = RULES / "schema-tracking-plan.json"
FIXTURE = RULES / "example-ga4-tracking-plan.json"

REQUIRED_FILES = {
    ROOT / "README.md",
    ROOT / "LICENSE",
    ROOT / "SECURITY.md",
    ROOT / "CONTRIBUTING.md",
    ROOT / "requirements.txt",
    ROOT / "pyproject.toml",
    SKILL / "SKILL.md",
    SKILL / "release.json",
    SKILL / "agents" / "openai.yaml",
    RULES / "execution-contract.md",
    RULES / "completion-gates.md",
    RULES / "policy-ga4-boundaries.md",
    RULES / "policy-authenticated-user-context.md",
    RULES / "policy-datalayer-contract.md",
    RULES / "policy-language-and-values.md",
    RULES / "library-ga4-event-scenarios.json",
    RULES / "library-ga4-recommended-events.json",
    RULES / "library-parameters.json",
    SCHEMA,
    FIXTURE,
    COMMANDS / "validation-commands.md",
    COMMANDS / "workbook-generation.md",
    MAINTENANCE / "references" / "official-catalog-maintenance.md",
    MAINTENANCE / "references" / "fresh-agent-evaluation.md",
    MAINTENANCE / "evaluations" / "fresh-agent-evaluation-cases.json",
    SKILL / "scripts" / "validate_tracking_plan.py",
    SKILL / "scripts" / "init_tracking_plan.py",
    SKILL / "scripts" / "generate_tracking_plan_workbook.py",
    SKILL / "scripts" / "tracking_plan_screenshots.py",
    SKILL / "scripts" / "browser_environment.py",
    SKILL / "scripts" / "inspect_browser_environment.py",
    SKILL / "scripts" / "tracking_plan_validation_common.py",
    SKILL / "scripts" / "tracking_plan_validation_datalayer.py",
    SKILL / "scripts" / "tracking_plan_validation_delivery.py",
    SKILL / "scripts" / "tracking_plan_validation_event_rules.py",
    SKILL / "scripts" / "tracking_plan_validation_governance.py",
    SKILL / "scripts" / "tracking_plan_validation_quality.py",
    SKILL / "scripts" / "tracking_plan_validation_screenshot_capture.py",
    SKILL / "scripts" / "tracking_plan_workbook_layout.py",
    SKILL / "scripts" / "official_ga4_catalog.py",
    SKILL / "scripts" / "official_source_receipt.py",
    SKILL / "scripts" / "resolve_tracking_plan.py",
    SKILL / "scripts" / "inspect_tracking_plan_template.py",
    SKILL / "scripts" / "adapt_tracking_plan_workbook.py",
    SKILL / "scripts" / "check_official_catalog.py",
    MAINTENANCE / "scripts" / "migrate_tracking_plan.py",
    ROOT / "scripts" / "inspect_tracking_plan_template.py",
    ROOT / "scripts" / "adapt_tracking_plan_workbook.py",
    ROOT / "scripts" / "inspect_browser_environment.py",
    ROOT / "scripts" / "init_tracking_plan.py",
    ROOT / "scripts" / "resolve_tracking_plan.py",
    ROOT / "scripts" / "validate_fresh_agent_evals.py",
    ROOT / "scripts" / "validate_eval_manifest.py",
}

EXPECTED_TABS = ["00 Overview", "01 GTM Protocol", "02 Parameter Reference", "03 Event Matrix", "04 DataLayer Examples", "05 Screenshot Register"]
TEXT_SUFFIXES = {".md", ".py", ".json", ".yaml", ".yml", ".toml", ".txt", ".ps1"}
LOCAL_ARTIFACT_PARTS = {
    "deliverables",
    "generated",
    "release",
    "tracking-plan-corpus-analysis",
    "fresh-agent-evaluation-output",
    "__pycache__",
}
SECRET_PATTERNS = {
    "Google API key": re.compile(r"AIza[0-9A-Za-z_-]{30,}"),
    "GitHub token": re.compile(r"gh[pousr]_[A-Za-z0-9]{30,}"),
    "private key": re.compile(r"-----BEGIN (?:RSA |EC |OPENSSH )?PRIVATE KEY-----"),
    "email address": re.compile(r"\b[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}\b", re.I),
    "GTM container ID": re.compile(r"\bGTM-[A-Z0-9]{6,}\b"),
    "GA4 measurement ID": re.compile(r"\bG-[A-Z0-9]{8,}\b"),
}
REQUIRED_RUNTIME_DEPENDENCIES = {"greenlet", "jsonschema", "openpyxl", "pillow", "playwright"}


def fail(message: str) -> None:
    raise RuntimeError(message)


def load_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8-sig"))


def run(command: list[str], label: str) -> subprocess.CompletedProcess[str]:
    result = subprocess.run(command, cwd=ROOT, text=True, capture_output=True, encoding="utf-8")
    if result.returncode:
        fail(f"{label} failed\n{result.stdout}\n{result.stderr}")
    return result


def check_required_files() -> None:
    missing = sorted(str(path.relative_to(ROOT)) for path in REQUIRED_FILES if not path.exists())
    if missing:
        fail(f"Missing required files: {', '.join(missing)}")


def check_skill_metadata() -> None:
    text = (SKILL / "SKILL.md").read_text(encoding="utf-8")
    if not text.startswith("---\n") or "name: ga4-tracking-plan" not in text or "description:" not in text:
        fail("SKILL.md frontmatter is invalid")
    if len(text.splitlines()) > 500:
        fail("SKILL.md exceeds the 500-line progressive-disclosure limit")
    agent = (SKILL / "agents" / "openai.yaml").read_text(encoding="utf-8")
    for expected in ("GA4 Web Analyst", "$ga4-tracking-plan"):
        if expected not in agent:
            fail(f"agents/openai.yaml is missing {expected}")


def dependency_name(requirement: str) -> str:
    return re.split(r"[<>=!~;\s\[]", requirement.strip(), maxsplit=1)[0].lower().replace("_", "-")


def check_release_metadata_and_dependencies() -> None:
    project = tomllib.loads((ROOT / "pyproject.toml").read_text(encoding="utf-8"))["project"]
    version = str(project.get("version", ""))
    if not re.fullmatch(r"\d+\.\d+\.\d+", version):
        fail(f"pyproject.toml contains an invalid semantic version: {version}")
    release = load_json(SKILL / "release.json")
    if release.get("name") != "ga4-tracking-plan" or release.get("version") != version:
        fail("skill/release.json must identify ga4-tracking-plan and match pyproject.toml version")
    if release.get("python_requires") != project.get("requires-python"):
        fail("skill/release.json python_requires must match pyproject.toml")
    try:
        date.fromisoformat(str(release.get("released_on", "")))
    except ValueError:
        fail("skill/release.json released_on must be an ISO date")
    fixture = load_json(FIXTURE)
    if release.get("schema_version") != fixture.get("schema_version"):
        fail("skill/release.json schema_version must match the generic fixture")
    project_dependencies = {dependency_name(value) for value in project.get("dependencies", [])}
    requirement_dependencies = {
        dependency_name(line)
        for line in (ROOT / "requirements.txt").read_text(encoding="utf-8").splitlines()
        if line.strip() and not line.lstrip().startswith("#")
    }
    if project_dependencies != requirement_dependencies:
        fail("requirements.txt and pyproject.toml runtime dependencies must match")
    missing = sorted(REQUIRED_RUNTIME_DEPENDENCIES - project_dependencies)
    if missing:
        fail(f"Runtime dependency declarations are missing: {', '.join(missing)}")
    try:
        from playwright.sync_api import sync_playwright  # noqa: F401
    except Exception as error:
        fail(f"Playwright runtime import failed: {type(error).__name__}: {error}")


def check_reference_navigation() -> None:
    text = (SKILL / "SKILL.md").read_text(encoding="utf-8")
    referenced = set(re.findall(r"`((?:references|scripts|assets)/[^`]+)`", text))
    for relative in referenced:
        path = SKILL / relative
        if "*" in relative:
            continue
        if not path.exists():
            fail(f"SKILL.md references missing resource: {relative}")


def check_ga4_only_scope() -> None:
    forbidden = "pia" + "no"
    for path in ROOT.rglob("*"):
        if not path.is_file() or ".git" in path.parts or path.suffix.lower() not in TEXT_SUFFIXES:
            continue
        text = path.read_text(encoding="utf-8-sig", errors="ignore").lower()
        if forbidden in text:
            fail(f"GA4-only package contains unsupported platform residue in {path.relative_to(ROOT)}")
    schema_text = SCHEMA.read_text(encoding="utf-8")
    for forbidden_field in ("analytics_platforms", "platform_mappings", "implementation_payloads", "qa_cases", "qa_id"):
        if f'"{forbidden_field}"' in schema_text:
            fail(f"GA4 v2 schema still contains removed field {forbidden_field}")


def check_schema_and_fixture() -> None:
    schema = load_json(SCHEMA)
    Draft202012Validator.check_schema(schema)
    fixture = load_json(FIXTURE)
    errors = sorted(Draft202012Validator(schema).iter_errors(fixture), key=lambda item: list(item.path))
    if errors:
        fail("Generic GA4 fixture does not match schema:\n" + "\n".join(error.message for error in errors))
    if fixture.get("schema_version") != "3.1.0":
        fail("Generic fixture must use schema_version 3.1.0")
    final_screenshot_statuses = schema["$defs"]["screenshotEvidence"]["properties"]["status"]["enum"]
    if final_screenshot_statuses != ["captured", "shared_evidence", "not_needed", "blocked"]:
        fail(f"Screenshot evidence enum contains non-final states: {final_screenshot_statuses}")
    event_ids = {event["event_id"] for event in fixture["events"]}
    covered = {event_id for evidence in fixture["screenshot_evidence"] for event_id in evidence["event_ids"]}
    if event_ids != covered:
        fail("Screenshot evidence must cover every fixture event exactly through explicit references")


def validator_issues(plan: dict, temp_dir: Path) -> list[dict]:
    path = temp_dir / "candidate.json"
    path.write_text(json.dumps(plan, indent=2), encoding="utf-8")
    result = subprocess.run(
        [sys.executable, "-B", "scripts/validate_tracking_plan.py", str(path), "--format", "json"],
        cwd=ROOT,
        text=True,
        capture_output=True,
        encoding="utf-8",
    )
    if result.returncode not in {0, 1}:
        fail(f"Tracking plan validator failed\n{result.stdout}\n{result.stderr}")
    return json.loads(result.stdout)


def expect_code(base: dict, temp_dir: Path, code: str, mutate) -> None:
    candidate = copy.deepcopy(base)
    mutate(candidate)
    codes = {issue["code"] for issue in validator_issues(candidate, temp_dir)}
    if code not in codes:
        fail(f"Validator did not emit expected {code}; got {sorted(codes)}")


def check_validator() -> None:
    fixture = load_json(FIXTURE)
    with tempfile.TemporaryDirectory() as raw:
        temp_dir = Path(raw)
        issues = validator_issues(fixture, temp_dir)
        if issues:
            fail(f"Generic fixture has validator issues: {issues}")

        expect_code(
            fixture,
            temp_dir,
            "LEGACY_UA_FIELD",
            lambda plan: plan["parameters"][0].update({"parameter_name": "eventCategory"}),
        )
        expect_code(
            fixture,
            temp_dir,
            "SCREENSHOT_EVENT_MISSING",
            lambda plan: plan.update({"screenshot_evidence": plan["screenshot_evidence"][1:]}),
        )
        expect_code(
            fixture,
            temp_dir,
            "SCREENSHOT_FILE_MISSING",
            lambda plan: plan["screenshot_evidence"][0].update({"status": "captured", "file_name": ""}),
        )
        expect_code(
            fixture,
            temp_dir,
            "PLAYWRIGHT_MCP_ATTEMPT_MISSING",
            lambda plan: plan["screenshot_capture"]["playwright_mcp_attempt"].update({"status": "not_required"}),
        )
        expect_code(
            fixture,
            temp_dir,
            "SCREENSHOT_CAPTURE_BLOCKED_MISMATCH",
            lambda plan: plan["screenshot_evidence"][0].update({"status": "captured", "file_name": "page.png"}),
        )
        expect_code(
            fixture,
            temp_dir,
            "EVENT_PARAMETER_OWNER_MISSING",
            lambda plan: plan["events"][0]["parameter_bindings"][0].update({"availability": "to_confirm", "data_owner": "TBD"}),
        )


def check_workbook(path: Path) -> None:
    wb = load_workbook(path, read_only=False, data_only=True)
    if wb.sheetnames != EXPECTED_TABS:
        fail(f"Unexpected workbook tabs: {wb.sheetnames}")

    overview = wb["00 Overview"]
    overview_text = " ".join(
        map(str, filter(None, map(attrgetter("value"), chain.from_iterable(overview.iter_rows()))))
    ).lower()
    forbidden = next(filter(overview_text.__contains__, ("main users", "reviewed by", "template used", "qa cases", "automation cue")), None)
    if forbidden:
        fail(f"Overview contains non-essential field: {forbidden}")

    parameter_headers = set(map(attrgetter("value"), wb["02 Parameter Reference"][3]))
    missing_headers = {
        "Variable name", "Display name", "Value rules", "Availability by event", "Data owner(s)", "Register in GA4"
    } - parameter_headers
    if missing_headers:
        fail(f"Parameter Reference is missing {', '.join(sorted(missing_headers))}")

    matrix = wb["03 Event Matrix"]
    if matrix.freeze_panes != "C6" or not str(matrix.auto_filter.ref).startswith("A5:J"):
        fail("Event Matrix must keep its parameter columns frozen and all event slots filterable")
    matrix_text = " ".join(map(str, filter(None, map(attrgetter("value"), chain.from_iterable(matrix.iter_rows())))))
    forbidden = next(filter(matrix_text.__contains__, ("event_id", "qa_id", "screenshot_id", "primary_platform")), None)
    if forbidden:
        fail(f"Event Matrix exposes internal field {forbidden}")

    datalayer_headers = list(map(attrgetter("value"), wb["04 DataLayer Examples"][3]))
    expected_datalayer_headers = ["Journey", "Event", "Evidence", "Trigger", "dataLayer.push example", "GTM and GA4 mapping", "Implementation notes"]
    if datalayer_headers[:7] != expected_datalayer_headers:
        fail(f"Unexpected DataLayer Examples headers: {datalayer_headers}")
    if wb["04 DataLayer Examples"].freeze_panes != "A4" or not str(wb["04 DataLayer Examples"].auto_filter.ref).startswith("A3:G"):
        fail("DataLayer Examples must keep headers frozen and all seven columns filterable")

    datalayer_event_names = set(
        filter(None, map(attrgetter("value"), wb["04 DataLayer Examples"]["B"][3:]))
    )
    missing_events = {event["event_name"] for event in load_json(FIXTURE)["events"]} - datalayer_event_names
    if missing_events:
        fail(f"DataLayer Examples is missing events: {', '.join(sorted(missing_events))}")

    screenshot_headers = list(map(attrgetter("value"), wb["05 Screenshot Register"][3]))
    expected_headers = ["Journey", "Event(s)", "Screenshot preview", "Page / component", "URL / route", "Capture objective", "Status", "Notes"]
    if screenshot_headers[:8] != expected_headers:
        fail(f"Unexpected Screenshot Register headers: {screenshot_headers}")
    screenshot_sheet = wb["05 Screenshot Register"]
    if (screenshot_sheet.column_dimensions["C"].width or 0) < 70:
        fail("Screenshot Register preview column is too narrow for readable 480 x 270 evidence")
    if any((screenshot_sheet.row_dimensions[row].height or 0) < 200 for row in range(4, screenshot_sheet.max_row + 1)):
        fail("Screenshot Register rows are too short for readable 480 x 270 evidence")


def check_generated_outputs() -> None:
    with tempfile.TemporaryDirectory() as raw:
        temp_dir = Path(raw)
        workbook = temp_dir / "plan.xlsx"
        csv_path = temp_dir / "plan.csv"
        run([sys.executable, "-B", "scripts/generate_tracking_plan_workbook.py", str(FIXTURE), "--output", str(workbook)], "Workbook generation")
        run([sys.executable, "-B", "scripts/export_tracking_plan_csv.py", str(FIXTURE), "--output", str(csv_path)], "CSV generation")
        check_workbook(workbook)
        csv_text = csv_path.read_text(encoding="utf-8-sig")
        for expected in ("analysis_use", "evidence_status", "availability", "data_owner"):
            if expected not in csv_text.splitlines()[0]:
                fail(f"CSV output is missing {expected}")

def check_catalog() -> None:
    run([sys.executable, "-B", "scripts/check_official_catalog.py", "--offline"], "Offline official catalog check")


def check_plan_initializer() -> None:
    with tempfile.TemporaryDirectory() as raw:
        output = Path(raw) / "plan.json"
        run(
            [
                sys.executable,
                "-B",
                "scripts/init_tracking_plan.py",
                "https://example.com/start",
                "--output",
                str(output),
                "--journey-name",
                "Example journey",
                "--screenshots",
                "not_requested",
            ],
            "Plan initializer",
        )
        issues = validator_issues(load_json(output), Path(raw))
        issue_codes = {issue["code"] for issue in issues}
        expected = {"PLAYWRIGHT_EXPLORATION_ATTEMPT_MISSING", "OFFICIAL_SOURCE_RECEIPT_INVALID"}
        if issue_codes != expected:
            fail(f"Initialized plan should remain blocked on live browser discovery and official-source review: {issues}")


def check_evaluation_manifest() -> None:
    run([sys.executable, "-B", "scripts/validate_eval_manifest.py"], "Evaluation manifest structure")


def check_migration() -> None:
    fixture = load_json(FIXTURE)
    legacy = copy.deepcopy(fixture)
    legacy["schema_version"] = "1.1.0"
    legacy["analytics_platforms"] = ["ga4"]
    legacy["qa_cases"] = []
    legacy.pop("screenshot_capture", None)
    with tempfile.TemporaryDirectory() as raw:
        temp_dir = Path(raw)
        source = temp_dir / "legacy.json"
        output = temp_dir / "migrated.json"
        source.write_text(json.dumps(legacy), encoding="utf-8")
        run([sys.executable, "-B", "scripts/migrate_tracking_plan.py", str(source), "--output", str(output)], "Contract migration")
        migrated = load_json(output)
        if migrated.get("schema_version") != "3.1.0" or "analytics_platforms" in migrated or "qa_cases" in migrated:
            fail("Contract migration did not produce a clean v3 plan")
        attempt = migrated.get("screenshot_capture", {}).get("playwright_mcp_attempt", {}).get("status")
        if attempt != "not_recorded":
            fail("Contract migration must mark an unrecorded Playwright MCP attempt explicitly")


def check_release_package() -> None:
    with tempfile.TemporaryDirectory() as raw:
        output = Path(raw) / "package.zip"
        mismatch = subprocess.run(
            [sys.executable, "-B", "scripts/create_release_package.py", "--version", "v0.0.0", "--output", str(output)],
            cwd=ROOT,
            text=True,
            capture_output=True,
            encoding="utf-8",
        )
        if mismatch.returncode == 0 or output.exists():
            fail("Release packager must reject a version that differs from pyproject.toml")
        run([sys.executable, "-B", "scripts/create_release_package.py", "--version", "validation", "--output", str(output)], "Release package")
        with zipfile.ZipFile(output) as archive:
            names = archive.namelist()
        if not any(name.endswith("skill/SKILL.md") for name in names):
            fail("Release package is missing skill/SKILL.md")
        for required in ("pyproject.toml", "skill/release.json"):
            if required not in names:
                fail(f"Release package is missing {required}")
        with zipfile.ZipFile(output) as archive:
            packaged_release = json.loads(archive.read("skill/release.json"))
            packaged_project = tomllib.loads(archive.read("pyproject.toml").decode("utf-8"))["project"]
        if packaged_release.get("version") != packaged_project.get("version"):
            fail("Release package contains inconsistent version metadata")
        required_commands = {
            "scripts/_run_skill_script.py",
            "scripts/check_installed_skill_sync.py",
            "scripts/validate_tracking_plan.py",
        }
        if not required_commands <= set(names):
            fail("Release package is missing self-contained root command wrappers")
        for name in names:
            if any(part in LOCAL_ARTIFACT_PARTS for part in Path(name).parts):
                fail(f"Release package contains banned path: {name}")


def tracked_paths() -> set[Path]:
    result = subprocess.run(
        ["git", "ls-files", "-z"],
        cwd=ROOT,
        capture_output=True,
        check=True,
    )
    return {Path(value.decode("utf-8")) for value in result.stdout.split(b"\0") if value}


def check_privacy_and_cleanliness() -> None:
    for relative in tracked_paths():
        if any(part in LOCAL_ARTIFACT_PARTS for part in relative.parts):
            fail(f"Repository tracks a local artifact path: {relative}")
    for path in ROOT.rglob("*"):
        if not path.is_file() or ".git" in path.parts:
            continue
        relative = path.relative_to(ROOT)
        if any(part in LOCAL_ARTIFACT_PARTS for part in relative.parts) or path.suffix.lower() in {".pyc", ".pyo"}:
            continue
        if path.suffix.lower() == ".xlsx":
            check_xlsx_privacy(path, relative)
            continue
        if path.suffix.lower() not in TEXT_SUFFIXES:
            continue
        raw = path.read_bytes()
        if raw.startswith(b"\xef\xbb\xbf"):
            fail(f"Text file contains UTF-8 BOM: {relative}")
        text = raw.decode("utf-8", errors="ignore")
        for label, pattern in SECRET_PATTERNS.items():
            for match in pattern.finditer(text):
                value = match.group(0)
                if value in {"GTM-XXXX", "GTM-XXXXXX", "G-XXXXXXXXXX"}:
                    continue
                fail(f"Potential {label} in {relative}: {value[:24]}")


def check_xlsx_privacy(path: Path, relative: Path) -> None:
    workbook = load_workbook(path, read_only=False, data_only=False)
    hidden = [sheet.title for sheet in workbook.worksheets if sheet.sheet_state != "visible"]
    if hidden:
        fail(f"Workbook contains hidden sheets in {relative}: {', '.join(hidden)}")
    with zipfile.ZipFile(path) as archive:
        names = archive.namelist()
        if any("externalLinks" in name for name in names):
            fail(f"Workbook contains external links: {relative}")
        for name in names:
            if not name.endswith((".xml", ".rels")):
                continue
            text = archive.read(name).decode("utf-8", errors="ignore")
            if name.endswith(".rels") and 'TargetMode="External"' in text:
                relationships = ElementTree.fromstring(text)
                blocked = [
                    item.get("Target", "")
                    for item in relationships
                    if item.get("TargetMode") == "External" and not item.get("Type", "").endswith("/hyperlink")
                ]
                if blocked:
                    fail(f"Workbook contains an external data relationship in {relative}: {blocked[0]}")
            for label, pattern in SECRET_PATTERNS.items():
                for match in pattern.finditer(text):
                    value = match.group(0)
                    if value in {"GTM-XXXX", "GTM-XXXXXX", "G-XXXXXXXXXX"}:
                        continue
                    fail(f"Potential {label} in {relative} ({name}): {value[:24]}")


CHECKS = [
    check_required_files,
    check_skill_metadata,
    check_release_metadata_and_dependencies,
    check_reference_navigation,
    check_ga4_only_scope,
    check_schema_and_fixture,
    check_validator,
    check_generated_outputs,
    check_catalog,
    check_plan_initializer,
    check_evaluation_manifest,
    check_migration,
    check_release_package,
    check_privacy_and_cleanliness,
]


def main() -> int:
    for check in CHECKS:
        check()
        print(f"OK {check.__name__}")
    print("Package validation passed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
