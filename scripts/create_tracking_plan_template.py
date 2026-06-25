from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


OUT = Path(__file__).resolve().parents[1] / "files" / "ga4_tracking_plan_template_v2_1.xlsx"

NAVY = "1F4E78"
BLUE = "D9EAF7"
LIGHT_BLUE = "EAF4FB"
GREEN = "DDEED6"
YELLOW = "FFF2CC"
GRAY = "F2F2F2"
WHITE = "FFFFFF"
RED = "F4CCCC"
DARK = "404040"

THIN = Side(style="thin", color="D9E2F3")
MEDIUM = Side(style="medium", color="9EADCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BLOCK_BORDER = Border(top=MEDIUM, bottom=THIN, left=THIN, right=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")


def set_widths(ws, widths):
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def style_cells(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = WRAP
            cell.border = BORDER
    ws.sheet_view.showGridLines = False


def title(ws, text, subtitle, max_col):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    ws.cell(1, 1, text)
    ws.cell(2, 1, subtitle)
    ws.cell(1, 1).fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(1, 1).font = Font(color=WHITE, bold=True, size=15)
    ws.cell(2, 1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws.cell(2, 1).font = Font(color=DARK)
    ws.cell(1, 1).alignment = Alignment(vertical="center")
    ws.cell(2, 1).alignment = WRAP
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 42


def section(ws, row, label, max_col):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row, 1, label)
    cell.fill = PatternFill("solid", fgColor=GREEN)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(vertical="center")
    for col in range(1, max_col + 1):
        ws.cell(row, col).border = BLOCK_BORDER


def header(ws, row, max_col, fill=NAVY):
    for col in range(1, max_col + 1):
        cell = ws.cell(row, col)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(color=WHITE if fill == NAVY else "000000", bold=True)
        cell.alignment = CENTER
        cell.border = BORDER


def add_rows(ws, rows):
    for row in rows:
        ws.append(row)


def build_overview(wb):
    ws = wb.create_sheet("00 Overview")
    title(
        ws,
        "GA4 Tracking Plan Template",
        "Human-ready tracking plan for web analysts, developers, media teams, QA, and stakeholders.",
        8,
    )
    add_rows(
        ws,
        [
            ["Document", "GA4 Tracking Plan Template", "Client / Project", "TBD", "Owner", "TBD", "Status", "Draft"],
            ["Purpose", "Define GA4 events, parameters, screenshots, and GTM/dataLayer implementation rules in one workbook.", "", "", "", "", "", ""],
            ["Main principle", "Group events from the same journey together. Each variable has one row; each event has a value/rule column and a testing-status column. When the same event can cover several similar components, keep one event slot and list possible values per variable.", "", "", "", "", "", ""],
            [],
            ["Analytics Information", "", "", "", "", "", "", ""],
            ["Account name", "Container name", "GTM container ID", "GA4 property / stream", "Associated domains", "Environment(s)", "Status", "Comments"],
            ["TBD", "TBD", "GTM-XXXXXXX", "G-XXXXXXXXXX", "example.com", "dev / staging / production", "Draft", ""],
            [],
            ["Version History", "", "", "", "", "", "", ""],
            ["Version", "Date", "Author", "Status", "Summary of changes", "Reviewed by", "Approval date", "Notes"],
            ["v2.0", "2026-06-25", "TBD", "Draft", "Production-ready template structure after comparison with original workbook", "TBD", "TBD", ""],
            [],
            ["Workbook Sheets", "", "", "", "", "", "", ""],
            ["Sheet", "Purpose", "Main users", "Notes", "", "", "", ""],
            ["00 Overview", "Document information, analytics setup, versioning, and conventions", "All", "Start here", "", "", "", ""],
            ["01 GTM Protocol", "Shared GTM/dataLayer implementation protocol", "Developers, analysts", "Readable redesign of the original protocol tab", "", "", "", ""],
            ["02 Parameter Reference", "Human-readable variable dictionary", "Analysts, developers, QA", "Human-readable columns only", "", "", "", ""],
            ["03 Event Matrix", "Page, ecommerce, and interaction events", "Analysts, developers, media, QA", "Primary working sheet", "", "", "", ""],
            ["04 Screenshot Register", "Visual evidence and interaction context", "Analysts, developers, QA, stakeholders", "Paste screenshots or link external captures", "", "", "", ""],
            [],
            ["Conventions", "", "", "", "", "", "", ""],
            ["Pattern / status", "Meaning", "Example", "Where used", "", "", "", ""],
            ["OK", "Implemented and validated", "OK", "Test status columns", "", "", "", ""],
            ["KO", "Implemented but wrong, missing, duplicated, or using unexpected value", "KO", "Test status columns", "", "", "", ""],
            ["Cannot test", "Blocked by environment, access, consent, data, or unavailable journey", "Cannot test", "Test status columns", "", "", "", ""],
            ["-", "Not applicable for this event", "-", "Value/rule columns", "", "", "", ""],
            ["value_a | value_b", "Possible values", "auto | home | unknown", "Value rules", "", "", "", ""],
            ["normalized_value", "Controlled analytics values should be lowercase ASCII snake_case with accents removed", "pret_a_porter_femme", "Value rules", "", "", "", ""],
            ["ecommerce", "Ecommerce events stay in ecommerce-only blocks and use official GA4 parameter names", "items[].item_id", "Event Matrix", "", "", "", ""],
            ["%value%", "Dynamic value", "%search_term%", "Value rules/examples", "", "", "", ""],
            ["object.key", "Nested object field", "page_data.template", "Variable names", "", "", "", ""],
            ["items[].field", "Field inside each object of an array", "items[].item_id", "Ecommerce item variables", "", "", "", ""],
        ],
    )
    for row in [7, 11, 15, 23]:
        header(ws, row, 8)
    for row in [6, 10, 14, 22]:
        section(ws, row, ws.cell(row, 1).value, 8)
    set_widths(ws, [24, 36, 28, 28, 34, 24, 18, 44])
    ws.freeze_panes = "A7"
    style_cells(ws)


def build_gtm_protocol(wb):
    ws = wb.create_sheet("01 GTM Protocol")
    title(
        ws,
        "Google Tag Manager Implementation Protocol",
        "Readable redesign of the original GTM Protocol tab. This sheet is the shared implementation contract.",
        5,
    )
    ws.append(["Section", "Topic", "Protocol / instruction", "Code / example", "Notes"])
    header(ws, 3, 5)
    rows = [
        ["1.A", "GTM base script", "Copy the GTM base script onto every page as soon as possible after the opening head tag. Replace GTM-XXXX by the project container ID.", "<!-- Google Tag Manager -->\n<script>(function(w,d,s,l,i){w[l]=w[l]||[];w[l].push({'gtm.start':\nnew Date().getTime(),event:'gtm.js'});var f=d.getElementsByTagName(s)[0],\nj=d.createElement(s),dl=l!='dataLayer'?'&l='+l:'';j.async=true;j.src=\n'https://www.googletagmanager.com/gtm.js?id='+i+dl;f.parentNode.insertBefore(j,f);\n})(window,document,'script','dataLayer','GTM-XXXX');</script>\n<!-- End Google Tag Manager -->", "Container ID is listed in Overview > Analytics Information."],
        ["1.A", "Single Page Applications", "For SPA or JavaScript apps, load the GTM base script only once, outside route-specific bundles.", "Load once in the HTML shell/root layout.", "Do not reload GTM on every route."],
        ["1.B", "dataLayer push", "Use dataLayer.push to pass values to GTM. Each push sends a JavaScript object containing the event and its context.", "dataLayer.push({\n  'event': 'event_name',\n  'key1': value1,\n  'key2': value2\n});", "Values can be string, integer, float, object, array, boolean, null, or undefined when specified."],
        ["1.B", "event key", "Most tracking pushes must include the event key. It tells GTM that something trackable occurred and names the event.", "dataLayer.push({\n  'event': 'login',\n  'event_data': {\n    'method': 'email'\n  }\n});", "Flush pushes do not include an event key."],
        ["1.B", "Flush previous values", "GTM can persist values between pushes. Flush reusable objects before sending a new object when the protocol requires it.", "dataLayer.push({'page_data': null});\ndataLayer.push({\n  'event': 'page_view',\n  'page_data': {\n    'template': 'homepage',\n    'language': 'fr'\n  }\n});", "Flush in a separate push because one JavaScript object cannot contain the same key twice."],
        ["1.B", "Do not override dataLayer", "Never assign a new value directly to dataLayer after GTM loads. This can break the command queue.", "Do not use:\ndataLayer = [{'key1': value1}]\n\nUse:\ndataLayer.push({'key1': value1});", "If you need to push before GTM loads, instantiate it safely first."],
        ["1.B", "Ensure dataLayer exists", "When a push can happen before the GTM base script loads, create the array safely before pushing.", "window.dataLayer = window.dataLayer || [];\ndataLayer.push({\n  'key1': value1\n});", "This preserves existing queued values."],
        ["1.C", "Common keys", "Use consistent object keys across the plan.", "event = trigger key\nuser_data = user context\npage_data = page context\necommerce = ecommerce context\nevent_data = interaction context", "The Event Matrix lists the exact variables for each event."],
        ["1.C", "Flush logic", "Common default: event and user_data do not need flushing; page_data, ecommerce, and event_data usually need flushing before a new value.", "dataLayer.push({'event_data': null});\ndataLayer.push({\n  'event': 'search',\n  'event_data': {'search_term': 'insurance'}\n});", "Confirm final behavior with the implementation team."],
        ["1.C", "Controlled value format", "For controlled analytics values, use lowercase ASCII snake_case, replace spaces with underscores, and remove accents/diacritics. Keep raw values only when explicitly required, such as URLs, product IDs, item names, or user-entered search terms after PII checks.", "Nouveautes -> nouveautes\nPret-a-porter femme -> pret_a_porter_femme\n60+ -> 60_plus", "This keeps GA4 dimensions easier to filter and join across reports."],
        ["1.D", "Event ordering", "When several events fire after a page or view is displayed, send page_view first, then page-specific ecommerce or interaction events.", "Home: page_view -> view_promotion\nProduct: page_view -> view_item\nCategory: page_view -> view_item_list\nCheckout start: page_view -> begin_checkout", "page_view carries page context for the view."],
        ["1.D", "Route changes", "For SPA route changes, update/flush page_data and send page_view before route-specific events.", "route change -> flush page_data -> page_view -> view_item / interaction event", "Avoid duplicate page_view on initial load."],
        ["Ecommerce", "Official GA4 ecommerce format", "In the Event Matrix, document ecommerce events with official GA4 parameter names such as currency, value, transaction_id, items, and items[].item_id. Keep GTM dataLayer wrapper paths such as ecommerce.currency in implementation notes, not as replacements for official GA4 parameters.", "gtag parameter: currency\nGTM dataLayer source: ecommerce.currency", "Do not mix event_data.* rows into ecommerce event definitions."],
        ["Ecommerce", "Required ecommerce parameters", "Include every official required or conditionally required parameter for each selected ecommerce event. If required item data is unavailable, mark the ecommerce event as not implementable for that context and use a non-ecommerce intent event when analytically useful.", "items required for most ecommerce events\none of items[].item_id or items[].item_name required\ntransaction_id required for purchase/refund\ncurrency required when value is sent", "Follow the latest Google ecommerce documentation."],
        ["QA", "Testing records", "Record the test result next to each event value/rule in the Event Matrix.", "OK / KO / Cannot test", "Use comments or notes when KO or Cannot test needs explanation."],
        ["QA", "Repeated event definitions", "When the same GA4 event and parameter structure applies to several components, test one consolidated event definition with multiple allowed values instead of creating many duplicate event columns.", "select_content with content_type: category_navigation | size_shortcut | catalogue_order_entry", "Split only when trigger logic, data availability, or business meaning is materially different."],
    ]
    add_rows(ws, rows)
    set_widths(ws, [12, 28, 62, 58, 42])
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:E{ws.max_row}"
    style_cells(ws)


def build_parameter_reference(wb):
    ws = wb.create_sheet("02 Parameter Reference")
    title(ws, "Parameter Reference", "Human-readable dictionary for variables/parameters used in the tracking plan.", 7)
    ws.append(["Variable name", "Display name", "Type", "Description", "Value rules", "Example values", "Comments"])
    header(ws, 3, 7)
    rows = [
        ["event", "Event name", "string", "GA4 event name sent in the dataLayer push.", "Use official GA4 names when possible; otherwise use clear snake_case custom names.", "page_view, search, view_item, begin_quote", ""],
        ["user_data", "User data object", "object", "Object containing user context when allowed.", "Never include direct personal data.", "{login_status:'logged'}", ""],
        ["user_data.user_id", "User ID", "string", "Stable anonymous/non-PII user identifier when allowed.", "Do not use email, phone, policy number, CRM ID if it is personal or not approved.", "u_12345_hash", "Use only with privacy approval."],
        ["user_data.login_status", "Login status", "string", "Whether the user is logged in.", "logged | not_logged | unknown", "logged", ""],
        ["user_data.type", "User type", "string", "Business user type or audience group.", "Use controlled values.", "prospect, customer", ""],
        ["user_data.segment", "User segment", "string", "Audience segment when available and allowed.", "Avoid high-cardinality or sensitive segments.", "subscriber", ""],
        ["page_data", "Page data object", "object", "Object containing page context.", "Flush before each page_view when needed.", "{template:'homepage'}", ""],
        ["page_data.location", "Page URL", "string", "Complete page URL.", "Remove sensitive query parameters.", "https://example.com/quote", ""],
        ["page_data.referrer", "Page referrer", "string", "Complete referrer URL when available.", "May be empty depending on browser/session.", "https://www.google.com/", ""],
        ["page_data.title", "Page title", "string", "Browser or CMS page title.", "Use stable page title where possible.", "Homepage", ""],
        ["page_data.name", "Page name", "string", "Business-friendly page name.", "Use controlled naming.", "Home", ""],
        ["page_data.template", "Page template", "string", "Page template/type.", "homepage | listing_page | product_page | cart | checkout | form | account", "homepage", ""],
        ["page_data.environment", "Environment", "string", "Website environment.", "production | staging | development", "production", ""],
        ["page_data.language", "Language", "string", "Page language.", "ISO language code where possible.", "fr", ""],
        ["page_data.country", "Country", "string", "Country/market context.", "ISO country code where possible.", "FR", ""],
        ["page_data.content_group1", "Content group 1", "string", "Top-level content grouping.", "Use a stable taxonomy.", "insurance", ""],
        ["page_data.content_group2", "Content group 2", "string", "Second-level content grouping.", "Use a stable taxonomy.", "auto", ""],
        ["page_data.search_term", "Search term", "string", "Search query submitted by the user.", "Scrub personal data before sending.", "insurance quote", ""],
        ["page_data.search_results_number", "Search results count", "integer", "Number of search results.", "Use 0 for no results when available.", "12", ""],
        ["event_data", "Event data object", "object", "Object containing interaction-specific data.", "Flush before each interaction event when needed.", "{method:'email'}", ""],
        ["event_data.method", "Method", "string", "Method used for login, sign-up, share, or similar actions.", "email | password | google | phone | unknown", "email", ""],
        ["event_data.content_type", "Content type", "string", "Type of selected content.", "article | product_card | faq_question | support_link | banner", "faq_question", ""],
        ["event_data.content_id", "Content ID", "string", "Stable ID of selected content.", "Prefer stable ID/slug over long title.", "faq_contact_001", ""],
        ["event_data.content_name", "Content name", "string", "Readable content name.", "Use normalized lowercase ASCII snake_case for controlled values; avoid high-cardinality GA4 registration when many values.", "how_to_contact_support", ""],
        ["event_data.search_term", "Search term", "string", "Search query for GA4 search event.", "Scrub personal data.", "insurance quote", ""],
        ["event_data.cta_location", "CTA location", "string", "Location of the call-to-action.", "header | hero | card | footer | modal", "hero", ""],
        ["event_data.cta_text", "CTA text", "string", "Visible CTA text.", "Use normalized lowercase ASCII snake_case when used as an analytics value; keep raw visible text only in screenshot notes when needed.", "get_my_quote", ""],
        ["event_data.link_url", "Link URL", "string", "Clicked link destination.", "Remove sensitive query parameters.", "https://example.com/contact", ""],
        ["event_data.form_name", "Form name", "string", "Name of the form.", "Use stable form IDs/names.", "lead_form", ""],
        ["event_data.form_step", "Form step", "string", "Current funnel/form step.", "Use stable step names or numbers.", "step_1_vehicle", ""],
        ["event_data.error_type", "Error type", "string", "Generic error category.", "Do not send raw personal/user-entered text.", "validation_error", ""],
        ["event_data.video_title", "Video title", "string", "Video title for video interactions.", "Use CMS title or stable ID.", "Product explainer", ""],
        ["ecommerce", "GTM ecommerce object", "object", "GTM/dataLayer source object containing GA4 ecommerce data.", "Flush before each ecommerce event. Event Matrix rows should use official GA4 parameter names without the ecommerce. prefix.", "{currency:'EUR',items:[...]}", "Implementation source path, not a GA4 parameter name."],
        ["currency", "Currency", "string", "Official GA4 ecommerce parameter. Currency code for monetary values.", "ISO 4217. Required when value is sent.", "EUR", "GTM source path often ecommerce.currency."],
        ["value", "Value", "number", "Official GA4 ecommerce parameter. Monetary event value.", "Use sum(price * quantity) for item events; currency is required when value is sent.", "99.99", "GTM source path often ecommerce.value."],
        ["coupon", "Coupon", "string", "Official GA4 ecommerce event or item parameter.", "Do not send personal codes if sensitive.", "WELCOME10", "Event-level and item-level coupon are independent."],
        ["transaction_id", "Transaction ID", "string", "Official GA4 ecommerce parameter for purchase/refund deduplication.", "Required for purchase/refund; must not be personal data.", "T12345", ""],
        ["affiliation", "Affiliation", "string", "Official GA4 item parameter.", "Use controlled values.", "online_store", "Item scope."],
        ["shipping", "Shipping amount", "number", "Official GA4 purchase/refund parameter.", "Use decimal number.", "4.99", ""],
        ["tax", "Tax amount", "number", "Official GA4 purchase/refund parameter.", "Use decimal number.", "19.99", ""],
        ["payment_type", "Payment type", "string", "Official GA4 checkout parameter.", "card | paypal | transfer | other", "card", ""],
        ["shipping_tier", "Shipping tier", "string", "Official GA4 shipping parameter.", "standard | express | pickup", "standard", ""],
        ["items", "Items array", "array<object>", "Official GA4 ecommerce parameter containing item objects.", "Required for most ecommerce events. Up to 200 items.", "[{item_id:'sku123'}]", "GTM source path often ecommerce.items."],
        ["items[].index", "Item index", "integer", "Official GA4 item parameter. Item position in a list.", "Use 1-based or 0-based consistently and document the choice.", "1", ""],
        ["items[].item_id", "Item ID", "string", "Official GA4 item parameter. Stable product/item ID.", "One of item_id or item_name is required when items is used.", "sku123", ""],
        ["items[].item_name", "Item name", "string", "Official GA4 item parameter. Readable product/item name.", "One of item_id or item_name is required when items is used.", "product_name", ""],
        ["items[].item_category", "Item category", "string", "Official GA4 item parameter. Primary item category.", "Use product taxonomy.", "phones", ""],
        ["items[].item_category2", "Item category 2", "string", "Official GA4 item parameter. Second category level.", "Use product taxonomy.", "smartphones", ""],
        ["items[].item_category3", "Item category 3", "string", "Official GA4 item parameter. Third category level.", "Use product taxonomy.", "android", ""],
        ["items[].item_category4", "Item category 4", "string", "Official GA4 item parameter. Fourth category level.", "Use product taxonomy.", "premium", ""],
        ["items[].item_category5", "Item category 5", "string", "Official GA4 item parameter. Fifth category level.", "Use product taxonomy.", "summer", ""],
        ["items[].item_brand", "Item brand", "string", "Official GA4 item parameter. Product brand.", "Use controlled brand value.", "apple", ""],
        ["items[].item_variant", "Item variant", "string", "Official GA4 item parameter. Variant or option.", "Use stable variant values.", "blue_128gb", ""],
        ["items[].price", "Item price", "number", "Official GA4 item parameter. Item unit price.", "Use decimal number.", "799.99", ""],
        ["items[].quantity", "Item quantity", "integer", "Official GA4 item parameter. Item quantity.", "Use integer.", "1", ""],
        ["items[].discount", "Item discount", "number", "Official GA4 item parameter. Unit monetary discount.", "Use decimal number.", "10.00", ""],
        ["items[].item_list_id", "Item list ID", "string", "Official GA4 item parameter. Stable item list ID.", "Use for list impression/click events.", "homepage_reco", ""],
        ["items[].item_list_name", "Item list name", "string", "Official GA4 item parameter. Readable list name.", "Use normalized controlled values.", "homepage_recommendations", ""],
        ["items[].location_id", "Location ID", "string", "Official GA4 item parameter. Physical location identifier when relevant.", "Use only when relevant.", "store_001", ""],
        ["promotion_id", "Promotion ID", "string", "Official GA4 ecommerce promotion parameter.", "Use same ID for impression and click.", "promo_summer_2026", "Event scope; item-level also exists for promotion events."],
        ["promotion_name", "Promotion name", "string", "Official GA4 ecommerce promotion parameter.", "Use normalized controlled values.", "summer_offer", "Event scope; item-level also exists for promotion events."],
        ["creative_name", "Creative name", "string", "Official GA4 ecommerce promotion parameter.", "Use stable campaign naming.", "banner_blue", "Event scope; item-level also exists for promotion events."],
        ["creative_slot", "Creative slot", "string", "Official GA4 ecommerce promotion parameter.", "Use section + position.", "homepage_hero_1", "Event scope; item-level also exists for promotion events."],
    ]
    add_rows(ws, rows)
    set_widths(ws, [32, 28, 18, 56, 54, 34, 44])
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:G{ws.max_row}"
    style_cells(ws)


def event_blocks():
    return [
        ["J-001 - Page and content discovery", "", "Page view", "", "Scroll", "", "Content selection", "", "", "", "", ""],
        ["event_name", "string", "page_view", "", "scroll", "", "select_content", "", "", "", "", ""],
        ["event_type", "string", "page", "", "interaction", "", "interaction", "", "", "", "", ""],
        ["trigger", "string", "Page load or SPA route change", "", "User reaches configured scroll threshold", "", "User selects a content/card/link element", "", "", "", "", ""],
        ["business_question", "string", "Which pages are viewed?", "", "How deeply do users engage with the page?", "", "Which content modules drive interest?", "", "", "", "", ""],
        ["screenshot_id", "string", "SCR-PAGE-001", "", "SCR-SCROLL-001", "", "SCR-CONTENT-001", "", "", "", "", ""],
        ["event", "string", "page_view", "", "scroll", "", "select_content", "", "", "", "", ""],
        ["page_data.location", "string", "Current URL", "", "Current URL", "", "Current URL", "", "", "", "", ""],
        ["page_data.title", "string", "Current page title", "", "Current page title", "", "Current page title", "", "", "", "", ""],
        ["page_data.template", "string", "Page template", "", "Page template", "", "Page template", "", "", "", "", ""],
        ["event_data.content_type", "string", "-", "", "-", "", "article | card | faq_question | support_link", "", "", "", "", ""],
        ["event_data.content_id", "string", "-", "", "-", "", "Stable content ID", "", "", "", "", ""],
        ["event_data.content_name", "string", "-", "", "-", "", "Readable content name", "", "", "", "", ""],
        ["J-002 - Search and lead journey", "", "Site search", "", "Quote/form start", "", "Lead generated", "", "", "", "", ""],
        ["event_name", "string", "search", "", "begin_quote", "", "generate_lead", "", "", "", "", ""],
        ["event_type", "string", "interaction", "", "interaction", "", "interaction", "", "", "", "", ""],
        ["trigger", "string", "User submits a search query", "", "User starts quote or lead form journey", "", "Lead/form successfully submitted", "", "", "", "", ""],
        ["business_question", "string", "What do users search for?", "", "Which entry points start lead journeys?", "", "Which starts become leads?", "", "", "", "", ""],
        ["screenshot_id", "string", "SCR-SEARCH-001", "", "SCR-LEAD-START-001", "", "SCR-LEAD-SUCCESS-001", "", "", "", "", ""],
        ["event", "string", "search", "", "begin_quote", "", "generate_lead", "", "", "", "", ""],
        ["event_data.search_term", "string", "%search_term% scrubbed", "", "-", "", "-", "", "", "", "", ""],
        ["event_data.cta_location", "string", "-", "", "header | hero | card | footer", "", "-", "", "", "", "", ""],
        ["event_data.cta_text", "string", "-", "", "Visible CTA text", "", "-", "", "", "", "", ""],
        ["event_data.form_name", "string", "-", "", "lead_form", "", "lead_form", "", "", "", "", ""],
        ["event_data.form_step", "string", "-", "", "start", "", "success", "", "", "", "", ""],
        ["J-003 - Ecommerce journey", "", "View item list", "", "Select item", "", "View item", "", "Add to cart", "", "Purchase", ""],
        ["event_name", "string", "view_item_list", "", "select_item", "", "view_item", "", "add_to_cart", "", "purchase", ""],
        ["event_type", "string", "ecommerce", "", "ecommerce", "", "ecommerce", "", "ecommerce", "", "ecommerce", ""],
        ["trigger", "string", "Product list displayed", "", "Product clicked from list", "", "Product detail viewed", "", "Product added to cart", "", "Order confirmation displayed", ""],
        ["business_question", "string", "Which lists are seen?", "", "Which products are selected?", "", "Which products are viewed?", "", "Which products enter cart?", "", "What revenue is generated?", ""],
        ["screenshot_id", "string", "SCR-LIST-001", "", "SCR-SELECT-ITEM-001", "", "SCR-VIEW-ITEM-001", "", "SCR-ADD-CART-001", "", "SCR-PURCHASE-001", ""],
        ["event", "string", "view_item_list", "", "select_item", "", "view_item", "", "add_to_cart", "", "purchase", ""],
        ["currency", "string", "Currency if value is sent", "", "Currency if value is sent", "", "Currency if value is sent", "", "Currency if value is sent", "", "Currency if value is sent", ""],
        ["value", "number", "-", "", "-", "", "Item/list value if available", "", "Cart addition value", "", "Transaction value", ""],
        ["transaction_id", "string", "-", "", "-", "", "-", "", "-", "", "Transaction ID required", ""],
        ["items", "array<object>", "Visible items required", "", "Clicked item required", "", "Viewed item required", "", "Added item required", "", "Purchased items required", ""],
        ["items[].item_id", "string", "Item ID; one of item_id or item_name required", "", "Item ID; one of item_id or item_name required", "", "Item ID; one of item_id or item_name required", "", "Item ID; one of item_id or item_name required", "", "Item ID; one of item_id or item_name required", ""],
        ["items[].item_name", "string", "Item name; one of item_id or item_name required", "", "Item name; one of item_id or item_name required", "", "Item name; one of item_id or item_name required", "", "Item name; one of item_id or item_name required", "", "Item name; one of item_id or item_name required", ""],
        ["items[].item_category", "string", "Item category", "", "Item category", "", "Item category", "", "Item category", "", "Item category", ""],
        ["items[].price", "number", "Item price if available", "", "Item price if available", "", "Item price", "", "Item price", "", "Item price", ""],
        ["items[].quantity", "integer", "-", "", "-", "", "-", "", "Quantity", "", "Quantity", ""],
        ["J-004 - Promotion journey", "", "Promotion impression", "", "Promotion click", "", "", "", "", "", "", ""],
        ["event_name", "string", "view_promotion", "", "select_promotion", "", "", "", "", "", "", ""],
        ["event_type", "string", "ecommerce", "", "ecommerce", "", "", "", "", "", "", ""],
        ["trigger", "string", "Promotion visible at agreed threshold", "", "User clicks promotion", "", "", "", "", "", "", ""],
        ["business_question", "string", "Which promotions are exposed?", "", "Which promotions drive clicks?", "", "", "", "", "", "", ""],
        ["screenshot_id", "string", "SCR-PROMO-VIEW-001", "", "SCR-PROMO-CLICK-001", "", "", "", "", "", "", ""],
        ["event", "string", "view_promotion", "", "select_promotion", "", "", "", "", "", "", ""],
        ["promotion_id", "string", "Stable promotion ID", "", "Same promotion ID", "", "", "", "", "", "", ""],
        ["promotion_name", "string", "Promotion name", "", "Promotion name", "", "", "", "", "", "", ""],
        ["creative_name", "string", "Creative name", "", "Creative name", "", "", "", "", "", "", ""],
        ["creative_slot", "string", "Creative slot", "", "Creative slot", "", "", "", "", "", "", ""],
        ["items", "array<object>", "Promotion item/offer array required for view_promotion", "", "Promotion item/offer array if available", "", "", "", "", "", "", ""],
        ["items[].item_id", "string", "One of item_id or item_name required", "", "One of item_id or item_name required if items is sent", "", "", "", "", "", "", ""],
        ["items[].item_name", "string", "One of item_id or item_name required", "", "One of item_id or item_name required if items is sent", "", "", "", "", "", "", ""],
    ]


def build_event_matrix(wb):
    ws = wb.create_sheet("03 Event Matrix")
    title(
        ws,
        "Event Matrix",
        "All page, ecommerce, and interaction events. Reuse one event slot for repeated same-name events when possible; list possible values per variable.",
        12,
    )
    for start_col, label in [(3, "Event slot 1"), (5, "Event slot 2"), (7, "Event slot 3"), (9, "Event slot 4"), (11, "Event slot 5")]:
        ws.merge_cells(start_row=4, start_column=start_col, end_row=4, end_column=start_col + 1)
        cell = ws.cell(4, start_col, label)
        cell.fill = PatternFill("solid", fgColor=GREEN)
        cell.font = Font(bold=True)
        cell.alignment = CENTER
    ws.append(["Variable / parameter", "Type", "Value / rule", "Test status", "Value / rule", "Test status", "Value / rule", "Test status", "Value / rule", "Test status", "Value / rule", "Test status"])
    header(ws, 5, 12)
    add_rows(ws, event_blocks())
    for r in range(6, ws.max_row + 1):
        is_block = str(ws.cell(r, 1).value or "").startswith("J-")
        for c in range(1, 13):
            cell = ws.cell(r, c)
            cell.fill = PatternFill("solid", fgColor=GREEN if is_block else WHITE)
            cell.border = BLOCK_BORDER if is_block else BORDER
            if is_block:
                cell.font = Font(bold=True)
    status_dv = DataValidation(type="list", formula1='"OK,KO,Cannot test"', allow_blank=True)
    ws.add_data_validation(status_dv)
    for col in ["D", "F", "H", "J", "L"]:
        status_dv.add(f"{col}6:{col}2000")
        ws.conditional_formatting.add(f"{col}6:{col}2000", CellIsRule(operator="equal", formula=['"OK"'], fill=PatternFill("solid", fgColor=GREEN)))
        ws.conditional_formatting.add(f"{col}6:{col}2000", CellIsRule(operator="equal", formula=['"KO"'], fill=PatternFill("solid", fgColor=RED)))
        ws.conditional_formatting.add(f"{col}6:{col}2000", CellIsRule(operator="equal", formula=['"Cannot test"'], fill=PatternFill("solid", fgColor=YELLOW)))
    set_widths(ws, [34, 18, 40, 16, 40, 16, 40, 16, 40, 16, 40, 16])
    ws.freeze_panes = "C6"
    ws.auto_filter.ref = f"A5:L{ws.max_row}"
    style_cells(ws)


def build_screenshot_register(wb):
    ws = wb.create_sheet("04 Screenshot Register")
    title(
        ws,
        "Screenshot Register",
        "Attach or link screenshots for page views and interaction events. Use screenshot IDs in the Event Matrix.",
        9,
    )
    ws.append(["Screenshot ID", "Journey", "Event name", "Capture type", "URL / route", "What the screenshot must show", "File path or link", "Visual evidence area", "Notes"])
    header(ws, 3, 9)
    rows = [
        ["SCR-PAGE-001", "Page and content discovery", "page_view", "Page view", "https://example.com/", "Full page after load", "", "Paste screenshot here", ""],
        ["SCR-SCROLL-001", "Page and content discovery", "scroll", "Page state", "https://example.com/", "Page at configured scroll threshold", "", "Paste screenshot here", ""],
        ["SCR-CONTENT-001", "Page and content discovery", "select_content", "Interaction", "https://example.com/", "Selected content/card/link visible", "", "Paste screenshot here", ""],
        ["SCR-SEARCH-001", "Search and lead journey", "search", "Interaction", "https://example.com/", "Search field and submitted query context, without PII", "", "Paste screenshot here", ""],
        ["SCR-LEAD-START-001", "Search and lead journey", "begin_quote / form start", "Interaction", "https://example.com/", "CTA/form start action visible", "", "Paste screenshot here", ""],
        ["SCR-LEAD-SUCCESS-001", "Search and lead journey", "generate_lead", "Confirmation", "TBD", "Successful lead/form confirmation state", "", "Paste screenshot here", ""],
        ["SCR-LIST-001", "Ecommerce journey", "view_item_list", "Page state", "TBD", "Product/item list visible", "", "Paste screenshot here", ""],
        ["SCR-ADD-CART-001", "Ecommerce journey", "add_to_cart", "Interaction", "TBD", "Add-to-cart action visible", "", "Paste screenshot here", ""],
        ["SCR-PURCHASE-001", "Ecommerce journey", "purchase", "Confirmation", "TBD", "Order confirmation state, no personal data visible", "", "Paste screenshot here", ""],
        ["SCR-PROMO-VIEW-001", "Promotion journey", "view_promotion", "Page state", "TBD", "Promotion visible in viewport", "", "Paste screenshot here", ""],
        ["SCR-PROMO-CLICK-001", "Promotion journey", "select_promotion", "Interaction", "TBD", "Promotion click target visible", "", "Paste screenshot here", ""],
    ]
    add_rows(ws, rows)
    for r in range(4, ws.max_row + 1):
        ws.row_dimensions[r].height = 82
        ws.cell(r, 8).fill = PatternFill("solid", fgColor=GRAY)
        ws.cell(r, 8).alignment = CENTER
    set_widths(ws, [24, 30, 24, 18, 34, 54, 42, 38, 40])
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:I{ws.max_row}"
    style_cells(ws)


def main():
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    build_overview(wb)
    build_gtm_protocol(wb)
    build_parameter_reference(wb)
    build_event_matrix(wb)
    build_screenshot_register(wb)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and len(cell.value) > 80:
                    ws.row_dimensions[cell.row].height = max(ws.row_dimensions[cell.row].height or 15, 42)
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
    wb.active = 0
    wb.save(OUT)

    check = load_workbook(OUT, read_only=True, data_only=True)
    print(OUT)
    for sheet in check.sheetnames:
        ws = check[sheet]
        print(f"{sheet}: {ws.max_row} rows x {ws.max_column} cols")


if __name__ == "__main__":
    main()
