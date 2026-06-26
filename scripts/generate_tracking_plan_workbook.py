from __future__ import annotations

import argparse
import json
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


NAVY = "1F4E78"
BLUE = "D9EAF7"
LIGHT_BLUE = "EAF4FB"
GREEN = "DDEED6"
YELLOW = "FFF2CC"
GRAY = "F2F2F2"
WHITE = "FFFFFF"
RED = "F4CCCC"
DARK = "404040"

THIN = Side(style="thin", color="D9E2F3")
MEDIUM = Side(style="medium", color="9EADCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BLOCK_BORDER = Border(top=MEDIUM, bottom=THIN, left=THIN, right=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")

EVENT_SLOT_COUNT = 5
STATUS_OPTIONS = "OK,KO,Cannot test"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate a human GA4 tracking-plan workbook from the canonical JSON contract.")
    parser.add_argument("plan", type=Path, help="Path to a JSON tracking plan using tracking_plan_schema.json.")
    parser.add_argument("--output", "-o", type=Path, required=True, help="Output XLSX path.")
    return parser.parse_args()


def load_plan(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def set_widths(ws, widths: list[int]) -> None:
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def style_cells(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = WRAP
            cell.border = BORDER
    ws.sheet_view.showGridLines = False


def title(ws, text: str, subtitle: str, max_col: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    ws.cell(1, 1, text)
    ws.cell(2, 1, subtitle)
    ws.cell(1, 1).fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(1, 1).font = Font(color=WHITE, bold=True, size=15)
    ws.cell(2, 1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws.cell(2, 1).font = Font(color=DARK)
    ws.cell(1, 1).alignment = Alignment(vertical="center")
    ws.cell(2, 1).alignment = WRAP
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 42


def section(ws, row: int, label: str, max_col: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row, 1, label)
    cell.fill = PatternFill("solid", fgColor=GREEN)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(vertical="center")
    for col in range(1, max_col + 1):
        ws.cell(row, col).border = BLOCK_BORDER


def header(ws, row: int, max_col: int, fill: str = NAVY) -> None:
    for col in range(1, max_col + 1):
        cell = ws.cell(row, col)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(color=WHITE if fill == NAVY else "000000", bold=True)
        cell.alignment = CENTER
        cell.border = BORDER


def join_values(values: list[Any] | None) -> str:
    if not values:
        return ""
    return " | ".join(str(value) for value in values)


def compact_json(value: Any) -> str:
    if value in (None, "", []):
        return "-"
    if isinstance(value, str):
        return value
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def event_family(event: dict[str, Any]) -> str:
    return "Ecommerce" if event.get("classification") == "recommended_ecommerce" else "Interactions"


def event_type_for_matrix(event: dict[str, Any]) -> str:
    if event.get("classification") == "recommended_ecommerce":
        return "ecommerce"
    if event.get("event_name") == "page_view":
        return "page"
    return "interaction"


def parameter_value(event: dict[str, Any], parameter: str) -> str:
    payload = event.get("ga4_payload", {})
    params = payload.get("parameters", {})
    items = payload.get("items", [])
    data_layer = event.get("data_layer", {})

    if parameter == "event":
        return str(data_layer.get("event_key") or event.get("event_name") or "")
    if parameter in params:
        return compact_json(params[parameter])
    if parameter == "items":
        return compact_json(items or "Required when ecommerce context is sent")
    if parameter.startswith("items[].") and items:
        key = parameter.split(".", 1)[1]
        values = [item.get(key) for item in items if isinstance(item, dict) and item.get(key) not in (None, "")]
        return join_values(values) or "Required when items is sent"

    push = data_layer.get("push", {})
    lookup_path = parameter.split(".")
    current: Any = push
    for part in lookup_path:
        if isinstance(current, dict) and part in current:
            current = current[part]
        else:
            current = None
            break
    if current is not None:
        return compact_json(current)
    return "-"


def apply_workbook_settings(wb: Workbook) -> None:
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and len(cell.value) > 80:
                    ws.row_dimensions[cell.row].height = max(ws.row_dimensions[cell.row].height or 15, 42)
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
    wb.active = 0


def build_overview(wb: Workbook, plan: dict[str, Any]) -> None:
    doc = plan["document"]
    ws = wb.create_sheet("00 Overview")
    title(ws, doc["title"], "Generated from the canonical GA4 tracking-plan JSON contract.", 8)
    rows = [
        ["Document", doc["title"], "Plan ID", plan["plan_id"], "Owner", doc["owner"], "Status", doc["status"]],
        ["Version", doc["version"], "Created date", doc["created_date"], "Template source", doc["template_source"], "Schema", plan["schema_version"]],
        ["Notes", doc.get("notes", ""), "", "", "", "", "", ""],
        [],
        ["Measurement Brief", "", "", "", "", "", "", ""],
        ["Journey", "Scope", "URL / route", "Page type", "Expected actions", "Analysis needs", "Priority", "Open questions"],
    ]
    for brief in plan["measurement_brief"]:
        rows.append([
            brief["journey_name"],
            brief["scope"],
            brief["url_or_route"],
            brief["page_type"],
            join_values(brief["expected_user_actions"]),
            join_values(brief["analysis_needs"]),
            brief["priority"],
            join_values(brief["open_questions"]),
        ])
    rows.extend([
        [],
        ["Version History", "", "", "", "", "", "", ""],
        ["Version", "Date", "Author", "Status", "Summary of changes", "Reviewed by", "Approval date", "Notes"],
        [doc["version"], doc["created_date"], doc["owner"], doc["status"], "Generated tracking plan draft", "TBD", "TBD", ""],
        [],
        ["Assumptions", "", "", "", "", "", "", ""],
        ["Assumption", "", "", "", "", "", "", ""],
    ])
    for assumption in plan["assumptions"]:
        rows.append([assumption, "", "", "", "", "", "", ""])
    rows.extend([
        [],
        ["Documentation Sources Checked", "", "", "", "", "", "", ""],
        ["Name", "URL", "Source type", "Checked for", "", "", "", ""],
    ])
    for source in plan["documentation_sources_checked"]:
        rows.append([source["name"], source["url"], source["source_type"], source["checked_for"], "", "", "", ""])
    for row in rows:
        ws.append(row)

    for row in range(1, ws.max_row + 1):
        label = ws.cell(row, 1).value
        if label in {"Measurement Brief", "Version History", "Assumptions", "Documentation Sources Checked"}:
            section(ws, row, str(label), 8)
        if label in {"Journey", "Version", "Assumption", "Name"}:
            header(ws, row, 8)
    set_widths(ws, [26, 38, 32, 22, 42, 46, 18, 42])
    ws.freeze_panes = "A7"
    style_cells(ws)


def build_gtm_protocol(wb: Workbook) -> None:
    ws = wb.create_sheet("01 GTM Protocol")
    title(ws, "Google Tag Manager Implementation Protocol", "Shared implementation contract for dataLayer, GTM, GA4, and QA.", 5)
    rows = [
        ["Section", "Topic", "Protocol / instruction", "Code / example", "Notes"],
        ["1.A", "dataLayer push", "Use dataLayer.push for event and context values. Do not overwrite the dataLayer object after GTM loads.", "window.dataLayer = window.dataLayer || [];\ndataLayer.push({ event: \"event_name\" });", "The Event Matrix lists the exact values."],
        ["1.B", "Flush reusable objects", "Flush page_data, ecommerce, or event_data before a new event when values could persist.", "dataLayer.push({ ecommerce: null });", "Use a separate push for flushing."],
        ["1.C", "Controlled values", "Use lowercase ASCII snake_case, replace spaces with underscores, and remove accents for controlled analytics values.", "pret_a_porter_femme", "Keep product IDs, ISO codes, numeric values, and safe raw terms when required."],
        ["1.D", "Ecommerce", "Use official GA4 ecommerce parameters in the plan and map GTM wrapper paths only in implementation notes.", "GA4: items[].item_id\nGTM source: ecommerce.items[].item_id", "Do not mix ecommerce events with generic interaction rows."],
        ["1.E", "Testing records", "Record test status as OK, KO, or Cannot test.", "OK / KO / Cannot test", "The QA Cases sheet provides the validation contract."],
    ]
    for row in rows:
        ws.append(row)
    header(ws, 3, 5)
    set_widths(ws, [12, 28, 62, 58, 42])
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:E{ws.max_row}"
    style_cells(ws)


def build_parameter_reference(wb: Workbook, plan: dict[str, Any]) -> None:
    ws = wb.create_sheet("02 Parameter Reference")
    title(ws, "Parameter Reference", "Human-readable dictionary for variables and parameters used in the plan.", 11)
    headers = [
        "Variable name",
        "Display name",
        "Scope",
        "Type",
        "Classification",
        "Requirement",
        "Description",
        "Value rules",
        "Example values",
        "Source",
        "Comments",
    ]
    ws.append(headers)
    header(ws, 3, len(headers))
    for param in plan["parameters"]:
        comments = []
        if param.get("register_custom_definition"):
            comments.append("Register custom definition")
        if param.get("cardinality_risk") != "low":
            comments.append(f"Cardinality risk: {param.get('cardinality_risk')}")
        if param.get("pii_risk") != "low":
            comments.append(f"PII risk: {param.get('pii_risk')}")
        ws.append([
            param["parameter_name"],
            param["display_name"],
            param["scope"],
            param["type"],
            param["classification"],
            param["required"],
            param["description"],
            param["value_rules"],
            param["example_value"],
            param["source"],
            "; ".join(comments),
        ])
    set_widths(ws, [32, 28, 18, 18, 32, 18, 50, 52, 34, 32, 42])
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:K{ws.max_row}"
    style_cells(ws)


def build_event_matrix(wb: Workbook, plan: dict[str, Any]) -> None:
    ws = wb.create_sheet("03 Event Matrix")
    title(ws, "Event Matrix", "Journey-based event matrix. Ecommerce events are kept in separate ecommerce-only blocks.", 12)
    for start_col, label in [(3, "Event slot 1"), (5, "Event slot 2"), (7, "Event slot 3"), (9, "Event slot 4"), (11, "Event slot 5")]:
        ws.merge_cells(start_row=4, start_column=start_col, end_row=4, end_column=start_col + 1)
        cell = ws.cell(4, start_col, label)
        cell.fill = PatternFill("solid", fgColor=GREEN)
        cell.font = Font(bold=True)
        cell.alignment = CENTER
    ws.append(["Variable / parameter", "Type", "Value / rule", "Test status", "Value / rule", "Test status", "Value / rule", "Test status", "Value / rule", "Test status", "Value / rule", "Test status"])
    header(ws, 5, 12)

    parameter_types = {param["parameter_name"]: param["type"] for param in plan["parameters"]}
    journey_names = {brief["journey_id"]: brief["journey_name"] for brief in plan["measurement_brief"]}
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for event in plan["events"]:
        grouped[(event["journey_id"], event_family(event))].append(event)

    block_index = 1
    for (journey_id, family), events in grouped.items():
        for chunk_index in range(0, len(events), EVENT_SLOT_COUNT):
            chunk = events[chunk_index:chunk_index + EVENT_SLOT_COUNT]
            block_title = f"J-{block_index:03d} - {journey_names.get(journey_id, journey_id)} - {family}"
            row = [block_title, ""]
            for event in chunk:
                row.extend([event["event_name"], ""])
            row.extend([""] * (12 - len(row)))
            ws.append(row)

            standard_rows = [
                ("event_id", "string", lambda event: event["event_id"]),
                ("qa_id", "string", lambda event: event["qa"]["qa_id"]),
                ("event_name", "string", lambda event: event["event_name"]),
                ("event_type", "string", event_type_for_matrix),
                ("classification", "string", lambda event: event["classification"]),
                ("official_ga4_match", "string", lambda event: event["official_ga4_match"]),
                ("trigger", "string", lambda event: event["trigger"]),
                ("business_question", "string", lambda event: event["business_question"]),
                ("key_event", "boolean", lambda event: str(event["key_event"]).lower()),
                ("event", "string", lambda event: event["data_layer"].get("event_key", event["event_name"])),
            ]
            for variable, value_type, resolver in standard_rows:
                matrix_row = [variable, value_type]
                for event in chunk:
                    matrix_row.extend([resolver(event), "Cannot test"])
                matrix_row.extend([""] * (12 - len(matrix_row)))
                ws.append(matrix_row)

            parameters = []
            for event in chunk:
                for parameter in event["parameters"]:
                    if parameter not in parameters:
                        parameters.append(parameter)
            for parameter in parameters:
                matrix_row = [parameter, parameter_types.get(parameter, "string")]
                for event in chunk:
                    value = parameter_value(event, parameter) if parameter in event["parameters"] else "-"
                    matrix_row.extend([value, "Cannot test"])
                matrix_row.extend([""] * (12 - len(matrix_row)))
                ws.append(matrix_row)
            block_index += 1

    for row in range(6, ws.max_row + 1):
        is_block = str(ws.cell(row, 1).value or "").startswith("J-")
        for col in range(1, 13):
            cell = ws.cell(row, col)
            cell.fill = PatternFill("solid", fgColor=GREEN if is_block else WHITE)
            cell.border = BLOCK_BORDER if is_block else BORDER
            if is_block:
                cell.font = Font(bold=True)
    status_dv = DataValidation(type="list", formula1=f'"{STATUS_OPTIONS}"', allow_blank=True)
    ws.add_data_validation(status_dv)
    for col in ["D", "F", "H", "J", "L"]:
        status_dv.add(f"{col}6:{col}2000")
        ws.conditional_formatting.add(f"{col}6:{col}2000", CellIsRule(operator="equal", formula=['"OK"'], fill=PatternFill("solid", fgColor=GREEN)))
        ws.conditional_formatting.add(f"{col}6:{col}2000", CellIsRule(operator="equal", formula=['"KO"'], fill=PatternFill("solid", fgColor=RED)))
        ws.conditional_formatting.add(f"{col}6:{col}2000", CellIsRule(operator="equal", formula=['"Cannot test"'], fill=PatternFill("solid", fgColor=YELLOW)))
    set_widths(ws, [34, 18, 40, 16, 40, 16, 40, 16, 40, 16, 40, 16])
    ws.freeze_panes = "C6"
    ws.auto_filter.ref = f"A5:L{ws.max_row}"
    style_cells(ws)


def build_screenshot_register(wb: Workbook, plan: dict[str, Any]) -> None:
    ws = wb.create_sheet("04 Screenshot Register")
    title(ws, "Screenshot Register", "Attach or link screenshots for page views and interaction events.", 9)
    ws.append(["Screenshot ID", "Journey", "Event name", "Capture type", "URL / route", "What the screenshot must show", "File path or link", "Visual evidence area", "Notes"])
    header(ws, 3, 9)
    journey_names = {brief["journey_id"]: brief["journey_name"] for brief in plan["measurement_brief"]}
    for event in plan["events"]:
        capture_type = "Page view" if event["event_name"] == "page_view" else "Interaction"
        if event["classification"] == "recommended_ecommerce":
            capture_type = "Ecommerce interaction"
        ws.append([
            f"SCR-{event['event_id']}",
            journey_names.get(event["journey_id"], event["journey_id"]),
            event["event_name"],
            capture_type,
            event["page_url_pattern"],
            event["trigger"],
            "",
            "Paste screenshot here",
            f"QA case: {event['qa']['qa_id']}",
        ])
    for row in range(4, ws.max_row + 1):
        ws.row_dimensions[row].height = 82
        ws.cell(row, 8).fill = PatternFill("solid", fgColor=GRAY)
        ws.cell(row, 8).alignment = CENTER
    set_widths(ws, [24, 30, 24, 22, 34, 54, 42, 38, 40])
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:I{ws.max_row}"
    style_cells(ws)


def build_qa_cases(wb: Workbook, plan: dict[str, Any]) -> None:
    ws = wb.create_sheet("05 QA Cases")
    title(ws, "QA Cases", "Validation contract for DebugView, GTM Preview, network checks, and release sign-off.", 11)
    ws.append([
        "QA ID",
        "Journey",
        "Event name",
        "Event ID",
        "Methods",
        "Steps",
        "Expected dataLayer",
        "Expected network / GA4 payload",
        "DebugView expectation",
        "Status",
        "Evidence / notes",
    ])
    header(ws, 3, 11)
    events_by_id = {event["event_id"]: event for event in plan["events"]}
    journey_names = {brief["journey_id"]: brief["journey_name"] for brief in plan["measurement_brief"]}
    for case in plan["qa_cases"]:
        event = events_by_id.get(case["event_id"], {})
        ws.append([
            case["qa_id"],
            journey_names.get(event.get("journey_id"), event.get("journey_id", "")),
            case["event_name"],
            case["event_id"],
            join_values(case["methods"]),
            "\n".join(case["steps"]),
            "\n".join(case["expected_data_layer"]),
            "\n".join(case["expected_network"]),
            case["debugview_expectation"],
            "Cannot test" if case["status"] == "not_started" else case["status"],
            case["evidence"],
        ])
    status_dv = DataValidation(type="list", formula1=f'"{STATUS_OPTIONS}"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add(f"J4:J{ws.max_row + 200}")
    ws.conditional_formatting.add(f"J4:J{ws.max_row + 200}", CellIsRule(operator="equal", formula=['"OK"'], fill=PatternFill("solid", fgColor=GREEN)))
    ws.conditional_formatting.add(f"J4:J{ws.max_row + 200}", CellIsRule(operator="equal", formula=['"KO"'], fill=PatternFill("solid", fgColor=RED)))
    ws.conditional_formatting.add(f"J4:J{ws.max_row + 200}", CellIsRule(operator="equal", formula=['"Cannot test"'], fill=PatternFill("solid", fgColor=YELLOW)))
    set_widths(ws, [20, 28, 28, 22, 28, 52, 52, 56, 44, 16, 44])
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:K{ws.max_row}"
    style_cells(ws)


def build_workbook(plan: dict[str, Any]) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    build_overview(wb, plan)
    build_gtm_protocol(wb)
    build_parameter_reference(wb, plan)
    build_event_matrix(wb, plan)
    build_screenshot_register(wb, plan)
    build_qa_cases(wb, plan)
    apply_workbook_settings(wb)
    return wb


def main() -> int:
    args = parse_args()
    plan = load_plan(args.plan)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook = build_workbook(plan)
    workbook.save(args.output)
    print(args.output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
